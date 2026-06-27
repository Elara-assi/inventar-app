from __future__ import annotations

import hashlib
import html
import json
import os
import re
import secrets
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse
from uuid import UUID

import httpx
from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage, ImageOps
from pydantic import BaseModel

from .db import db_pool_status, execute, fetch_all, fetch_one
from .logic import (
    WORKSHOP_REFERENCE_CATALOG,
    audit,
    build_ai_suggestion,
    create_rework_tasks,
    finalization_blockers,
    load_inventory_history_references,
    load_special_tool_references,
    next_inventory_id,
    normalize_bga_object_class,
    plausible_value_limit,
    tokenize_reference_text,
)
from .migrations import run_migrations
from .security import (
    bearer_token_from_request,
    create_access_token,
    create_mobile_session_token,
    current_user_from_request,
    decode_access_token,
    hash_password,
    request_id,
    verify_password,
)
from .settings import settings

AI_ACTIVE_STATUSES = {
    "ki_wartet",
    "ki_laeuft",
    "ki_schnell_wartet",
    "ki_schnell_laeuft",
    "ki_pruefung_wartet",
    "ki_pruefung_laeuft",
}
AI_ACTIVE_STATUS_SQL = ", ".join(f"'{status}'" for status in sorted(AI_ACTIVE_STATUSES))
AI_STALE_INTERVAL_SQL = "8 minutes"

app = FastAPI(title="Inventar API", version="0.1.0-phase1")
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origin_list(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def request_context_middleware(request: Request, call_next):
    rid = request.headers.get("x-request-id") or request_id()
    request.state.request_id = rid
    auth_error = authorize_request(request)
    if auth_error:
        return auth_error
    response = await call_next(request)
    response.headers["x-request-id"] = rid
    return response


PUBLIC_PATH_PREFIXES = (
    "/health",
    "/auth/login",
    "/sessions/join",
)

MOBILE_ALLOWED_PREFIXES = (
    "/meta/bootstrap",
    "/exports",
    "/items",
    "/damage-reports",
    "/offline-sync/photos",
    "/offline-sync/items",
    "/offline-sync/status",
    "/offline-sync/reconcile",
    "/mobile-diagnostics",
    "/uploads/photos",
    "/uploads/damage-photos",
)


def json_error(status_code: int, detail: str, request: Request | None = None):
    from fastapi.responses import JSONResponse

    response = JSONResponse({"detail": detail}, status_code=status_code)
    if request is not None:
        origin = request.headers.get("origin")
        allowed_origins = settings.cors_origin_list()
        if origin and ("*" in allowed_origins or origin in allowed_origins):
            response.headers["access-control-allow-origin"] = origin
            response.headers["access-control-allow-credentials"] = "true"
            response.headers["vary"] = "Origin"
        response.headers["x-request-id"] = getattr(request.state, "request_id", request_id())
    return response


def authorize_request(request: Request):
    if request.method == "OPTIONS":
        return None
    path = request.url.path
    if any(path.startswith(prefix) for prefix in PUBLIC_PATH_PREFIXES):
        return None
    token = bearer_token_from_request(request)
    if not token:
        return json_error(401, "Anmeldung erforderlich", request)
    try:
        payload = decode_access_token(token)
    except HTTPException as exc:
        return json_error(exc.status_code, str(exc.detail), request)
    request.state.auth = payload
    if payload.get("kind") == "mobile_session":
        device_id = str(payload.get("device_id") or "")
        session_id = str(payload.get("session_id") or "")
        device = fetch_one(
            "SELECT id, revoked_at FROM session_devices WHERE id = %s AND session_id = %s",
            (device_id, session_id),
        ) if device_id and session_id else None
        if not device or device.get("revoked_at"):
            return json_error(403, "Mobile Session wurde widerrufen. Bitte QR-Code neu koppeln.", request)
        if any(path.startswith(prefix) for prefix in MOBILE_ALLOWED_PREFIXES):
            return None
        return json_error(403, "Mobile Session darf diese Funktion nicht ausführen", request)
    return None

INVENTORY_TYPE_LABELS = {
    "bga": "Betriebs- und Geschäftsausstattung",
    "tires_wheels": "Reifen und Räder",
    "special_tools": "Spezialwerkzeuge",
}


def inventory_type_label(value: Any) -> str:
    return INVENTORY_TYPE_LABELS.get(str(value or "bga"), INVENTORY_TYPE_LABELS["bga"])


class LoginIn(BaseModel):
    email: str
    password: str = "!Scherer!"


class SessionIn(BaseModel):
    location_id: str | None = None
    location_name: str | None = None
    building_id: str | None = None
    building_name: str | None = None
    room_id: str | None = None
    room_name: str | None = None
    started_by: str | None = None
    inventory_type: str = "bga"


class JoinIn(BaseModel):
    token: str
    device_name: str = "Mobile device"
    device_fingerprint: str | None = None


class LocationIn(BaseModel):
    name: str
    code: str | None = None
    address: str | None = None


class UserIn(BaseModel):
    display_name: str
    email: str | None = None
    role_slug: str = "erfasser"


class RoomIn(BaseModel):
    building_id: str | None = None
    location_id: str | None = None
    location_name: str | None = None
    building_name: str | None = None
    name: str
    code: str | None = None
    room_type: str = "workspace"


class RoomPatch(BaseModel):
    building_id: str | None = None
    name: str | None = None
    code: str | None = None
    room_type: str | None = None


class DamageArticleSnapshot(BaseModel):
    article_no: str | None = None
    nr: str | None = None
    buchungskreis: str | None = None
    anlagenbezeichnung: str | None = None
    aktivdatum: str | int | float | None = None
    aktivdatum_iso: str | None = None
    alter: float | None = None


class DamageReportPayload(BaseModel):
    client_report_id: str
    session_id: str | None = None
    source_device_id: str | None = None
    article_no: str
    article: DamageArticleSnapshot
    entry_type: str = "catalog"
    free_reference: str | None = None
    team_name: str = "Team 1"
    description: str = ""
    uvv_sticker_present: str = "unklar"
    created_at: str | None = None
    updated_at: str | None = None


class DamagePhotoPayload(BaseModel):
    client_photo_id: str | None = None
    photo_type: str
    filename: str | None = None
    mime_type: str | None = None
    size: int | None = None


class DamageSyncPayload(BaseModel):
    report: DamageReportPayload
    photos: list[DamagePhotoPayload] = []


class ItemIn(BaseModel):
    session_id: str
    inventory_type: str = "bga"
    client_item_id: str | None = None
    source_device_id: str | None = None
    object_type: str | None = None
    object_class_id: str | None = None
    inventory_id: str | None = None
    temporary_id: str | None = None
    condition: str = "gebraucht"
    condition_note: str | None = None
    brand: str | None = None
    model: str | None = None
    serial_number: str | None = None
    created_by: str | None = None
    specification: str | None = None
    construction_year: str | None = None
    function_ok: str = "nicht_geprueft"
    uvv_status: str = "unklar"
    uvv_valid_until: date | None = None
    inspection_book_available: str = "unklar"
    remark: str | None = None
    type_plate_status: str = "nicht_geprueft"


class OfflineReconcilePackageIn(BaseModel):
    client_item_id: str
    client_photo_ids: list[str] = []


class OfflineReconcileIn(BaseModel):
    session_id: str
    source_device_id: str
    packages: list[OfflineReconcilePackageIn] = []


class ItemPatch(BaseModel):
    object_type: str | None = None
    object_class_id: str | None = None
    brand: str | None = None
    model: str | None = None
    serial_number: str | None = None
    condition: str | None = None
    condition_note: str | None = None
    value_estimate: float | None = None
    estimated_age_years: float | None = None
    age_source: str | None = None
    age_verification_status: str | None = None
    status: str | None = None
    review_status: str | None = None
    commercial_category: str | None = None
    accounting_status: str | None = None
    inventory_type: str | None = None
    specification: str | None = None
    construction_year: str | None = None
    function_ok: str | None = None
    uvv_status: str | None = None
    uvv_valid_until: date | None = None
    inspection_book_available: str | None = None
    remark: str | None = None
    type_plate_status: str | None = None


class TireWheelDataBase(BaseModel):
    set_type: str | None = None
    season: str | None = None
    manufacturer: str | None = None
    profile_model: str | None = None
    tire_size: str | None = None
    load_index: str | None = None
    speed_index: str | None = None
    dot: str | None = None
    production_week: int | None = None
    production_year: int | None = None
    tread_depth_front_left: float | None = None
    tread_depth_front_right: float | None = None
    tread_depth_rear_left: float | None = None
    tread_depth_rear_right: float | None = None
    tread_depth_single: float | None = None
    rim_present: bool | None = None
    rim_type: str | None = None
    rim_condition: str | None = None
    tire_condition: str | None = None
    damage_note: str | None = None
    set_complete: bool | None = None
    storage_location: str | None = None
    remark: str | None = None


class TireWheelDataCreate(TireWheelDataBase):
    set_type: str = "satz"
    season: str = "unklar"


class TireWheelDataUpdate(TireWheelDataBase):
    pass


class TireWheelDataResponse(TireWheelDataBase):
    id: str | None = None
    item_id: str
    created_at: datetime | None = None
    updated_at: datetime | None = None


class LearningExampleIn(BaseModel):
    notes: str | None = None


class ReopenIn(BaseModel):
    reason: str | None = None


def ensure_upload_dirs() -> None:
    for sub in ["originals", "stamped", "audio", "exports", "temp", "damage"]:
        Path(settings.upload_root, sub).mkdir(parents=True, exist_ok=True)


def demo_user_id() -> str | None:
    row = fetch_one("SELECT id FROM users WHERE email = %s", (settings.demo_user_email,))
    return str(row["id"]) if row else None


def default_tenant_id() -> str | None:
    row = fetch_one("SELECT id FROM tenants WHERE slug = %s", (settings.default_tenant_slug,))
    return str(row["id"]) if row else None


def safe_upload_path(value: str) -> Path:
    root = Path(settings.upload_root).resolve()
    path = Path(value).resolve()
    if root not in path.parents and path != root:
        raise HTTPException(status_code=403, detail="Uploadpfad nicht erlaubt")
    return path


def detect_image_mime(data: bytes) -> str | None:
    if data.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if data.startswith(b"RIFF") and b"WEBP" in data[:16]:
        return "image/webp"
    if len(data) > 12 and data[4:8] == b"ftyp":
        brand = data[8:12].lower()
        if brand in {b"heic", b"heix", b"hevc", b"hevx", b"mif1", b"msf1"}:
            return "image/heic"
    return None


def safe_photo_suffix(filename: str | None, mime_type: str | None) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}:
        return suffix
    return {
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/heic": ".heic",
        "image/heif": ".heif",
    }.get(mime_type or "", ".jpg")


def safe_audio_suffix(filename: str | None, content_type: str | None) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".webm", ".m4a", ".mp3", ".mp4", ".wav", ".ogg", ".aac", ".txt"}:
        return suffix
    declared = (content_type or "").split(";")[0].strip().lower()
    return {
        "audio/webm": ".webm",
        "audio/mp4": ".m4a",
        "audio/mpeg": ".mp3",
        "audio/wav": ".wav",
        "audio/x-wav": ".wav",
        "audio/ogg": ".ogg",
        "audio/aac": ".aac",
        "video/mp4": ".m4a",
        "text/plain": ".txt",
    }.get(declared, ".bin")


def validate_photo_upload(file: UploadFile, data: bytes) -> tuple[str, str]:
    if not data:
        raise HTTPException(status_code=400, detail="Foto-Datei ist leer")
    if len(data) > settings.max_upload_bytes:
        raise HTTPException(status_code=413, detail="Foto ist zu groß")
    declared = (file.content_type or "").split(";")[0].strip().lower()
    detected = detect_image_mime(data)
    allowed = settings.allowed_upload_mime_list()
    effective = detected or declared
    if effective not in allowed:
        raise HTTPException(status_code=415, detail="Nur JPEG, PNG, WebP oder HEIC Fotos sind erlaubt")
    if declared and declared not in allowed:
        raise HTTPException(status_code=415, detail="Foto-Dateityp ist nicht erlaubt")
    return effective, safe_photo_suffix(file.filename, effective)


def require_item_session(item_id: str) -> dict[str, Any]:
    row = fetch_one(
        """
        SELECT i.id, i.session_id, s.status AS session_status
        FROM inventory_items i
        JOIN inventory_sessions s ON s.id = i.session_id
        WHERE i.id = %s
        """,
        (item_id,),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    return row


def require_item_session_open(item_id: str) -> dict[str, Any]:
    row = require_item_session(item_id)
    if row["session_status"] != "open":
        raise HTTPException(status_code=409, detail="Raum ist abgeschlossen")
    return row


def next_sequence_number(session_id: str) -> int:
    row = fetch_one(
        "SELECT COALESCE(max(sequence_number), 0)::int + 1 AS next_number FROM inventory_items WHERE session_id = %s",
        (session_id,),
    )
    return int((row or {"next_number": 1})["next_number"])


def photo_type_aliases(kind: str) -> tuple[str, ...]:
    if kind == "object":
        return ("object", "object_front", "tire_overview")
    if kind == "nameplate":
        return ("nameplate", "type_plate")
    if kind == "condition":
        return ("condition", "condition_detail", "object_back")
    if kind == "tire_overview":
        return ("tire_overview",)
    return (kind,)


def has_photo_type(item_id: str, kind: str) -> bool:
    aliases = photo_type_aliases(kind)
    placeholders = ", ".join(["%s"] * len(aliases))
    row = fetch_one(
        f"SELECT 1 AS ok FROM item_photos WHERE item_id = %s AND photo_type IN ({placeholders}) LIMIT 1",
        (item_id, *aliases),
    )
    return bool(row)


def has_ai_object_photo(item_id: str) -> bool:
    return has_photo_type(item_id, "object")


def upsert_rework_task(item_id: str, role: str, field: str, priority: str = "normal", comment: str | None = None) -> None:
    upsert_rework_task_for_type(item_id, "bga_check", role, field, priority, comment)


def upsert_rework_task_for_type(
    item_id: str,
    task_type: str,
    role: str,
    field: str,
    priority: str = "normal",
    comment: str | None = None,
) -> None:
    execute(
        """
        INSERT INTO accounting_tasks (item_id, task_type, assigned_role, missing_field, priority, comment)
        SELECT %s, %s, %s, %s, %s, %s
        WHERE NOT EXISTS (
          SELECT 1 FROM accounting_tasks
          WHERE item_id = %s AND task_type = %s AND missing_field = %s AND status = 'open'
        )
        RETURNING id
        """,
        (item_id, task_type, role, field, priority, comment or field, item_id, task_type, field),
    )


def complete_satisfied_bga_rework_tasks(item_id: str, item: dict[str, Any]) -> None:
    function_resolved = item.get("function_ok") == "ja"
    uvv_resolved = item.get("uvv_status") in {"vorhanden", "nicht_uvv_pflichtig"}
    inspection_book_resolved = item.get("inspection_book_available") in {"ja", "nein", "nicht_erforderlich"}

    if function_resolved and uvv_resolved:
        execute(
            """
            UPDATE accounting_tasks
            SET status = 'completed', completed_at = now()
            WHERE item_id = %s
              AND status = 'open'
              AND assigned_role IN ('Erfasser', 'Technik')
              AND lower(COALESCE(missing_field, '')) LIKE '%%funktion%%'
              AND lower(COALESCE(missing_field, '')) LIKE '%%uvv%%'
            RETURNING id
            """,
            (item_id,),
        )
    if function_resolved:
        execute(
            """
            UPDATE accounting_tasks
            SET status = 'completed', completed_at = now()
            WHERE item_id = %s
              AND status = 'open'
              AND assigned_role IN ('Erfasser', 'Technik')
              AND lower(COALESCE(missing_field, '')) LIKE '%%funktion%%'
              AND lower(COALESCE(missing_field, '')) NOT LIKE '%%uvv%%'
            RETURNING id
            """,
            (item_id,),
        )
    if uvv_resolved:
        execute(
            """
            UPDATE accounting_tasks
            SET status = 'completed', completed_at = now()
            WHERE item_id = %s
              AND status = 'open'
              AND assigned_role IN ('Erfasser', 'Technik')
              AND lower(COALESCE(missing_field, '')) LIKE '%%uvv%%'
              AND lower(COALESCE(missing_field, '')) NOT LIKE '%%funktion%%'
            RETURNING id
            """,
            (item_id,),
        )
    if inspection_book_resolved:
        execute(
            """
            UPDATE accounting_tasks
            SET status = 'completed', completed_at = now()
            WHERE item_id = %s
              AND status = 'open'
              AND assigned_role IN ('Erfasser', 'Technik')
              AND lower(COALESCE(missing_field, '')) LIKE '%%prüfbuch%%'
            RETURNING id
            """,
            (item_id,),
        )


def parse_dot_week_year(value: str | None) -> tuple[int | None, int | None, bool]:
    raw = re.sub(r"\D", "", value or "")
    if len(raw) < 4:
        return None, None, False
    token = raw[-4:]
    week = int(token[:2])
    year_suffix = int(token[2:])
    current_year = date.today().year
    current_suffix = current_year % 100
    year = 2000 + year_suffix if year_suffix <= current_suffix + 1 else 1900 + year_suffix
    if week < 1 or week > 53:
        return None, None, False
    if year < 1980 or year > current_year + 1:
        return None, None, False
    return week, year, True


def normalize_tire_wheel_payload(data: dict[str, Any]) -> dict[str, Any]:
    if "dot" in data:
        week, year, valid = parse_dot_week_year(data.get("dot"))
        data["production_week"] = week if valid else None
        data["production_year"] = year if valid else None
    return data


def tire_tread_values(data: dict[str, Any]) -> list[float]:
    fields = [
        "tread_depth_front_left",
        "tread_depth_front_right",
        "tread_depth_rear_left",
        "tread_depth_rear_right",
        "tread_depth_single",
    ]
    values = []
    for field in fields:
        value = data.get(field)
        if value is None:
            continue
        try:
            values.append(float(value))
        except (TypeError, ValueError):
            continue
    return values


def tire_tread_missing(data: dict[str, Any]) -> bool:
    if data.get("set_type") == "einzelreifen":
        return data.get("tread_depth_single") is None
    if data.get("set_type") == "satz":
        return any(data.get(field) is None for field in [
            "tread_depth_front_left",
            "tread_depth_front_right",
            "tread_depth_rear_left",
            "tread_depth_rear_right",
        ])
    return True


def tire_warning_limit(season: str | None) -> float:
    return 3.0 if season == "sommer" else 4.0


def run_bga_rework_check(item_id: str) -> None:
    item = fetch_one("SELECT * FROM inventory_items WHERE id = %s", (item_id,))
    if not item or item.get("inventory_type") != "bga":
        return
    execute(
        """
        UPDATE accounting_tasks
        SET status = 'completed', completed_at = now()
        WHERE item_id = %s AND task_type = 'bga_check' AND status = 'open'
        RETURNING id
        """,
        (item_id,),
    )

    if not has_photo_type(item_id, "object"):
        upsert_rework_task(item_id, "Erfasser", "Objektfoto fehlt", "hoch", "Pflichtfoto aus der Zählliste fehlt.")
    if not str(item.get("object_type") or "").strip():
        upsert_rework_task(item_id, "Erfasser", "Bezeichnung fehlt", "hoch", "Bezeichnung aus der Zählliste fehlt.")
    if item.get("condition") in {None, "", "unklar"}:
        upsert_rework_task(item_id, "Erfasser", "Zustand unklar", "normal", "Zustand muss für die Aufnahme eingeordnet werden.")
    if item.get("function_ok") == "nein":
        upsert_rework_task(item_id, "Technik", "Funktion nicht in Ordnung", "hoch", "Funktionsstatus muss technisch geklärt werden.")

    complete_satisfied_bga_rework_tasks(item_id, item)

    open_critical = fetch_one(
        """
        SELECT
          count(*)::int AS count,
          count(*) FILTER (WHERE assigned_role = 'Erfasser')::int AS erfasser_count,
          count(*) FILTER (WHERE assigned_role = 'Technik')::int AS technik_count
        FROM accounting_tasks
        WHERE item_id = %s AND status = 'open' AND assigned_role IN ('Erfasser', 'Technik')
        """,
        (item_id,),
    )
    if open_critical and int(open_critical["count"]) > 0:
        next_review_status = "nacharbeit_erfasser" if int(open_critical.get("erfasser_count") or 0) > 0 else "nacharbeit_technik"
        execute(
            """
            UPDATE inventory_items
            SET review_status = CASE
                  WHEN review_status = 'finalisiert' THEN review_status
                  ELSE %s
                END,
                status = CASE WHEN status = 'finalisiert' THEN status ELSE 'nacharbeit_noetig' END,
                updated_at = now()
            WHERE id = %s
            RETURNING id
            """,
            (next_review_status, item_id),
        )
    else:
        execute(
            """
            UPDATE inventory_items
            SET review_status = CASE WHEN review_status IN ('erfasst', 'nacharbeit_erfasser', 'nacharbeit_technik') THEN 'finalisierbar' ELSE review_status END,
                updated_at = now()
            WHERE id = %s
            RETURNING id
            """,
            (item_id,),
        )


def run_tire_wheel_rework_check(item_id: str) -> None:
    item = fetch_one("SELECT * FROM inventory_items WHERE id = %s", (item_id,))
    if not item or item.get("inventory_type") != "tires_wheels":
        return
    execute(
        """
        UPDATE accounting_tasks
        SET status = 'completed', completed_at = now()
        WHERE item_id = %s AND task_type = 'tire_wheel_check' AND status = 'open'
        RETURNING id
        """,
        (item_id,),
    )
    data = fetch_one("SELECT * FROM item_tire_wheel_data WHERE item_id = %s", (item_id,)) or {}

    def add(role: str, field: str, priority: str = "normal", comment: str | None = None) -> None:
        upsert_rework_task_for_type(item_id, "tire_wheel_check", role, field, priority, comment)

    if not has_photo_type(item_id, "tire_overview"):
        add("Erfasser", "Gesamtfoto Reifen/Radsatz fehlt", "hoch", "Pflichtfoto für Reifen/Räder fehlt.")

    dot = str(data.get("dot") or "").strip()
    _, _, dot_valid = parse_dot_week_year(dot)
    if not dot or not dot_valid:
        add("Erfasser", "DOT fehlt oder unklar", "hoch", "DOT muss als WWYY plausibel erfasst werden.")

    if not str(data.get("tire_size") or "").strip():
        add("Erfasser", "Reifengröße fehlt", "hoch", "Reifengröße muss erfasst oder per Foto nachgewiesen werden.")

    if data.get("season") in {None, "", "unklar"}:
        add("Erfasser", "Saison unklar", "normal", "Sommer, Winter oder Ganzjahr muss eingeordnet werden.")

    if tire_tread_missing(data):
        add("Erfasser", "Profiltiefe fehlt", "hoch", "Profiltiefe muss für Satz oder Einzelreifen erfasst werden.")
    else:
        values = tire_tread_values(data)
        if any(value < 1.6 for value in values):
            add("Technik", "Profiltiefe unter gesetzlicher Mindestgrenze", "hoch", "Mindestens ein Reifen liegt unter 1,6 mm.")
        elif any(value < tire_warning_limit(data.get("season")) for value in values):
            add("Prüfer", "Profiltiefe unter fachlichem Warnwert", "normal", "Reifen liegt unter dem internen Warnwert.")

    if data.get("set_complete") is False:
        add("Erfasser", "Satz unvollständig", "hoch", "Radsatz ist nicht vollständig erfasst.")

    if str(data.get("damage_note") or "").strip():
        add("Technik", "Schaden vorhanden", "hoch", "Schaden muss fachlich bewertet werden.")

    if data.get("rim_present") is True and data.get("rim_type") in {None, "", "unklar"}:
        add("Erfasser", "Felgentyp unklar", "normal", "Felgentyp Stahl, Alu oder unklar prüfen.")

    if data.get("rim_present") is True and str(data.get("rim_condition") or "").strip().lower() in {"", "unklar"}:
        add("Erfasser", "Felgenzustand unklar", "normal", "Felgenzustand muss eingeordnet werden.")

    open_critical = fetch_one(
        """
        SELECT count(*)::int AS count
        FROM accounting_tasks
        WHERE item_id = %s AND status = 'open' AND assigned_role IN ('Erfasser', 'Technik')
        """,
        (item_id,),
    )
    if open_critical and int(open_critical["count"]) > 0:
        execute(
            """
            UPDATE inventory_items
            SET review_status = CASE
                  WHEN review_status = 'finalisiert' THEN review_status
                  ELSE 'nacharbeit_erfasser'
                END,
                status = CASE WHEN status = 'finalisiert' THEN status ELSE 'nacharbeit_noetig' END,
                updated_at = now()
            WHERE id = %s
            RETURNING id
            """,
            (item_id,),
        )
    else:
        execute(
            """
            UPDATE inventory_items
            SET review_status = CASE WHEN review_status IN ('erfasst', 'nacharbeit_erfasser', 'nacharbeit_technik') THEN 'finalisierbar' ELSE review_status END,
                updated_at = now()
            WHERE id = %s
            RETURNING id
            """,
            (item_id,),
        )


def run_inventory_rework_check(item_id: str) -> None:
    item = fetch_one("SELECT inventory_type FROM inventory_items WHERE id = %s", (item_id,))
    if not item:
        return
    if item.get("inventory_type") == "bga":
        run_bga_rework_check(item_id)
    elif item.get("inventory_type") == "tires_wheels":
        run_tire_wheel_rework_check(item_id)


def build_process_hints(row: dict[str, Any], ai_result: dict[str, Any], tasks: list[dict[str, Any]]) -> list[dict[str, str]]:
    hints: list[dict[str, str]] = []
    seen: set[str] = set()

    def add(kind: str, label: str, severity: str = "info") -> None:
        if kind in seen:
            return
        seen.add(kind)
        hints.append({"kind": kind, "label": label, "severity": severity})

    status = row.get("status")
    if status in {"ki_wartet", "ki_laeuft", "ki_schnell_wartet", "ki_schnell_laeuft"}:
        add("ki", "KI läuft", "info")
    if status == "ki_schnell_fertig":
        add("ki_quick_done", "Schnell-KI fertig", "ok")
    if status == "ki_pruefung_offen":
        add("ki_review_open", "Prüf-KI offen", "warn")
    if status in {"ki_pruefung_wartet", "ki_pruefung_laeuft"}:
        add("ki_review", "Prüf-KI läuft", "info")
    if status == "ki_pruefung_fertig":
        add("ki_review_done", "Prüf-KI fertig", "ok")
    if row.get("confidence_score") and float(row["confidence_score"]) >= 0.7:
        add("ki_confident", "KI sicher", "ok")

    special_matches = ai_result.get("special_tool_matches") or []
    history_matches = ai_result.get("inventory_history_matches") or []
    if special_matches:
        add("reference", "Referenztreffer", "info")
    if history_matches:
        add("history", "Historie", "info")

    uvv_resolved = row.get("uvv_status") in {"vorhanden", "nicht_uvv_pflichtig"}
    function_ok_resolved = row.get("function_ok") == "ja"
    inspection_book_resolved = row.get("inspection_book_available") in {"ja", "nicht_erforderlich"}

    for match in history_matches:
        if match.get("uvv_due") and not uvv_resolved:
            add("uvv", "UVV prüfen", "danger")
        if match.get("maintenance_due"):
            add("maintenance", "Wartung prüfen", "warn")
        if match.get("inspection_book_missing") and not inspection_book_resolved:
            add("inspection_book", "Prüfbuch fehlt", "warn")
        if match.get("missing"):
            add("target_missing", "Soll/Fehlt prüfen", "danger")
        if match.get("defective") and not function_ok_resolved:
            add("defective", "Defekt prüfen", "danger")

    for task in tasks:
        role = task.get("assigned_role")
        if role == "Technik":
            add("technical_rework", "Technik-Nacharbeit", "warn")
        elif role in {"Buchhaltung", "Auswertung"}:
            add("later_review", "Spätere Auswertung", "info")
        elif role == "Erfasser":
            add("capture_rework", "Erfasser-Nacharbeit", "warn")

    return hints


@app.on_event("startup")
def startup() -> None:
    run_migrations()
    ensure_upload_dirs()
    try:
        execute(
            f"""
            UPDATE inventory_items
            SET status = 'ki_pruefung_offen',
                updated_at = now()
            WHERE status IN ({AI_ACTIVE_STATUS_SQL})
              AND status <> 'finalisiert'
            """,
        )
    except Exception as exc:
        print(f"AI startup cleanup skipped: {type(exc).__name__}: {str(exc)[:160]}")


@app.get("/health")
def health() -> dict[str, Any]:
    db_ok = False
    migrations_ok = False
    upload_root_ok = False
    upload_free_mb = None
    try:
        db_ok = bool(fetch_one("SELECT 1 AS ok"))
        migrations_ok = bool(fetch_one("SELECT to_regclass('public.schema_migrations') IS NOT NULL AS ok")["ok"])
    except Exception:
        db_ok = False
        migrations_ok = False
    try:
        ensure_upload_dirs()
        test_path = Path(settings.upload_root, "temp", ".health-write")
        test_path.write_text("ok", encoding="utf-8")
        test_path.unlink(missing_ok=True)
        upload_root_ok = True
        usage = shutil.disk_usage(settings.upload_root)
        upload_free_mb = round(usage.free / 1024 / 1024)
    except Exception:
        upload_root_ok = False
    return {
        "ok": bool(db_ok and upload_root_ok),
        "database": db_ok,
        "migrations": migrations_ok,
        "upload_root": upload_root_ok,
        "upload_free_mb": upload_free_mb,
        "db_pool": db_pool_status(),
        "auth_secret_configured": settings.auth_secret_configured,
        "phase": "enterprise-foundation",
    }


def resolve_location(location_id: str | None, location_name: str | None) -> dict[str, Any]:
    name = location_name.strip() if location_name else None
    if name:
        location = fetch_one("SELECT * FROM locations WHERE lower(name) = lower(%s)", (name,))
        if not location:
            location = execute(
                """
                INSERT INTO locations (tenant_id, name, code)
                VALUES (%s, %s, %s)
                RETURNING *
                """,
                (default_tenant_id(), name, f"BETR-{secrets.token_hex(2).upper()}"),
            )
            audit("location_created", "location", str(location["id"]), location)
        return location
    if location_id:
        location = fetch_one("SELECT * FROM locations WHERE id = %s", (location_id,))
        if location:
            return location
    location = fetch_one("SELECT * FROM locations ORDER BY created_at LIMIT 1")
    if location:
        return location
    location = execute(
        """
        INSERT INTO locations (tenant_id, name, code)
        VALUES (%s, 'Betrieb', %s)
        RETURNING *
        """,
        (default_tenant_id(), f"LOC-{secrets.token_hex(2).upper()}"),
    )
    audit("location_created", "location", str(location["id"]), location)
    return location


@app.post("/auth/login")
def login(body: LoginIn) -> dict[str, Any]:
    login_name = body.email.strip()
    user = fetch_one(
        """
        SELECT u.*, t.slug AS tenant_slug, array_remove(array_agg(r.slug), null) AS roles
        FROM users u
        LEFT JOIN tenants t ON t.id = u.tenant_id
        LEFT JOIN user_roles ur ON ur.user_id = u.id
        LEFT JOIN roles r ON r.id = ur.role_id
        WHERE (lower(u.email) = lower(%s) OR lower(u.display_name) = lower(%s)) AND u.active = true
        GROUP BY u.id, t.slug
        ORDER BY CASE WHEN lower(u.email) = lower(%s) THEN 0 ELSE 1 END, u.created_at ASC
        LIMIT 1
        """,
        (login_name, login_name, login_name),
    )
    if not user or not verify_password(body.password, user.get("password_hash")):
        raise HTTPException(status_code=401, detail="Login Name oder Passwort ist falsch")
    execute("UPDATE users SET last_login_at = now(), updated_at = now() WHERE id = %s RETURNING id", (user["id"],))
    token = create_access_token(user)
    audit("login", "user", str(user["id"]), {"login_name": login_name, "email": user.get("email")})
    return {"access_token": token, "token_type": "bearer", "user": user}


@app.get("/auth/me")
def me(request: Request) -> dict[str, Any]:
    user = current_user_from_request(request)
    if not user:
        user = fetch_one(
            """
            SELECT u.*, t.slug AS tenant_slug, array_remove(array_agg(r.slug), null) AS roles
            FROM users u
            LEFT JOIN tenants t ON t.id = u.tenant_id
            LEFT JOIN user_roles ur ON ur.user_id = u.id
            LEFT JOIN roles r ON r.id = ur.role_id
            WHERE u.email = %s
            GROUP BY u.id, t.slug
            """,
            (settings.demo_user_email,),
        )
    return {"user": user}


@app.get("/meta/bootstrap")
def bootstrap() -> dict[str, Any]:
    return {
        "users": fetch_all(
            """
            SELECT u.id, u.email, u.display_name, array_remove(array_agg(r.slug), null) AS roles
            FROM users u
            LEFT JOIN user_roles ur ON ur.user_id = u.id
            LEFT JOIN roles r ON r.id = ur.role_id
            WHERE u.active = true
            GROUP BY u.id
            ORDER BY u.display_name
            """
        ),
        "locations": fetch_all("SELECT * FROM locations ORDER BY name"),
        "buildings": fetch_all("SELECT * FROM buildings ORDER BY name"),
        "rooms": fetch_all("SELECT * FROM rooms ORDER BY name"),
        "object_classes": fetch_all("SELECT * FROM object_classes ORDER BY name"),
        "brands": [row["name"] for row in fetch_all("SELECT name FROM brand_lexicon ORDER BY name")],
    }


def mobile_join_bootstrap(session: dict[str, Any]) -> dict[str, Any]:
    location = fetch_one(
        "SELECT id, name, code FROM locations WHERE id = %s",
        (session.get("location_id"),),
    ) if session.get("location_id") else None
    building = fetch_one(
        "SELECT id, name, location_id FROM buildings WHERE id = %s",
        (session.get("building_id"),),
    ) if session.get("building_id") else None
    room = fetch_one(
        "SELECT id, name, building_id, code, room_type FROM rooms WHERE id = %s",
        (session.get("room_id"),),
    ) if session.get("room_id") else None
    return {
        "users": [],
        "locations": [location] if location else [],
        "buildings": [building] if building else [],
        "rooms": [room] if room else [],
        "object_classes": fetch_all("SELECT id, name, slug FROM object_classes ORDER BY name"),
    }


def template_from_object_class(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": f"class:{row['id']}",
        "source": "Objektklasse",
        "label": row["name"],
        "subtitle": row.get("description") or "Standard-Vorlage",
        "object_type": row["name"],
        "object_class_id": str(row["id"]),
        "object_class_slug": row.get("slug"),
        "brand": None,
        "model": None,
    }


@app.get("/item-templates")
def item_templates(q: str = "", room: str = "", limit: int = 8) -> list[dict[str, Any]]:
    limit = max(1, min(limit, 12))
    query = " ".join([q, room]).strip()
    query_tokens = tokenize_reference_text(query)
    object_classes = fetch_all("SELECT * FROM object_classes ORDER BY name")
    class_by_slug = {row.get("slug"): row for row in object_classes}
    templates: list[dict[str, Any]] = [template_from_object_class(row) for row in object_classes]

    for entry in WORKSHOP_REFERENCE_CATALOG:
        oc = class_by_slug.get(entry.get("slug"))
        templates.append(
            {
                "id": f"workshop:{entry.get('slug')}",
                "source": "Werkstattprofil",
                "label": entry.get("object_class"),
                "subtitle": ", ".join((entry.get("typical_brands") or [])[:4]),
                "object_type": entry.get("object_class"),
                "object_class_id": str(oc["id"]) if oc else None,
                "object_class_slug": entry.get("slug"),
                "brand": None,
                "model": None,
            }
        )

    for record in load_special_tool_references()[:600]:
        label = record.get("designation_de") or record.get("designation_en") or record.get("vag_no") or record.get("order_no")
        if not label:
            continue
        templates.append(
            {
                "id": f"special:{record.get('row_id') or record.get('vag_no') or record.get('order_no')}",
                "source": "Spezialwerkzeug",
                "label": label,
                "subtitle": " · ".join(str(value) for value in [record.get("vag_no"), record.get("order_no"), record.get("source_file")] if value),
                "object_type": label,
                "object_class_id": None,
                "object_class_slug": None,
                "brand": None,
                "model": record.get("vag_no") or record.get("order_no"),
            }
        )

    for record in load_inventory_history_references()[:400]:
        label = record.get("designation_de") or record.get("tool_no")
        if not label:
            continue
        templates.append(
            {
                "id": f"history:{record.get('row_id') or record.get('tool_no')}",
                "source": "Alte Aufnahme",
                "label": label,
                "subtitle": " · ".join(str(value) for value in [record.get("tool_no"), record.get("action"), record.get("list_name")] if value),
                "object_type": label,
                "object_class_id": None,
                "object_class_slug": None,
                "brand": None,
                "model": record.get("tool_no"),
            }
        )

    seen: set[str] = set()
    scored: list[tuple[int, dict[str, Any]]] = []
    for template in templates:
        key = "|".join(str(template.get(field) or "").lower() for field in ["label", "subtitle", "source"])
        if key in seen:
            continue
        seen.add(key)
        haystack = " ".join(str(template.get(field) or "") for field in ["label", "subtitle", "source", "object_class_slug"])
        tokens = tokenize_reference_text(haystack)
        if query_tokens:
            score = len(query_tokens & tokens)
            if str(template.get("label") or "").lower().startswith(q.lower().strip()):
                score += 4
            if score <= 0:
                continue
        else:
            score = 1 if template["source"] in {"Objektklasse", "Werkstattprofil"} else 0
            if score <= 0:
                continue
        scored.append((score, template))
    scored.sort(key=lambda item: (item[0], item[1].get("source") == "Objektklasse"), reverse=True)
    return [template for _, template in scored[:limit]]


@app.post("/users")
def create_user(body: UserIn) -> dict[str, Any]:
    name = body.display_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name fehlt")
    email = (body.email or f"{name.lower().replace(' ', '.')}@example.local").strip().lower()
    existing = fetch_one("SELECT * FROM users WHERE lower(email) = lower(%s)", (email,))
    if existing:
        return existing
    initial_password = secrets.token_urlsafe(18)
    user = execute(
        """
        INSERT INTO users (tenant_id, email, display_name, password_hash, password_reset_required)
        VALUES (%s, %s, %s, %s, true)
        RETURNING *
        """,
        (default_tenant_id(), email, name, hash_password(initial_password)),
    )
    role = fetch_one("SELECT id FROM roles WHERE slug = %s", (body.role_slug,))
    if role:
        execute(
            "INSERT INTO user_roles (user_id, role_id) VALUES (%s, %s) ON CONFLICT DO NOTHING RETURNING user_id",
            (user["id"], role["id"]),
        )
    audit("user_created", "user", str(user["id"]), {"email": email, "role_slug": body.role_slug, "password_reset_required": True})
    return {**user, "initial_password": initial_password}


@app.post("/locations")
def create_location(body: LocationIn) -> dict[str, Any]:
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Betriebsname fehlt")
    existing = fetch_one("SELECT * FROM locations WHERE lower(name) = lower(%s)", (name,))
    if existing:
        return existing
    code = (body.code or f"BETR-{secrets.token_hex(2).upper()}").strip().upper()
    row = execute(
        """
        INSERT INTO locations (tenant_id, name, code, address)
        VALUES (%s, %s, %s, %s)
        RETURNING *
        """,
        (default_tenant_id(), name, code, body.address),
    )
    audit("location_created", "location", str(row["id"]), row)
    return row


@app.post("/rooms")
def create_room(body: RoomIn) -> dict[str, Any]:
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Raumname fehlt")
    building_id = body.building_id
    building_name = body.building_name.strip() if body.building_name else None
    if building_name:
        location = resolve_location(body.location_id, body.location_name)
        building = fetch_one(
            "SELECT * FROM buildings WHERE location_id = %s AND lower(name) = lower(%s)",
            (location["id"], building_name),
        )
        if not building:
            building = execute(
                """
                INSERT INTO buildings (tenant_id, location_id, name, code)
                VALUES (%s, %s, %s, %s)
                RETURNING *
                """,
                (location.get("tenant_id") or default_tenant_id(), location["id"], building_name, f"GEB-{secrets.token_hex(3).upper()}"),
            )
            audit("building_created", "building", str(building["id"]), building)
        building_id = str(building["id"])
    if not building_id or not fetch_one("SELECT id FROM buildings WHERE id = %s", (building_id,)):
        raise HTTPException(status_code=400, detail="Gebäude auswählen oder neues Gebäude eingeben")
    code = (body.code or f"RAUM-{secrets.token_hex(3).upper()}").strip()
    row = execute(
        """
        INSERT INTO rooms (tenant_id, building_id, name, code, room_type)
        VALUES ((SELECT tenant_id FROM buildings WHERE id = %s), %s, %s, %s, %s)
        RETURNING *
        """,
        (building_id, building_id, name, code, body.room_type),
    )
    audit("room_created", "room", str(row["id"]), row)
    return row


@app.patch("/rooms/{room_id}")
def update_room(room_id: str, body: RoomPatch) -> dict[str, Any]:
    current = fetch_one("SELECT * FROM rooms WHERE id = %s", (room_id,))
    if not current:
        raise HTTPException(status_code=404, detail="Raum nicht gefunden")
    building_id = body.building_id if body.building_id is not None else current["building_id"]
    if not fetch_one("SELECT id FROM buildings WHERE id = %s", (building_id,)):
        raise HTTPException(status_code=400, detail="Gebäude nicht gefunden")
    name = body.name.strip() if body.name is not None else current["name"]
    code = body.code.strip() if body.code is not None else current["code"]
    room_type = body.room_type.strip() if body.room_type is not None else current["room_type"]
    if not name:
        raise HTTPException(status_code=400, detail="Raumname fehlt")
    row = execute(
        """
        UPDATE rooms
        SET building_id = %s, name = %s, code = %s, room_type = %s
        WHERE id = %s
        RETURNING *
        """,
        (building_id, name, code, room_type, room_id),
    )
    audit("room_changed", "room", room_id, {"building_id": building_id, "name": name, "code": code, "room_type": room_type})
    return row


@app.delete("/rooms/{room_id}")
def delete_room(room_id: str, force: bool = False) -> dict[str, Any]:
    current = fetch_one("SELECT * FROM rooms WHERE id = %s", (room_id,))
    if not current:
        raise HTTPException(status_code=404, detail="Raum nicht gefunden")
    usage = fetch_one(
        """
        SELECT
          (SELECT count(*) FROM inventory_sessions WHERE room_id = %s) AS session_count,
          (SELECT count(*) FROM inventory_items WHERE room_id = %s) AS item_count
        """,
        (room_id, room_id),
    )
    has_usage = bool(usage and (usage["session_count"] or usage["item_count"]))
    if has_usage and not force:
        raise HTTPException(
            status_code=409,
            detail="Raum enthält noch Sessions oder Gegenstände. Nutze Löschen mit Bestätigung, um Testdaten mit zu entfernen.",
        )
    if has_usage:
        audit("room_delete_cascade_requested", "room", room_id, {"room": current, "usage": usage})
        execute("DELETE FROM inventory_sessions WHERE room_id = %s RETURNING id", (room_id,))
    row = execute("DELETE FROM rooms WHERE id = %s RETURNING *", (room_id,))
    audit("room_deleted", "room", room_id, {"room": row, "deleted_dependents": usage if has_usage else None})
    return {"deleted": True, "room": row, "deleted_dependents": usage if has_usage else None}


@app.post("/sessions")
def create_session(body: SessionIn) -> dict[str, Any]:
    token = secrets.token_urlsafe(18)
    inventory_type = body.inventory_type if body.inventory_type in {"bga", "tires_wheels", "special_tools"} else "bga"
    room_id = body.room_id
    building_id = body.building_id
    location_id = body.location_id
    location_name = body.location_name.strip() if body.location_name else None
    building_name = body.building_name.strip() if body.building_name else None

    if body.room_name and not room_id:
        room_name = body.room_name.strip()
        if not room_name:
            raise HTTPException(status_code=400, detail="Raumname fehlt")
        if building_name:
            location = resolve_location(location_id, location_name)
            location_id = str(location["id"])
            building = fetch_one(
                "SELECT * FROM buildings WHERE location_id = %s AND lower(name) = lower(%s)",
                (location_id, building_name),
            )
            if not building:
                building = execute(
                    """
                    INSERT INTO buildings (tenant_id, location_id, name, code)
                    VALUES (%s, %s, %s, %s)
                    RETURNING *
                    """,
                    (location.get("tenant_id") or default_tenant_id(), location_id, building_name, f"FREI-{secrets.token_hex(3).upper()}"),
                )
                audit("building_created", "building", str(building["id"]), building)
            building_id = str(building["id"])
        elif not building_id:
            location = resolve_location(location_id, location_name)
            location_id = str(location["id"])
            building = fetch_one("SELECT * FROM buildings WHERE location_id = %s ORDER BY created_at LIMIT 1", (location_id,))
            if not building:
                building = execute(
                    """
                    INSERT INTO buildings (tenant_id, location_id, name, code)
                    VALUES (%s, %s, 'Hauptgebäude', %s)
                    RETURNING *
                    """,
                    (location.get("tenant_id") or default_tenant_id(), location["id"], f"HG-{secrets.token_hex(2).upper()}"),
                )
                audit("building_created", "building", str(building["id"]), building)
            building_id = str(building["id"])
        else:
            building = fetch_one("SELECT * FROM buildings WHERE id = %s", (building_id,))
            if not building:
                raise HTTPException(status_code=400, detail="Gebäude nicht gefunden")
        location_id = location_id or str(building["location_id"])
        room = fetch_one(
            "SELECT * FROM rooms WHERE building_id = %s AND lower(name) = lower(%s)",
            (building_id, room_name),
        )
        if not room:
            room = execute(
                """
                INSERT INTO rooms (tenant_id, building_id, name, code, room_type)
                VALUES (%s, %s, %s, %s, 'workspace')
                RETURNING *
                """,
                (building.get("tenant_id") or default_tenant_id(), building_id, room_name, f"FREI-{secrets.token_hex(3).upper()}"),
            )
            audit("room_created", "room", str(room["id"]), room)
        room_id = str(room["id"])

    if not room_id:
        raise HTTPException(status_code=400, detail="Raum auswählen oder freien Raum eingeben")

    room = fetch_one(
        """
        SELECT r.*, b.location_id, COALESCE(r.tenant_id, b.tenant_id, l.tenant_id) AS tenant_id
        FROM rooms r
        JOIN buildings b ON b.id = r.building_id
        JOIN locations l ON l.id = b.location_id
        WHERE r.id = %s
        """,
        (room_id,),
    )
    if not room:
        raise HTTPException(status_code=400, detail="Raum nicht gefunden")
    building_id = str(room["building_id"])
    location_id = str(room["location_id"])

    row = execute(
        """
        INSERT INTO inventory_sessions (tenant_id, location_id, building_id, room_id, join_token, join_token_expires_at, started_by, inventory_type)
        VALUES (%s, %s, %s, %s, %s, now() + interval '36 hours', %s, %s)
        RETURNING *
        """,
        (room.get("tenant_id") or default_tenant_id(), location_id, building_id, room_id, token, body.started_by, inventory_type),
    )
    audit("session_started", "inventory_session", str(row["id"]), row)
    return row


@app.get("/sessions")
def list_sessions() -> list[dict[str, Any]]:
    return fetch_all(
        """
        SELECT s.*, l.name AS location_name, b.name AS building_name, r.name AS room_name,
          starter.display_name AS started_by_name,
          COALESCE(s.inventory_type, 'bga') AS inventory_type,
          count(i.id)::int AS item_count
        FROM inventory_sessions s
        JOIN locations l ON l.id = s.location_id
        JOIN buildings b ON b.id = s.building_id
        JOIN rooms r ON r.id = s.room_id
        LEFT JOIN users starter ON starter.id = s.started_by
        LEFT JOIN inventory_items i ON i.session_id = s.id
        GROUP BY s.id, l.name, b.name, r.name, starter.display_name
        ORDER BY s.created_at DESC
        """
    )


@app.get("/sessions/{session_id}")
def get_session(session_id: str) -> dict[str, Any]:
    row = fetch_one(
        """
        SELECT s.*, COALESCE(s.inventory_type, 'bga') AS inventory_type,
          l.name AS location_name, l.code AS location_code, b.name AS building_name, r.name AS room_name
        FROM inventory_sessions s
        JOIN locations l ON l.id = s.location_id
        JOIN buildings b ON b.id = s.building_id
        JOIN rooms r ON r.id = s.room_id
        WHERE s.id = %s
        """,
        (session_id,),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Session not found")
    return row


@app.post("/sessions/{session_id}/join-token")
def renew_join_token(session_id: str) -> dict[str, Any]:
    token = secrets.token_urlsafe(18)
    row = execute(
        "UPDATE inventory_sessions SET join_token = %s, join_token_expires_at = now() + interval '36 hours' WHERE id = %s RETURNING *",
        (token, session_id),
    )
    audit("join_token_created", "inventory_session", session_id, {"join_token": token})
    return row


def first_damage_access_room(tenant_id: str | None) -> dict[str, Any]:
    room = fetch_one(
        """
        SELECT r.id AS room_id, b.id AS building_id, l.id AS location_id
        FROM rooms r
        JOIN buildings b ON b.id = r.building_id
        JOIN locations l ON l.id = b.location_id
        WHERE COALESCE(r.tenant_id, b.tenant_id, l.tenant_id) IS NOT DISTINCT FROM %s
        ORDER BY r.created_at ASC
        LIMIT 1
        """,
        (tenant_id,),
    )
    if room:
        return room
    location = execute(
        """
        INSERT INTO locations (tenant_id, name, code)
        VALUES (%s, %s, %s)
        RETURNING *
        """,
        (tenant_id, "Schadenerfassung", f"SCHADEN-{secrets.token_hex(3).upper()}"),
    )
    building = execute(
        """
        INSERT INTO buildings (tenant_id, location_id, name, code)
        VALUES (%s, %s, %s, %s)
        RETURNING *
        """,
        (tenant_id, location["id"], "Gemeinsame Liste", "SCHADEN"),
    )
    created_room = execute(
        """
        INSERT INTO rooms (tenant_id, building_id, name, code, room_type)
        VALUES (%s, %s, %s, %s, 'workspace')
        RETURNING *
        """,
        (tenant_id, building["id"], "Schadensliste", "SCHADEN"),
    )
    return {
        "location_id": location["id"],
        "building_id": building["id"],
        "room_id": created_room["id"],
    }


@app.post("/damage-reports/access-session")
def ensure_damage_access_session(request: Request) -> dict[str, Any]:
    auth = getattr(request.state, "auth", {}) or {}
    if auth.get("kind") == "mobile_session" and auth.get("session_id"):
        return get_session(str(auth["session_id"]))
    tenant_id = request_tenant_id(request)
    existing = fetch_one(
        """
        SELECT s.*, (s.join_token_expires_at <= now() + interval '30 hours') AS token_needs_refresh
        FROM inventory_sessions s
        WHERE s.tenant_id IS NOT DISTINCT FROM %s AND s.status = 'open'
        ORDER BY s.created_at DESC
        LIMIT 1
        """,
        (tenant_id,),
    )
    if existing:
        if existing.get("token_needs_refresh"):
            token = secrets.token_urlsafe(18)
            execute(
                "UPDATE inventory_sessions SET join_token = %s, join_token_expires_at = now() + interval '36 hours' WHERE id = %s RETURNING *",
                (token, existing["id"]),
            )
        return get_session(str(existing["id"]))
    room = first_damage_access_room(tenant_id)
    token = secrets.token_urlsafe(18)
    started_by = auth.get("sub") if auth.get("kind") == "user" else None
    row = execute(
        """
        INSERT INTO inventory_sessions (
          tenant_id, location_id, building_id, room_id, join_token, join_token_expires_at, started_by, inventory_type
        )
        VALUES (%s, %s, %s, %s, %s, now() + interval '36 hours', %s, 'bga')
        RETURNING *
        """,
        (tenant_id, room["location_id"], room["building_id"], room["room_id"], token, started_by),
    )
    audit("damage_access_session_created", "inventory_session", str(row["id"]), row)
    return get_session(str(row["id"]))


@app.post("/sessions/join")
def join_session(body: JoinIn) -> dict[str, Any]:
    session = fetch_one(
        "SELECT * FROM inventory_sessions WHERE join_token = %s AND join_token_expires_at > now() AND status = 'open'",
        (body.token,),
    )
    if not session:
        raise HTTPException(status_code=404, detail="Join token invalid or expired")
    device = None
    if body.device_fingerprint:
        device = fetch_one(
            """
            SELECT *
            FROM session_devices
            WHERE session_id = %s AND device_fingerprint = %s
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (session["id"], body.device_fingerprint),
        )
        if device and device.get("revoked_at"):
            raise HTTPException(status_code=403, detail="Dieses Gerät wurde für diese Session widerrufen")
    if body.device_fingerprint:
        device = execute(
            """
            INSERT INTO session_devices (tenant_id, session_id, device_name, device_fingerprint, last_seen_at)
            VALUES (%s, %s, %s, %s, now())
            ON CONFLICT (session_id, device_fingerprint) WHERE device_fingerprint IS NOT NULL
            DO UPDATE
              SET device_name = EXCLUDED.device_name,
                  last_seen_at = now()
              WHERE session_devices.revoked_at IS NULL
            RETURNING *
            """,
            (session.get("tenant_id") or default_tenant_id(), session["id"], body.device_name, body.device_fingerprint),
        )
        if not device:
            raise HTTPException(status_code=403, detail="Dieses Gerät wurde für diese Session widerrufen")
        audit("device_seen", "session_device", str(device["id"]), device)
    else:
        device = execute(
            """
            INSERT INTO session_devices (tenant_id, session_id, device_name, device_fingerprint, last_seen_at)
            VALUES (%s, %s, %s, %s, now())
            RETURNING *
            """,
            (session.get("tenant_id") or default_tenant_id(), session["id"], body.device_name, body.device_fingerprint),
        )
        audit("device_joined", "session_device", str(device["id"]), device)
    return {
        "session": session,
        "device": device,
        "access_token": create_mobile_session_token(session, device),
        "token_type": "bearer",
        "bootstrap": mobile_join_bootstrap(session),
    }


@app.post("/sessions/{session_id}/close")
def close_session(session_id: str) -> dict[str, Any]:
    items = fetch_all("SELECT id FROM inventory_items WHERE session_id = %s", (session_id,))
    blockers = {str(i["id"]): finalization_blockers(str(i["id"])) for i in items}
    open_blockers = {k: v for k, v in blockers.items() if v}
    if open_blockers:
        raise HTTPException(status_code=409, detail={"message": "Offene Punkte für Raumabschluss", "blockers": open_blockers})
    row = execute(
        "UPDATE inventory_sessions SET status = 'closed', closed_at = now() WHERE id = %s RETURNING *",
        (session_id,),
    )
    execute(
        "UPDATE inventory_items SET locked_at = COALESCE(locked_at, now()), updated_at = now() WHERE session_id = %s RETURNING id",
        (session_id,),
    )
    audit("session_closed", "inventory_session", session_id, row)
    return row


@app.post("/sessions/{session_id}/reopen")
def reopen_session(session_id: str, body: ReopenIn | None = None) -> dict[str, Any]:
    current = fetch_one("SELECT * FROM inventory_sessions WHERE id = %s", (session_id,))
    if not current:
        raise HTTPException(status_code=404, detail="Session nicht gefunden")
    if current["status"] == "open":
        return current
    row = execute(
        """
        UPDATE inventory_sessions
        SET status = 'open',
            closed_at = NULL,
            closed_by = NULL,
            join_token_expires_at = now() + interval '36 hours'
        WHERE id = %s
        RETURNING *
        """,
        (session_id,),
    )
    execute(
        """
        UPDATE inventory_items
        SET locked_at = NULL,
            updated_at = now()
        WHERE session_id = %s
          AND status <> 'finalisiert'
        RETURNING id
        """,
        (session_id,),
    )
    reason = (body.reason if body else None) or "Raum wurde für Korrekturen wieder geöffnet"
    execute("UPDATE inventory_items SET reopened_reason = %s WHERE session_id = %s RETURNING id", (reason, session_id))
    audit("session_reopened", "inventory_session", session_id, row, reason=reason)
    return row


@app.delete("/sessions/{session_id}")
def delete_session(session_id: str) -> dict[str, Any]:
    current = fetch_one("SELECT * FROM inventory_sessions WHERE id = %s", (session_id,))
    if not current:
        raise HTTPException(status_code=404, detail="Session nicht gefunden")
    item_count = fetch_one("SELECT count(*)::int AS count FROM inventory_items WHERE session_id = %s", (session_id,))
    audit("session_deleted", "inventory_session", session_id, {"session": current, "item_count": item_count["count"] if item_count else 0})
    row = execute("DELETE FROM inventory_sessions WHERE id = %s RETURNING *", (session_id,))
    return {"deleted": True, "session": row}


@app.get("/sessions/{session_id}/devices")
def devices(session_id: str) -> list[dict[str, Any]]:
    return fetch_all("SELECT * FROM session_devices WHERE session_id = %s ORDER BY created_at DESC", (session_id,))


@app.post("/sessions/{session_id}/devices/{device_id}/revoke")
def revoke_device(session_id: str, device_id: str) -> dict[str, Any]:
    row = execute(
        "UPDATE session_devices SET revoked_at = now() WHERE id = %s AND session_id = %s RETURNING *",
        (device_id, session_id),
    )
    audit("device_revoked", "session_device", device_id, row)
    return row


@app.post("/items")
def create_item(body: ItemIn) -> dict[str, Any]:
    session = get_session(body.session_id)
    if session["status"] != "open":
        raise HTTPException(status_code=409, detail="Raum ist abgeschlossen")
    session_inventory_type = session.get("inventory_type") or "bga"
    if body.inventory_type and body.inventory_type != session_inventory_type:
        raise HTTPException(status_code=409, detail="Erfassungsart der Session kann nicht gemischt werden")
    if body.client_item_id and body.source_device_id:
        existing = fetch_one(
            """
            SELECT *
            FROM inventory_items
            WHERE session_id = %s AND source_device_id = %s AND client_item_id = %s
            LIMIT 1
            """,
            (body.session_id, body.source_device_id, body.client_item_id),
        )
        if existing:
            return existing
    inventory_id = body.inventory_id or next_inventory_id(session["location_code"])
    temporary_id = body.temporary_id or f"TEMP-{secrets.token_hex(4).upper()}"
    object_class_id = body.object_class_id
    if not object_class_id and session_inventory_type in {"bga", "tires_wheels", "special_tools"}:
        object_class = fetch_one("SELECT id FROM object_classes WHERE slug = %s", (session_inventory_type,))
        object_class_id = str(object_class["id"]) if object_class else None
    oc = fetch_one("SELECT * FROM object_classes WHERE id = %s", (object_class_id,)) if object_class_id else None
    sequence_number = next_sequence_number(body.session_id)
    row = execute(
        """
        INSERT INTO inventory_items (
          tenant_id, inventory_id, temporary_id, sequence_number, inventory_type,
          client_item_id, source_device_id,
          session_id, location_id, building_id, room_id,
          object_type, object_class_id, category, brand, model, serial_number, condition,
          condition_note, commercial_category, requires_accounting_review,
          accounting_relevance, created_by, specification, construction_year,
          function_ok, uvv_status, uvv_valid_until, inspection_book_available,
          remark, type_plate_status
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING *
        """,
        (
            session.get("tenant_id") or default_tenant_id(),
            inventory_id,
            temporary_id,
            sequence_number,
            session_inventory_type,
            body.client_item_id,
            body.source_device_id,
            body.session_id,
            session["location_id"],
            session["building_id"],
            session["room_id"],
            body.object_type,
            object_class_id,
            session_inventory_type,
            body.brand,
            body.model,
            body.serial_number,
            body.condition,
            body.condition_note,
            oc["default_commercial_category"] if oc else "ungeklaert",
            oc["requires_accounting_review"] if oc else False,
            oc["requires_accounting_review"] if oc else False,
            body.created_by,
            body.specification,
            body.construction_year,
            body.function_ok,
            body.uvv_status,
            body.uvv_valid_until,
            body.inspection_book_available,
            body.remark,
            body.type_plate_status,
        ),
    )
    execute(
        """
        INSERT INTO item_accounting_data (item_id, commercial_category, accounting_status)
        VALUES (%s, %s, %s)
        RETURNING id
        """,
        (row["id"], row["commercial_category"], row["accounting_status"]),
    )
    audit("item_created", "inventory_item", str(row["id"]), row)
    run_inventory_rework_check(str(row["id"]))
    return row


@app.get("/sessions/{session_id}/items")
def session_items(session_id: str) -> list[dict[str, Any]]:
    repair_stale_ai_statuses(session_id=session_id)
    rows = fetch_all(
        """
        SELECT i.*, oc.name AS object_class_name,
          EXISTS(SELECT 1 FROM item_photos p WHERE p.item_id = i.id AND p.photo_type IN ('object','object_front')) AS has_object_photo,
          EXISTS(SELECT 1 FROM item_photos p WHERE p.item_id = i.id AND p.photo_type IN ('nameplate','type_plate')) AS has_nameplate_photo,
          EXISTS(SELECT 1 FROM item_photos p WHERE p.item_id = i.id AND p.photo_type = 'dot') AS has_dot_photo,
          (
            SELECT p.id
            FROM item_photos p
            WHERE p.item_id = i.id AND p.photo_type IN ('object','object_front')
            ORDER BY p.uploaded_at DESC
            LIMIT 1
          ) AS object_photo_id
        FROM inventory_items i
        LEFT JOIN object_classes oc ON oc.id = i.object_class_id
        WHERE i.session_id = %s
        ORDER BY i.created_at DESC
        """,
        (session_id,),
    )
    for row in rows:
        row["blockers"] = finalization_blockers(str(row["id"]))
        row["open_tasks"] = fetch_all(
            "SELECT * FROM accounting_tasks WHERE item_id = %s AND status = 'open' ORDER BY created_at",
            (row["id"],),
        )
        ai_row = fetch_one(
            "SELECT result_json FROM ai_results WHERE item_id = %s AND ai_type <> 'deep_dive' ORDER BY created_at DESC LIMIT 1",
            (row["id"],),
        )
        deep_dive_row = fetch_one(
            """
            SELECT result_json
            FROM ai_results
            WHERE item_id = %s AND ai_type = 'deep_dive'
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (row["id"],),
        )
        ai_result = ai_row.get("result_json") if ai_row else {}
        deep_dive = deep_dive_row.get("result_json") if deep_dive_row else None
        row["ai_summary"] = {
            "confidence": ai_result.get("confidence"),
            "notes": ai_result.get("notes"),
            "bga_detection": ai_result.get("bga_detection"),
            "nameplate_extraction": ai_result.get("nameplate_extraction"),
            "suggested_fields": ai_result.get("suggested_fields"),
            "requires_manual_review": ai_result.get("requires_manual_review") or ai_result.get("requires_review"),
            "uncertainty_reason": ai_result.get("uncertainty_reason"),
            "special_tool_matches": (ai_result.get("special_tool_matches") or [])[:3],
            "inventory_history_matches": (ai_result.get("inventory_history_matches") or [])[:3],
            "deep_dive": deep_dive,
        }
        row["process_hints"] = build_process_hints(row, ai_result, row["open_tasks"])
        row["photos"] = fetch_all(
            """
            SELECT id, photo_type, uploaded_at
            FROM item_photos
            WHERE item_id = %s
            ORDER BY uploaded_at ASC
            LIMIT 5
            """,
            (row["id"],),
        )
    return rows


TIRE_WHEEL_FIELDS = [
    "set_type",
    "season",
    "manufacturer",
    "profile_model",
    "tire_size",
    "load_index",
    "speed_index",
    "dot",
    "production_week",
    "production_year",
    "tread_depth_front_left",
    "tread_depth_front_right",
    "tread_depth_rear_left",
    "tread_depth_rear_right",
    "tread_depth_single",
    "rim_present",
    "rim_type",
    "rim_condition",
    "tire_condition",
    "damage_note",
    "set_complete",
    "storage_location",
    "remark",
]


def require_tire_wheel_item(item_id: str, require_open: bool = False) -> dict[str, Any]:
    item = require_item_session_open(item_id) if require_open else fetch_one("SELECT * FROM inventory_items WHERE id = %s", (item_id,))
    if not item:
        raise HTTPException(status_code=404, detail="Gegenstand nicht gefunden")
    if item.get("inventory_type") != "tires_wheels":
        raise HTTPException(status_code=409, detail="Gegenstand ist kein Reifen/Räder-Datensatz")
    return item


def upsert_tire_wheel_data(item_id: str, data: dict[str, Any]) -> dict[str, Any]:
    payload = normalize_tire_wheel_payload({key: value for key, value in data.items() if key in TIRE_WHEEL_FIELDS})
    if not payload:
        row = fetch_one("SELECT * FROM item_tire_wheel_data WHERE item_id = %s", (item_id,))
        if row:
            return row
        payload = {}

    insert_fields = ["item_id", *payload.keys()]
    placeholders = ", ".join(["%s"] * len(insert_fields))
    update_fields = [field for field in payload.keys() if field != "item_id"]
    update_sql = ", ".join([f"{field} = EXCLUDED.{field}" for field in update_fields])
    if update_sql:
        update_sql += ", updated_at = now()"
    else:
        update_sql = "updated_at = now()"
    row = execute(
        f"""
        INSERT INTO item_tire_wheel_data ({", ".join(insert_fields)})
        VALUES ({placeholders})
        ON CONFLICT (item_id) DO UPDATE SET {update_sql}
        RETURNING *
        """,
        (item_id, *payload.values()),
    )
    audit("tire_wheel_data_saved", "inventory_item", item_id, row)
    run_tire_wheel_rework_check(item_id)
    return row


@app.get("/items/resolve-client")
def resolve_client_item(session_id: str, source_device_id: str, client_item_id: str) -> dict[str, Any]:
    row = fetch_one(
        """
        SELECT id, session_id, source_device_id, client_item_id
        FROM inventory_items
        WHERE session_id = %s AND source_device_id = %s AND client_item_id = %s
        LIMIT 1
        """,
        (session_id, source_device_id, client_item_id),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Client item mapping not found")
    return row


@app.get("/items/{item_id}")
def get_item(item_id: str) -> dict[str, Any]:
    row = fetch_one("SELECT * FROM inventory_items WHERE id = %s", (item_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    row["photos"] = fetch_all("SELECT * FROM item_photos WHERE item_id = %s ORDER BY uploaded_at", (item_id,))
    row["audio_notes"] = fetch_all("SELECT * FROM item_audio_notes WHERE item_id = %s ORDER BY uploaded_at", (item_id,))
    row["ai_results"] = fetch_all("SELECT * FROM ai_results WHERE item_id = %s ORDER BY created_at DESC", (item_id,))
    row["open_tasks"] = fetch_all(
        "SELECT * FROM accounting_tasks WHERE item_id = %s AND status = 'open' ORDER BY created_at",
        (item_id,),
    )
    row["accounting"] = fetch_one("SELECT * FROM item_accounting_data WHERE item_id = %s", (item_id,))
    if row.get("inventory_type") == "tires_wheels":
        row["tire_wheel_data"] = fetch_one("SELECT * FROM item_tire_wheel_data WHERE item_id = %s", (item_id,))
    row["blockers"] = finalization_blockers(item_id)
    return row


@app.get("/items/{item_id}/tires-wheels")
def get_tire_wheel_data(item_id: str) -> dict[str, Any]:
    require_tire_wheel_item(item_id)
    row = fetch_one("SELECT * FROM item_tire_wheel_data WHERE item_id = %s", (item_id,))
    return row or {"item_id": item_id}


@app.post("/items/{item_id}/tires-wheels")
def create_tire_wheel_data(item_id: str, body: TireWheelDataCreate) -> dict[str, Any]:
    require_tire_wheel_item(item_id, require_open=True)
    data = body.model_dump(exclude_unset=True)
    return upsert_tire_wheel_data(item_id, data)


@app.patch("/items/{item_id}/tires-wheels")
def patch_tire_wheel_data(item_id: str, body: TireWheelDataUpdate) -> dict[str, Any]:
    require_tire_wheel_item(item_id, require_open=True)
    data = body.model_dump(exclude_unset=True)
    return upsert_tire_wheel_data(item_id, data)


@app.patch("/items/{item_id}")
def patch_item(item_id: str, body: ItemPatch) -> dict[str, Any]:
    require_item_session(item_id)
    data = body.model_dump(exclude_unset=True)
    if not data:
        return get_item(item_id)
    if data.get("age_source") in (None, ""):
        data["age_source"] = "unbekannt"
    if data.get("age_verification_status") in (None, ""):
        data["age_verification_status"] = "offen"
    if data.get("construction_year"):
        derived_age = construction_year_age({"construction_year": data.get("construction_year")})
        if derived_age is not None:
            typed_age = numeric_or_none(data.get("estimated_age_years"))
            if typed_age is None or abs(typed_age - derived_age) < 0.01:
                data["estimated_age_years"] = derived_age
                data["age_source"] = "baujahr"
                data["age_verification_status"] = "geprueft"
    reviewer_id = demo_user_id()
    allowed = list(data.keys())
    sql_parts = [f"{key} = %s" for key in allowed]
    if reviewer_id:
        sql_parts.append("reviewed_by = COALESCE(reviewed_by, %s)")
    sql = ", ".join(sql_parts)
    row = execute(
        f"UPDATE inventory_items SET {sql}, updated_at = now() WHERE id = %s RETURNING *",
        tuple(data[key] for key in allowed) + ((reviewer_id,) if reviewer_id else ()) + (item_id,),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    audit("item_changed", "inventory_item", item_id, data)
    run_inventory_rework_check(item_id)
    return get_item(item_id)


@app.delete("/items/{item_id}")
def delete_item(item_id: str) -> dict[str, Any]:
    require_item_session_open(item_id)
    current = fetch_one("SELECT * FROM inventory_items WHERE id = %s", (item_id,))
    if not current:
        raise HTTPException(status_code=404, detail="Gegenstand nicht gefunden")
    audit("item_deleted", "inventory_item", item_id, current)
    row = execute("DELETE FROM inventory_items WHERE id = %s RETURNING *", (item_id,))
    return {"deleted": True, "item": row}


@app.post("/items/{item_id}/finalize")
def finalize_item(item_id: str) -> dict[str, Any]:
    require_item_session(item_id)
    blockers = finalization_blockers(item_id)
    if blockers:
        raise HTTPException(status_code=409, detail={"message": "Item has blockers", "blockers": blockers})
    user_id = demo_user_id()
    row = execute(
        """
        UPDATE inventory_items
        SET review_status = 'finalisiert', status = 'finalisiert',
            lifecycle_status = COALESCE(lifecycle_status, 'aktiv'),
            reviewed_by = COALESCE(reviewed_by, %s),
            finalized_by = COALESCE(finalized_by, %s),
            finalized_at = now(), updated_at = now()
        WHERE id = %s RETURNING *
        """,
        (user_id, user_id, item_id),
    )
    audit("item_finalized", "inventory_item", item_id, row)
    return row


@app.post("/items/{item_id}/request-rework")
def request_rework(item_id: str, body: dict[str, Any]) -> dict[str, Any]:
    require_item_session(item_id)
    assigned_role = body.get("assigned_role", "Prüfer")
    row = execute(
        """
        INSERT INTO accounting_tasks (item_id, task_type, assigned_role, missing_field, priority, comment)
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING *
        """,
        (
            item_id,
            body.get("task_type", "rework"),
            assigned_role,
            body.get("missing_field"),
            body.get("priority", "normal"),
            body.get("comment"),
        ),
    )
    review_status = {
        "Buchhaltung": "nacharbeit_buchhaltung",
        "Auswertung": "nacharbeit_buchhaltung",
        "Erfasser": "nacharbeit_erfasser",
        "Technik": "nacharbeit_technik",
    }.get(assigned_role, "nacharbeit_pruefer")
    execute("UPDATE inventory_items SET review_status = %s WHERE id = %s RETURNING id", (review_status, item_id))
    audit("rework_requested", "inventory_item", item_id, row)
    return row


@app.post("/items/{item_id}/change-status")
def change_status(item_id: str, body: dict[str, Any]) -> dict[str, Any]:
    require_item_session(item_id)
    row = execute(
        "UPDATE inventory_items SET status = %s, review_status = COALESCE(%s, review_status), updated_at = now() WHERE id = %s RETURNING *",
        (body.get("status"), body.get("review_status"), item_id),
    )
    audit("status_changed", "inventory_item", item_id, body)
    return row


@app.post("/items/{item_id}/ai/learning-example")
def save_learning_example(item_id: str, body: LearningExampleIn | None = None) -> dict[str, Any]:
    require_item_session_open(item_id)
    item = fetch_one(
        """
        SELECT i.*, oc.name AS object_class_name
        FROM inventory_items i
        LEFT JOIN object_classes oc ON oc.id = i.object_class_id
        WHERE i.id = %s
        """,
        (item_id,),
    )
    if not item:
        raise HTTPException(status_code=404, detail="Gegenstand nicht gefunden")
    ai_row = fetch_one(
        """
        SELECT result_json
        FROM ai_results
        WHERE item_id = %s AND ai_type <> 'deep_dive'
        ORDER BY created_at DESC
        LIMIT 1
        """,
        (item_id,),
    )
    photo_rows = fetch_all(
        "SELECT id FROM item_photos WHERE item_id = %s ORDER BY uploaded_at DESC LIMIT 5",
        (item_id,),
    )
    corrected = {
        "object_type": item.get("object_type"),
        "object_class": item.get("object_class_name"),
        "brand": item.get("brand"),
        "model": item.get("model"),
        "serial_number_present": bool(item.get("serial_number")),
        "serial_number": item.get("serial_number"),
        "specification": item.get("specification"),
        "construction_year": item.get("construction_year"),
        "condition": item.get("condition"),
        "value_estimate": item.get("value_estimate"),
        "estimated_age_years": item.get("estimated_age_years"),
        "age_source": item.get("age_source"),
        "age_verification_status": item.get("age_verification_status"),
        "commercial_category": item.get("commercial_category"),
        "reference_kind": "gepruefte_wertreferenz" if item.get("value_estimate") is not None or item.get("estimated_age_years") is not None else "ki_lernbeispiel",
    }
    note = (body.notes if body else None) or "Vom Prüfer als geprüfte Wertreferenz markiert"
    existing = fetch_one("SELECT id FROM ai_learning_examples WHERE item_id = %s AND approved = true ORDER BY created_at DESC LIMIT 1", (item_id,))
    if existing:
        row = execute(
            """
            UPDATE ai_learning_examples
            SET object_class_id = %s, object_class_name = %s, object_type = %s,
                brand = %s, model = %s, serial_number = %s, condition = %s,
                corrected_json = %s::jsonb, ai_suggestion_json = %s::jsonb,
                photo_ids = %s, notes = %s
            WHERE id = %s
            RETURNING *
            """,
            (
                item.get("object_class_id"),
                item.get("object_class_name"),
                item.get("object_type"),
                item.get("brand"),
                item.get("model"),
                item.get("serial_number"),
                item.get("condition"),
                json_string(corrected),
                json_string(ai_row.get("result_json") if ai_row else {}),
                [photo["id"] for photo in photo_rows],
                note,
                existing["id"],
            ),
        )
    else:
        row = execute(
            """
            INSERT INTO ai_learning_examples (
              item_id, session_id, object_class_id, object_class_name, object_type,
              brand, model, serial_number, condition, corrected_json, ai_suggestion_json,
              photo_ids, notes, approved
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s,%s,true)
            RETURNING *
            """,
            (
                item_id,
                item.get("session_id"),
                item.get("object_class_id"),
                item.get("object_class_name"),
                item.get("object_type"),
                item.get("brand"),
                item.get("model"),
                item.get("serial_number"),
                item.get("condition"),
                json_string(corrected),
                json_string(ai_row.get("result_json") if ai_row else {}),
                [photo["id"] for photo in photo_rows],
                note,
            ),
        )
    audit("ai_learning_example_created", "inventory_item", item_id, {"learning_example_id": str(row["id"]), **corrected})
    return row


@app.post("/items/{item_id}/photos")
async def upload_photo(
    item_id: str,
    photo_type: str = "object",
    client_photo_id: str | None = None,
    source_device_id: str | None = None,
    file: UploadFile = File(...),
) -> dict[str, Any]:
    row, _ = await save_item_photo(
        item_id=item_id,
        file=file,
        photo_type=photo_type,
        client_photo_id=client_photo_id,
        source_device_id=source_device_id,
    )
    return row


@app.get("/offline-sync/status")
def offline_sync_status(
    request: Request,
    session_id: str,
    source_device_id: str,
    client_item_id: str,
    client_photo_ids: str | None = None,
) -> dict[str, Any]:
    auth = getattr(request.state, "auth", {}) or {}
    if auth.get("kind") == "mobile_session" and str(auth.get("session_id") or "") != session_id:
        raise HTTPException(status_code=403, detail="Sync-Status gehört nicht zu dieser mobilen Session")
    session = get_session(session_id)
    expected_photo_ids = [
        value.strip()
        for value in (client_photo_ids or "").split(",")
        if value.strip()
    ]
    item = fetch_one(
        """
        SELECT id, session_id, source_device_id, client_item_id, object_type, sequence_number, status, review_status
        FROM inventory_items
        WHERE session_id = %s AND source_device_id = %s AND client_item_id = %s
        LIMIT 1
        """,
        (session_id, source_device_id, client_item_id),
    )
    if not item:
        return {
            "session_id": session_id,
            "session_status": session.get("status"),
            "source_device_id": source_device_id,
            "client_item_id": client_item_id,
            "server_item_id": None,
            "item_status": "missing",
            "photos": [],
            "known_client_photo_ids": [],
            "missing_client_photo_ids": expected_photo_ids,
        }
    photos = fetch_all(
        """
        SELECT id, item_id, photo_type, client_photo_id, original_path, uploaded_at
        FROM item_photos
        WHERE item_id = %s AND source_device_id = %s
        ORDER BY uploaded_at, id
        """,
        (item["id"], source_device_id),
    )
    all_known_ids = [str(photo["client_photo_id"]) for photo in photos if photo.get("client_photo_id")]
    all_known_set = set(all_known_ids)
    if expected_photo_ids:
        expected_set = set(expected_photo_ids)
        known_ids = [photo_id for photo_id in all_known_ids if photo_id in expected_set]
    else:
        known_ids = all_known_ids
    return {
        "session_id": session_id,
        "session_status": session.get("status"),
        "source_device_id": source_device_id,
        "client_item_id": client_item_id,
        "server_item_id": str(item["id"]),
        "item_status": "synced",
        "server_item": item,
        "photos": photos,
        "known_client_photo_ids": known_ids,
        "missing_client_photo_ids": [photo_id for photo_id in expected_photo_ids if photo_id not in all_known_set],
    }


@app.post("/offline-sync/reconcile")
def offline_sync_reconcile(request: Request, body: OfflineReconcileIn) -> dict[str, Any]:
    auth = getattr(request.state, "auth", {}) or {}
    if auth.get("kind") == "mobile_session" and str(auth.get("session_id") or "") != body.session_id:
        raise HTTPException(status_code=403, detail="Sync-Reconcile gehört nicht zu dieser mobilen Session")

    session = fetch_one(
        "SELECT id, status, inventory_type FROM inventory_sessions WHERE id = %s",
        (body.session_id,),
    )
    if not session:
        return {
            "session_id": body.session_id,
            "source_device_id": body.source_device_id,
            "session_status": "missing",
            "packages": [
                {
                    "client_item_id": package.client_item_id,
                    "status": "discardable",
                    "server_item_id": None,
                    "known_client_photo_ids": [],
                    "missing_client_photo_ids": list(dict.fromkeys(package.client_photo_ids)),
                    "session_status": "missing",
                }
                for package in body.packages
            ],
        }

    packages: list[dict[str, Any]] = []
    for package in body.packages:
        expected_photo_ids = [value for value in dict.fromkeys(package.client_photo_ids) if value]
        if not package.client_item_id.strip() or not body.source_device_id.strip():
            packages.append({
                "client_item_id": package.client_item_id,
                "status": "foreign_or_invalid",
                "server_item_id": None,
                "known_client_photo_ids": [],
                "missing_client_photo_ids": expected_photo_ids,
                "session_status": session.get("status"),
            })
            continue

        item = fetch_one(
            """
            SELECT id, session_id, source_device_id, client_item_id, object_type, sequence_number, status, review_status
            FROM inventory_items
            WHERE session_id = %s AND source_device_id = %s AND client_item_id = %s
            LIMIT 1
            """,
            (body.session_id, body.source_device_id, package.client_item_id),
        )
        if not item:
            packages.append({
                "client_item_id": package.client_item_id,
                "status": "session_closed" if session.get("status") != "open" else "missing",
                "server_item_id": None,
                "known_client_photo_ids": [],
                "missing_client_photo_ids": expected_photo_ids,
                "session_status": session.get("status"),
            })
            continue

        photos = fetch_all(
            """
            SELECT id, item_id, photo_type, client_photo_id, original_path, uploaded_at
            FROM item_photos
            WHERE item_id = %s AND source_device_id = %s
            ORDER BY uploaded_at, id
            """,
            (item["id"], body.source_device_id),
        )
        all_known_ids = [str(photo["client_photo_id"]) for photo in photos if photo.get("client_photo_id")]
        all_known_set = set(all_known_ids)
        known_ids = [photo_id for photo_id in all_known_ids if not expected_photo_ids or photo_id in set(expected_photo_ids)]
        missing_ids = [photo_id for photo_id in expected_photo_ids if photo_id not in all_known_set]
        status = "session_closed" if session.get("status") != "open" else ("missing_photos" if missing_ids else "synced")
        packages.append({
            "client_item_id": package.client_item_id,
            "status": status,
            "server_item_id": str(item["id"]),
            "server_item": item,
            "known_client_photo_ids": known_ids,
            "missing_client_photo_ids": missing_ids,
            "session_status": session.get("status"),
        })

    return {
        "session_id": body.session_id,
        "source_device_id": body.source_device_id,
        "session_status": session.get("status"),
        "packages": packages,
    }


@app.post("/offline-sync/photos")
async def offline_sync_photo(
    session_id: str = Form(...),
    source_device_id: str = Form(...),
    client_item_id: str = Form(...),
    client_photo_id: str = Form(...),
    photo_type: str = Form("object_front"),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    session = get_session(session_id)
    if session["status"] != "open":
        raise HTTPException(status_code=409, detail="Raum ist abgeschlossen")
    item = fetch_one(
        """
        SELECT *
        FROM inventory_items
        WHERE session_id = %s AND source_device_id = %s AND client_item_id = %s
        LIMIT 1
        """,
        (session_id, source_device_id, client_item_id),
    )
    if not item:
        raise HTTPException(status_code=409, detail="Objekt ist noch nicht synchronisiert")
    row, already_exists = await save_item_photo(
        item_id=str(item["id"]),
        file=file,
        photo_type=photo_type,
        client_photo_id=client_photo_id,
        source_device_id=source_device_id,
    )
    return {
        "server_item_id": str(item["id"]),
        "server_photo_id": str(row["id"]),
        "client_item_id": client_item_id,
        "client_photo_id": client_photo_id,
        "photo_type": row.get("photo_type") or photo_type,
        "status": "already_exists" if already_exists else "synced",
    }


async def save_item_photo(
    item_id: str,
    file: UploadFile,
    photo_type: str = "object",
    client_photo_id: str | None = None,
    source_device_id: str | None = None,
) -> tuple[dict[str, Any], bool]:
    ensure_upload_dirs()
    require_item_session_open(item_id)
    if client_photo_id and source_device_id:
        existing = fetch_one(
            """
            SELECT *
            FROM item_photos
            WHERE item_id = %s AND source_device_id = %s AND client_photo_id = %s
            LIMIT 1
            """,
            (item_id, source_device_id, client_photo_id),
        )
        if existing:
            return existing, True
    photo_count = fetch_one("SELECT count(*)::int AS count FROM item_photos WHERE item_id = %s", (item_id,))
    if photo_count and int(photo_count["count"]) >= 5:
        raise HTTPException(status_code=409, detail="Maximal 5 Fotos pro Gegenstand möglich")
    data = await file.read()
    mime_type, suffix = validate_photo_upload(file, data)
    digest = hashlib.sha256(data).hexdigest()
    duplicate = fetch_one(
        "SELECT * FROM item_photos WHERE item_id = %s AND original_hash = %s LIMIT 1",
        (item_id, digest),
    )
    if duplicate:
        return duplicate, True
    filename = f"{item_id}-{photo_type}-{digest[:12]}{suffix}"
    path = Path(settings.upload_root, "originals", filename)
    path.write_bytes(data)
    stamped = Path(settings.upload_root, "stamped", filename)
    stamped.write_bytes(data)
    row = execute(
        """
        INSERT INTO item_photos (
          tenant_id, item_id, photo_type, original_path, stamped_path, original_hash,
          client_photo_id, source_device_id, metadata_json
        )
        VALUES ((SELECT tenant_id FROM inventory_items WHERE id = %s), %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
        RETURNING *
        """,
        (
            item_id,
            item_id,
            photo_type,
            str(path),
            str(stamped),
            digest,
            client_photo_id,
            source_device_id,
            json.dumps({"phase": "1", "stamp": "pending-worker", "mime_type": mime_type, "size": len(data)}),
        ),
    )
    audit("photo_uploaded", "inventory_item", item_id, row)
    run_inventory_rework_check(item_id)
    return row, False


@app.get("/offline-sync/recent")
def offline_sync_recent(request: Request, limit: int = 15) -> list[dict[str, Any]]:
    """Zuletzt erfasste Objekte dieses Geräts in dieser Session.

    Grundlage für die Nacherfassung am Handy: ein gespeichertes Objekt
    wieder öffnen und Fotos/Felder ergänzen. Bewusst auf das eigene
    Gerät begrenzt (source_device_id), damit der Bundle-Upsert beim
    erneuten Speichern dasselbe Objekt trifft und keine Dublette entsteht.
    """
    auth = getattr(request.state, "auth", None) or {}
    if auth.get("kind") != "mobile_session":
        raise HTTPException(status_code=403, detail="Nur für gekoppelte Geräte")
    session_id = str(auth.get("session_id") or "")
    device_id = str(auth.get("device_id") or "")
    if not session_id or not device_id:
        raise HTTPException(status_code=403, detail="Session/Gerät unbekannt")
    limit = max(1, min(int(limit), 30))
    rows = fetch_all(
        """
        SELECT i.id, i.client_item_id, i.sequence_number, i.object_type, i.specification,
               i.serial_number, i.construction_year, i.condition, i.condition_note,
               i.function_ok, i.uvv_status, i.uvv_valid_until, i.inspection_book_available,
               i.remark, i.type_plate_status, i.object_class_id, i.captured_at, i.locked_at,
               COALESCE(p.photo_types, '{}') AS photo_types
        FROM inventory_items i
        LEFT JOIN LATERAL (
          SELECT array_agg(DISTINCT photo_type) AS photo_types
          FROM item_photos WHERE item_id = i.id
        ) p ON true
        WHERE i.session_id = %s AND i.source_device_id = %s
        ORDER BY i.captured_at DESC
        LIMIT %s
        """,
        (session_id, device_id, limit),
    )
    return rows


@app.post("/offline-sync/items")
async def offline_sync_item_bundle(payload: str = Form(...), files: list[UploadFile] = File(default=[])) -> dict[str, Any]:
    try:
        parsed = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Sync-Paket ist kein gültiges JSON") from exc
    item_payload = parsed.get("item") if isinstance(parsed.get("item"), dict) else parsed
    photo_meta = parsed.get("photos") if isinstance(parsed.get("photos"), list) else []
    try:
        body = ItemIn(**item_payload)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Sync-Paket ist unvollständig: {exc}") from exc
    if not body.client_item_id or not body.source_device_id:
        raise HTTPException(status_code=422, detail="client_item_id und source_device_id sind für Offline-Sync erforderlich")
    session = get_session(body.session_id)
    if session["status"] != "open":
        raise HTTPException(status_code=409, detail="Raum ist abgeschlossen")
    session_inventory_type = session.get("inventory_type") or "bga"
    if body.inventory_type and body.inventory_type != session_inventory_type:
        raise HTTPException(status_code=409, detail="Erfassungsart der Session kann nicht gemischt werden")

    existing_before = fetch_one(
        """
        SELECT *
        FROM inventory_items
        WHERE session_id = %s AND source_device_id = %s AND client_item_id = %s
        LIMIT 1
        """,
        (body.session_id, body.source_device_id, body.client_item_id),
    )
    item = existing_before or create_item(body)
    update_data = {
        "object_type": body.object_type,
        "brand": body.brand,
        "model": body.model,
        "serial_number": body.serial_number,
        "condition": body.condition,
        "condition_note": body.condition_note,
        "created_by": body.created_by,
        "specification": body.specification,
        "construction_year": body.construction_year,
        "function_ok": body.function_ok,
        "uvv_status": body.uvv_status,
        "uvv_valid_until": body.uvv_valid_until,
        "inspection_book_available": body.inspection_book_available,
        "remark": body.remark,
        "type_plate_status": body.type_plate_status,
    }
    if body.object_class_id:
        update_data["object_class_id"] = body.object_class_id
    fields = list(update_data.keys())
    item = execute(
        f"""
        UPDATE inventory_items
        SET {", ".join(f"{field} = %s" for field in fields)}, updated_at = now()
        WHERE id = %s
        RETURNING *
        """,
        tuple(update_data[field] for field in fields) + (item["id"],),
    )
    audit("offline_item_synced", "inventory_item", str(item["id"]), {"client_item_id": body.client_item_id, "photos": len(photo_meta)})
    run_inventory_rework_check(str(item["id"]))

    photo_results: list[dict[str, Any]] = []
    for index, meta in enumerate(photo_meta):
        client_photo_id = str(meta.get("client_photo_id") or "") if isinstance(meta, dict) else ""
        photo_type = str(meta.get("photo_type") or "object_front") if isinstance(meta, dict) else "object_front"
        if index >= len(files):
            photo_results.append({
                "client_photo_id": client_photo_id,
                "photo_type": photo_type,
                "status": "failed",
                "error": "Datei fehlt im Sync-Paket",
            })
            continue
        try:
            row, already_exists = await save_item_photo(
                item_id=str(item["id"]),
                file=files[index],
                photo_type=photo_type,
                client_photo_id=client_photo_id or None,
                source_device_id=body.source_device_id,
            )
            photo_results.append({
                "client_photo_id": client_photo_id,
                "photo_type": photo_type,
                "status": "already_exists" if already_exists else "synced",
                "server_photo_id": str(row["id"]),
            })
        except Exception as exc:
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            photo_results.append({
                "client_photo_id": client_photo_id,
                "photo_type": photo_type,
                "status": "failed",
                "error": detail,
            })
    return {
        "server_item_id": str(item["id"]),
        "client_item_id": body.client_item_id,
        "created_or_updated": "updated" if existing_before else "created",
        "photo_results": photo_results,
    }


@app.post("/mobile-diagnostics")
async def mobile_diagnostics(request: Request, body: dict[str, Any]) -> dict[str, Any]:
    ensure_upload_dirs()
    auth = getattr(request.state, "auth", {}) or {}
    body_session_id = str(((body.get("allgemein") or {}) if isinstance(body.get("allgemein"), dict) else {}).get("session_id") or "")
    auth_session_id = str(auth.get("session_id") or "")
    if auth.get("kind") == "mobile_session" and auth_session_id and body_session_id and body_session_id != auth_session_id:
        raise HTTPException(status_code=403, detail="Diagnose gehört nicht zu dieser mobilen Session")
    session_id = body_session_id or auth_session_id or "unknown"
    diagnostic_id = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(4)}"
    path = Path(settings.upload_root, "temp", f"mobile-sync-{diagnostic_id}.json")
    payload = {
        "diagnostic_id": diagnostic_id,
        "received_at": datetime.utcnow().isoformat() + "Z",
        "request_id": getattr(request.state, "request_id", None),
        "auth": {
            "kind": auth.get("kind"),
            "session_id": auth_session_id or None,
            "device_id": auth.get("device_id"),
        },
        "body": body,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    audit(
        "mobile_sync_diagnostic",
        "inventory_session",
        session_id if session_id != "unknown" else None,
        {
            "diagnostic_id": diagnostic_id,
            "path": str(path),
            "queue_summary": body.get("queue_summary") if isinstance(body.get("queue_summary"), dict) else None,
        },
    )
    return {"id": diagnostic_id, "stored": True}


@app.post("/items/{item_id}/audio")
async def upload_audio(
    item_id: str,
    transcript: str | None = None,
    file: UploadFile | None = File(None),
    transcript_form: str | None = Form(None, alias="transcript"),
) -> dict[str, Any]:
    ensure_upload_dirs()
    require_item_session_open(item_id)
    text = transcript_form if transcript_form is not None else transcript
    audio_hash: str | None = None
    path = Path(settings.upload_root, "audio", f"{item_id}-{secrets.token_hex(4)}.txt")
    if file:
        data = await file.read()
        if not data:
            raise HTTPException(status_code=400, detail="Audio-Datei ist leer")
        if len(data) > settings.max_upload_bytes:
            raise HTTPException(status_code=413, detail="Audio-Datei ist zu gross")
        suffix = safe_audio_suffix(file.filename, file.content_type)
        audio_hash = hashlib.sha256(data).hexdigest()
        path = safe_upload_path(str(Path(settings.upload_root, "audio", f"{item_id}-{audio_hash[:16]}{suffix}")))
        path.write_bytes(data)
        # Diktat-Pipeline: Audio ohne Transkript wartet auf den Whisper-Worker.
        transcript_status = "completed" if text else "pending"
    else:
        if not text:
            raise HTTPException(status_code=400, detail="Weder Audiodatei noch Transkript uebergeben")
        audio_hash = hashlib.sha256(text.encode("utf-8")).hexdigest()
        path.write_text(text, encoding="utf-8")
        transcript_status = "completed"
    row = execute(
        """
        INSERT INTO item_audio_notes (item_id, audio_path, transcript, transcript_status, audio_hash)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (item_id, audio_hash) WHERE audio_hash IS NOT NULL DO NOTHING
        RETURNING *
        """,
        (item_id, str(path), text, transcript_status, audio_hash),
    )
    if row is None:
        # Wiederholter Sync derselben Aufnahme (Offline-Queue): keine Dublette.
        return fetch_one(
            "SELECT * FROM item_audio_notes WHERE item_id = %s AND audio_hash = %s",
            (item_id, audio_hash),
        )
    audit("audio_saved", "inventory_item", item_id, row)
    return row


async def upload_audio(item_id: str, transcript: str | None = None, file: UploadFile | None = File(None)) -> dict[str, Any]:
    ensure_upload_dirs()
    require_item_session_open(item_id)
    path = Path(settings.upload_root, "audio", f"{item_id}-{secrets.token_hex(4)}.txt")
    if file:
        data = await file.read()
        if not data:
            raise HTTPException(status_code=400, detail="Audio-Datei ist leer")
        if len(data) > settings.max_upload_bytes:
            raise HTTPException(status_code=413, detail="Audio-Datei ist zu groß")
        suffix = safe_audio_suffix(file.filename, file.content_type)
        path = safe_upload_path(str(Path(settings.upload_root, "audio", f"{item_id}-{secrets.token_hex(16)}{suffix}")))
        path.write_bytes(data)
    else:
        path.write_text(transcript or "", encoding="utf-8")
    row = execute(
        """
        INSERT INTO item_audio_notes (item_id, audio_path, transcript, transcript_status)
        VALUES (%s, %s, %s, 'completed')
        RETURNING *
        """,
        (item_id, str(path), transcript),
    )
    audit("audio_saved", "inventory_item", item_id, row)
    return row


def ai_status_for_mode(mode: str, step: str) -> str:
    normalized = "review" if mode == "review" else "fast"
    matrix = {
        "fast": {
            "queued": "ki_schnell_wartet",
            "running": "ki_schnell_laeuft",
            "done": "ki_schnell_fertig",
            "open": "ki_pruefung_offen",
        },
        "review": {
            "queued": "ki_pruefung_wartet",
            "running": "ki_pruefung_laeuft",
            "done": "ki_pruefung_fertig",
            "open": "ki_pruefung_offen",
        },
    }
    return matrix[normalized][step]


def repair_stale_ai_statuses(session_id: str | None = None, item_id: str | None = None) -> list[dict[str, Any]]:
    conditions = [
        f"status IN ({AI_ACTIVE_STATUS_SQL})",
        f"updated_at < now() - interval '{AI_STALE_INTERVAL_SQL}'",
    ]
    params: list[Any] = []
    if session_id:
        conditions.append("session_id = %s")
        params.append(session_id)
    if item_id:
        conditions.append("id = %s")
        params.append(item_id)
    rows = fetch_all(
        f"""
        SELECT id, object_type, status, updated_at
        FROM inventory_items
        WHERE {" AND ".join(conditions)}
        ORDER BY updated_at ASC
        """,
        tuple(params),
    )
    repaired: list[dict[str, Any]] = []
    for row in rows:
        updated = execute(
            """
            UPDATE inventory_items
            SET status = 'ki_pruefung_offen',
                updated_at = now()
            WHERE id = %s
              AND status <> 'finalisiert'
            RETURNING id, object_type, status, updated_at
            """,
            (row["id"],),
        )
        if updated:
            repaired.append(updated)
            audit("ai_job_repaired", "inventory_item", str(row["id"]), {
                "previous_status": row.get("status"),
                "previous_updated_at": row.get("updated_at"),
                "reason": f"active longer than {AI_STALE_INTERVAL_SQL}",
            })
    return repaired


def current_active_ai_job(item_id: str) -> dict[str, Any] | None:
    repair_stale_ai_statuses(item_id=item_id)
    return fetch_one(
        f"""
        SELECT id, object_type, status, updated_at
        FROM inventory_items
        WHERE id = %s
          AND status IN ({AI_ACTIVE_STATUS_SQL})
          AND updated_at >= now() - interval '{AI_STALE_INTERVAL_SQL}'
        """,
        (item_id,),
    )


def ai_session_generation_for_item(item_id: str) -> int:
    row = fetch_one(
        """
        SELECT COALESCE(s.ai_cancel_generation, 0) AS ai_cancel_generation
        FROM inventory_items i
        JOIN inventory_sessions s ON s.id = i.session_id
        WHERE i.id = %s
        """,
        (item_id,),
    )
    return int((row or {}).get("ai_cancel_generation") or 0)


def ai_job_generation_for_item(item_id: str) -> dict[str, int]:
    row = fetch_one(
        """
        SELECT COALESCE(s.ai_cancel_generation, 0) AS session_generation,
               COALESCE(i.ai_cancel_generation, 0) AS item_generation
        FROM inventory_items i
        JOIN inventory_sessions s ON s.id = i.session_id
        WHERE i.id = %s
        """,
        (item_id,),
    )
    return {
        "session": int((row or {}).get("session_generation") or 0),
        "item": int((row or {}).get("item_generation") or 0),
    }


def expected_session_generation(expected_generation: Any) -> int | None:
    if expected_generation is None:
        return None
    if isinstance(expected_generation, dict):
        return int(expected_generation.get("session") or 0)
    return int(expected_generation)


def expected_item_generation(expected_generation: Any) -> int | None:
    if isinstance(expected_generation, dict):
        return int(expected_generation.get("item") or 0)
    return None


def ai_session_generation(session_id: str) -> int:
    row = fetch_one(
        "SELECT COALESCE(ai_cancel_generation, 0) AS ai_cancel_generation FROM inventory_sessions WHERE id = %s",
        (session_id,),
    )
    return int((row or {}).get("ai_cancel_generation") or 0)


def ai_job_cancelled(item_id: str, expected_generation: Any = None) -> bool:
    expected_session = expected_session_generation(expected_generation)
    if expected_session is None:
        return False
    current = ai_job_generation_for_item(item_id)
    if current["session"] != expected_session:
        return True
    expected_item = expected_item_generation(expected_generation)
    return expected_item is not None and current["item"] != expected_item


def latest_ai_context(item_id: str) -> str:
    row = fetch_one(
        """
        SELECT result_json
        FROM ai_results
        WHERE item_id = %s AND ai_type <> 'deep_dive'
        ORDER BY created_at DESC
        LIMIT 1
        """,
        (item_id,),
    )
    if not row:
        return ""
    result = row.get("result_json") or {}
    parts = [
        result.get("notes"),
        result.get("brand"),
        result.get("model"),
        result.get("object_type"),
        " ".join(str(value) for value in result.get("missing_fields") or []),
    ]
    return " ".join(str(part) for part in parts if part)


GENERIC_DEEP_DIVE_TERMS = {
    "betriebs",
    "geschaeftsausstattung",
    "geschäftsausstattung",
    "betriebsausstattung",
    "inventar",
    "inventur",
    "anlagevermoegen",
    "anlagevermögen",
    "gwg",
    "buchhaltung",
    "abschreibung",
}

GENERIC_SOURCE_DOMAINS = (
    "haufe.de",
    "lexware.de",
    "bwl-lexikon.de",
    "buchhaltungsbutler.de",
    "dasfinanzen.de",
    "rechnungswesenforum.de",
)

SHOP_SOURCE_HINTS = (
    "ebay.",
    "kleinanzeigen.",
    "amazon.",
    "idealo.",
    "geizhals.",
    "backmarket.",
    "rebuy.",
    "asgoodasnew.",
    "refurbed.",
    "picclick.",
    "pccomponentes.",
    "notebooksbilliger.",
)


def normalize_deep_dive_text(value: Any) -> str:
    text = str(value or "").strip()
    text = (
        text.replace("Ã¤", "ä")
        .replace("Ã¶", "ö")
        .replace("Ã¼", "ü")
        .replace("ÃŸ", "ß")
        .replace("â‚¬", "€")
    )
    return re.sub(r"\s+", " ", text).strip()


def search_tokens(value: str) -> list[str]:
    text = normalize_deep_dive_text(value).lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    tokens = re.findall(r"[a-z0-9]{3,}", text)
    return [token for token in tokens if token not in GENERIC_DEEP_DIVE_TERMS]


def deep_dive_product_terms(item: dict[str, Any], ai_context: str = "") -> list[str]:
    raw_parts = [
        item.get("brand"),
        item.get("model"),
        item.get("object_type"),
        item.get("specification"),
        item.get("object_class_name"),
        item.get("serial_number"),
    ]
    context = normalize_deep_dive_text(ai_context)
    # Nur kurze, produktnahe Kontext-Hinweise übernehmen. Lange KI-Sätze
    # verwässern SearXNG sonst mit Steuer-/BGA-Artikeln statt Produktquellen.
    context_terms = []
    for term in re.findall(r"\b[A-ZÄÖÜa-zäöüß0-9][A-ZÄÖÜa-zäöüß0-9+./-]{2,}\b", context):
        if len(term) <= 28 and term.lower() not in GENERIC_DEEP_DIVE_TERMS:
            context_terms.append(term)
        if len(context_terms) >= 4:
            break
    parts: list[str] = []
    seen: set[str] = set()
    for part in raw_parts + context_terms:
        text = normalize_deep_dive_text(part)
        if not text:
            continue
        key = text.lower()
        if key in seen or key in GENERIC_DEEP_DIVE_TERMS:
            continue
        seen.add(key)
        parts.append(text)
    return parts or ["Betriebs- und Geschäftsausstattung"]


def deep_dive_research_basis(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "designation": normalize_deep_dive_text(item.get("object_type")),
        "specification": normalize_deep_dive_text(item.get("specification")),
        "brand": normalize_deep_dive_text(item.get("brand")),
        "model": normalize_deep_dive_text(item.get("model")),
        "serial_number": normalize_deep_dive_text(item.get("serial_number")),
        "construction_year": normalize_deep_dive_text(item.get("construction_year")),
        "condition": normalize_deep_dive_text(item.get("condition")),
        "object_class": normalize_deep_dive_text(item.get("object_class_name")),
    }


def deep_dive_queries(item: dict[str, Any], ai_context: str = "") -> list[str]:
    terms = deep_dive_product_terms(item, ai_context)
    product = " ".join(terms[:5]).strip()
    object_type = normalize_deep_dive_text(item.get("object_type") or item.get("object_class_name") or "")
    brand_model = " ".join(term for term in [item.get("brand"), item.get("model")] if term).strip()
    specification = normalize_deep_dive_text(item.get("specification"))
    construction_year = normalize_deep_dive_text(item.get("construction_year"))
    if brand_model and object_type and object_type.lower() not in brand_model.lower():
        base = f"{brand_model} {object_type}"
    else:
        base = brand_model or product or object_type
    exact_base = base
    if specification and specification.lower() not in exact_base.lower():
        exact_base = f"{exact_base} {specification[:80]}".strip()
    queries = []
    if construction_year:
        queries.append(f"{exact_base} {construction_year} gebraucht Preis")
    queries.extend([
        f"{exact_base} gebraucht Preis",
        f"{base} Gebrauchtpreis",
        f"{base} Datenblatt Erscheinungsjahr",
    ])
    if object_type and object_type.lower() not in base.lower():
        queries.append(f"{base} {object_type} Preis gebraucht")
    compact: list[str] = []
    seen: set[str] = set()
    for query in queries:
        query = normalize_deep_dive_text(query)
        key = query.lower()
        if query and key not in seen:
            seen.add(key)
            compact.append(query)
    return compact[:4]


def deep_dive_query(item: dict[str, Any], ai_context: str = "") -> str:
    return deep_dive_queries(item, ai_context)[0]


def source_host(url: str) -> str:
    try:
        return (urlparse(url).netloc or "").lower()
    except Exception:
        return ""


def is_relevant_source(source: dict[str, Any], item: dict[str, Any], ai_context: str = "") -> bool:
    text = " ".join(
        normalize_deep_dive_text(source.get(key))
        for key in ["title", "snippet", "url"]
    ).lower()
    host = source_host(str(source.get("url") or ""))
    product_terms = [
        token for token in search_tokens(" ".join(deep_dive_product_terms(item, ai_context)))
        if token not in {"modell", "model", "objekt", "geraet", "gerät"}
    ]
    if any(domain in host for domain in GENERIC_SOURCE_DOMAINS) and not any(term in text for term in product_terms[:4]):
        return False
    if not product_terms:
        return True
    strong_terms = [term for term in product_terms if any(ch.isdigit() for ch in term) or len(term) >= 5]
    matches = sum(1 for term in strong_terms[:8] if term in text)
    required_matches = 2 if len(strong_terms) >= 2 else 1
    if matches >= required_matches:
        return True
    category_terms = search_tokens(f"{item.get('object_type') or ''} {item.get('object_class_name') or ''}")
    if (
        any(hint in host for hint in SHOP_SOURCE_HINTS)
        and any(term in text for term in product_terms[:8])
        and (not category_terms or any(term in text for term in category_terms[:5]))
    ):
        return True
    return False


def collect_search_sources(item: dict[str, Any], ai_context: str = "", limit: int = 6) -> tuple[list[dict[str, Any]], str, str | None, list[str]]:
    collected: list[dict[str, Any]] = []
    errors: list[str] = []
    provider_used = settings.search_provider or "searxng"
    seen: set[str] = set()
    queries = deep_dive_queries(item, ai_context)
    for query in queries:
        sources, provider, error = search_sources(query, limit=limit)
        provider_used = provider
        if error:
            errors.append(error)
        for source in sources:
            url = str(source.get("url") or "")
            if not url or url in seen:
                continue
            if not is_relevant_source(source, item, ai_context):
                continue
            seen.add(url)
            source["query"] = query
            source["rank"] = len(collected) + 1
            collected.append(source)
            if len(collected) >= limit:
                return collected, provider_used, "; ".join(dict.fromkeys(errors)) or None, queries
    return collected, provider_used, "; ".join(dict.fromkeys(errors)) or None, queries


def classify_deep_dive_source(source: dict[str, Any]) -> str:
    host = source_host(str(source.get("url") or ""))
    text = " ".join(str(source.get(key) or "") for key in ["title", "snippet", "url"]).lower()
    if any(hint in host for hint in ("ebay.", "kleinanzeigen.", "willhaben.", "ricardo.", "rebuy.", "backmarket.", "refurbed.")):
        return "gebrauchtmarkt"
    if any(hint in host for hint in ("amazon.", "idealo.", "geizhals.", "notebooksbilliger.", "conrad.", "reichelt.")):
        return "haendler"
    if any(hint in text for hint in ("datenblatt", "data sheet", "datasheet", "manual", "bedienungsanleitung", ".pdf")):
        return "datenblatt"
    if any(hint in host for hint in ("forum", "reddit", "motor-talk", "gutefrage")):
        return "forum"
    if any(hint in host for hint in ("bosch", "siemens", "hp.", "dell.", "lenovo", "apple.", "nussbaum", "maha", "beissbarth")):
        return "hersteller"
    return "generisch"


def deep_dive_source_evidence(item: dict[str, Any], ai_context: str, sources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    product_terms = [
        term
        for term in search_tokens(" ".join(deep_dive_product_terms(item, ai_context)))
        if len(term) >= 3 and term not in {"modell", "model", "objekt", "geraet", "gerät"}
    ]
    evidence: list[dict[str, Any]] = []
    for source in sources:
        text = " ".join(normalize_deep_dive_text(source.get(key)).lower() for key in ["title", "snippet", "url"])
        matched_terms = [term for term in product_terms[:10] if term in text]
        kind = classify_deep_dive_source(source)
        score = min(
            1.0,
            (0.16 * len(matched_terms))
            + (0.25 if kind in {"hersteller", "datenblatt"} else 0)
            + (0.18 if kind in {"gebrauchtmarkt", "haendler"} else 0)
            + (0.12 if source.get("rank") in {1, 2} else 0),
        )
        evidence.append(
            {
                "title": source.get("title"),
                "url": source.get("url"),
                "host": source_host(str(source.get("url") or "")),
                "kind": kind,
                "query": source.get("query"),
                "rank": source.get("rank"),
                "snippet": source.get("snippet"),
                "matched_terms": matched_terms,
                "relevance_score": round(score, 2),
                "review_required": True,
            }
        )
    return evidence


def deep_dive_dossier(
    item: dict[str, Any],
    ai_context: str,
    sources: list[dict[str, Any]],
    estimated_value: Any = None,
    estimated_value_range: dict[str, Any] | None = None,
    estimated_value_confidence: float | None = None,
    value_source: str | None = None,
    price_candidates: list[dict[str, Any]] | None = None,
    valuation_state: str | None = None,
    reference_price_available: bool = False,
    reference_price_label: str | None = None,
    estimated_age_years: Any = None,
    age_candidates: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    basis = deep_dive_research_basis(item)
    confidence = 0.35
    if basis.get("model"):
        confidence += 0.2
    if basis.get("brand"):
        confidence += 0.12
    if basis.get("serial_number"):
        confidence += 0.08
    if sources:
        confidence += min(0.25, 0.06 * len(sources))
    return {
        "identified_product": {
            "designation": basis.get("designation") or basis.get("object_class"),
            "manufacturer": basis.get("brand"),
            "model": basis.get("model"),
            "serial_number": basis.get("serial_number"),
            "construction_year": basis.get("construction_year"),
            "specification": basis.get("specification"),
            "confidence": round(min(confidence, 0.92), 2),
            "review_required": True,
        },
        "technical_facts": {
            "object_class": basis.get("object_class"),
            "condition": basis.get("condition"),
            "specification": basis.get("specification"),
            "typeplate_context": ai_context,
            "review_required": True,
        },
        "source_evidence": deep_dive_source_evidence(item, ai_context, sources),
        "price_candidates": price_candidates or [],
        "age_candidates": age_candidates or [],
        "suggested_value": {
            "amount": estimated_value,
            "range": estimated_value_range or {"min": None, "max": None},
            "source": value_source or "offen",
            "confidence": estimated_value_confidence or 0.0,
            "review_required": True,
            "reference_price_available": reference_price_available,
            "reference_price_label": reference_price_label or "Kein Referenzpreis verfügbar",
        },
        "valuation_state": valuation_state or ("reference_available" if reference_price_available else "no_reference"),
        "reference_price_available": reference_price_available,
        "reference_price_label": reference_price_label or "Kein Referenzpreis verfügbar",
        "suggested_age_years": estimated_age_years,
        "confidence": round(min(confidence, 0.92), 2),
        "review_required": True,
    }


def decode_search_url(value: str) -> str:
    parsed = urlparse(value)
    query = parse_qs(parsed.query)
    if "uddg" in query and query["uddg"]:
        return unquote(query["uddg"][0])
    return value


def normalize_search_source(
    title: str | None,
    url: str | None,
    snippet: str | None,
    provider: str,
    rank: int,
) -> dict[str, Any] | None:
    clean_title = html.unescape(re.sub(r"\s+", " ", str(title or ""))).strip()
    clean_url = str(url or "").strip()
    clean_snippet = html.unescape(re.sub(r"\s+", " ", str(snippet or ""))).strip()
    if not clean_title or not clean_url:
        return None
    return {
        "title": clean_title[:220],
        "url": clean_url,
        "snippet": clean_snippet[:500],
        "source_provider": provider,
        "rank": rank,
    }


def parse_searxng_html_sources(text: str, limit: int) -> list[dict[str, Any]]:
    sources: list[dict[str, Any]] = []
    seen: set[str] = set()
    for article in re.finditer(
        r'<article\b[^>]*class="[^"]*\bresult\b[^"]*"[^>]*>(.*?)</article>',
        text or "",
        re.IGNORECASE | re.DOTALL,
    ):
        body = article.group(1)
        link_match = re.search(r"<h3>\s*<a[^>]+href=\"([^\"]+)\"[^>]*>(.*?)</a>\s*</h3>", body, re.IGNORECASE | re.DOTALL)
        if not link_match:
            link_match = re.search(r'<a[^>]+class="url_header"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', body, re.IGNORECASE | re.DOTALL)
        if not link_match:
            continue
        raw_url, raw_title = link_match.groups()
        url = html.unescape(raw_url).strip()
        if not url or url in seen:
            continue
        title = re.sub(r"<[^>]+>", "", raw_title)
        snippet = ""
        snippet_match = re.search(r'<p[^>]+class="content"[^>]*>(.*?)</p>', body, re.IGNORECASE | re.DOTALL)
        if snippet_match:
            snippet = re.sub(r"<[^>]+>", "", snippet_match.group(1))
        source = normalize_search_source(title, url, snippet, "searxng", len(sources) + 1)
        if not source:
            continue
        source["source_format"] = "html"
        seen.add(url)
        sources.append(source)
        if len(sources) >= limit:
            break
    return sources


def search_sources_searxng(
    query: str,
    limit: int = 5,
    language: str = "de",
    country: str = "DE",
) -> tuple[list[dict[str, Any]], str | None]:
    base_url = (settings.searxng_base_url or "").rstrip("/")
    if not base_url:
        return [], "SEARXNG_BASE_URL ist nicht gesetzt."
    errors: list[str] = []
    try:
        response = httpx.get(
            f"{base_url}/search",
            params={
                "q": query,
                "format": "json",
                "language": language,
                "categories": "general",
                "safesearch": "0",
            },
            headers={"User-Agent": "Mozilla/5.0 Inventar-App-Raumtest/0.1"},
            timeout=settings.search_timeout_seconds,
        )
        if response.status_code == 403:
            errors.append("JSON-Format nicht freigeschaltet (HTTP 403).")
            payload = {"results": []}
        else:
            response.raise_for_status()
            payload = response.json()
    except Exception as exc:
        errors.append(f"JSON: {type(exc).__name__}: {str(exc)[:220]}")
        payload = {"results": []}

    sources: list[dict[str, Any]] = []
    seen: set[str] = set()
    for result in payload.get("results") or []:
        url = str(result.get("url") or "").strip()
        if not url or url in seen:
            continue
        source = normalize_search_source(
            result.get("title"),
            url,
            result.get("content") or result.get("snippet"),
            "searxng",
            len(sources) + 1,
        )
        if not source:
            continue
        seen.add(url)
        sources.append(source)
        if len(sources) >= limit:
            break
    if sources:
        return sources, "; ".join(errors) or None

    try:
        html_response = httpx.get(
            f"{base_url}/search",
            params={
                "q": query,
                "language": language,
                "categories": "general",
                "safesearch": "0",
            },
            headers={"User-Agent": "Mozilla/5.0 Inventar-App-Raumtest/0.1"},
            timeout=settings.search_timeout_seconds,
        )
        html_response.raise_for_status()
        sources = parse_searxng_html_sources(html_response.text, limit)
    except Exception as exc:
        errors.append(f"HTML: {type(exc).__name__}: {str(exc)[:220]}")
    if not sources:
        errors.append("SearXNG hat keine verwertbaren Quellen geliefert.")
        return [], "; ".join(errors)
    # JSON ist bei manchen SearXNG-Instanzen deaktiviert. Wenn HTML-Treffer
    # sauber gelesen wurden, gilt die Websuche trotzdem als erfolgreich.
    non_json_errors = [error for error in errors if "JSON-Format" not in error]
    return sources, "; ".join(non_json_errors) or None


def search_sources_duckduckgo_fallback(
    query: str,
    limit: int = 5,
    language: str = "de",
    country: str = "DE",
) -> tuple[list[dict[str, Any]], str | None]:
    try:
        response = httpx.get(
            "https://duckduckgo.com/html/",
            params={"q": query, "kl": f"{language.lower()}-{country.lower()}"},
            headers={"User-Agent": "Mozilla/5.0 Inventar-App-Raumtest/0.1"},
            timeout=settings.search_timeout_seconds,
            follow_redirects=True,
        )
        response.raise_for_status()
    except Exception as exc:
        return [], f"{type(exc).__name__}: {str(exc)[:220]}"

    sources: list[dict[str, Any]] = []
    seen: set[str] = set()
    for match in re.finditer(
        r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>',
        response.text,
        re.IGNORECASE | re.DOTALL,
    ):
        raw_url, raw_title = match.groups()
        url = decode_search_url(html.unescape(raw_url))
        title = re.sub(r"<[^>]+>", "", raw_title)
        if not url or url in seen:
            continue
        seen.add(url)
        snippet = ""
        tail = response.text[match.end() : match.end() + 1800]
        snippet_match = re.search(
            r'<a[^>]+class="result__snippet"[^>]*>(.*?)</a>|<div[^>]+class="result__snippet"[^>]*>(.*?)</div>',
            tail,
            re.IGNORECASE | re.DOTALL,
        )
        if snippet_match:
            snippet = re.sub(r"<[^>]+>", "", snippet_match.group(1) or snippet_match.group(2) or "")
        source = normalize_search_source(title, url, snippet, "duckduckgo_html", len(sources) + 1)
        if not source:
            continue
        sources.append(source)
        if len(sources) >= limit:
            break
    if not sources:
        return [], "DuckDuckGo-Fallback hat keine verwertbaren Quellen geliefert."
    return sources, None


def search_sources(
    query: str,
    limit: int = 5,
    language: str = "de",
    country: str = "DE",
) -> tuple[list[dict[str, Any]], str, str | None]:
    preferred = (settings.search_provider or "searxng").strip().lower()
    errors: list[str] = []

    if preferred == "searxng":
        sources, error = search_sources_searxng(query, limit, language, country)
        if sources:
            return sources, "searxng", error
        if error:
            errors.append(f"SearXNG: {error}")
    elif preferred in {"duckduckgo", "duckduckgo_html"}:
        sources, error = search_sources_duckduckgo_fallback(query, limit, language, country)
        if sources:
            return sources, "duckduckgo_html", error
        if error:
            errors.append(f"DuckDuckGo: {error}")
    else:
        errors.append(f"Unbekannter SEARCH_PROVIDER '{preferred}'.")

    if preferred not in {"duckduckgo", "duckduckgo_html"}:
        fallback_sources, fallback_error = search_sources_duckduckgo_fallback(query, limit, language, country)
        if fallback_sources:
            error_text = "; ".join(errors) if errors else None
            return fallback_sources, "duckduckgo_html", error_text
        if fallback_error:
            errors.append(f"DuckDuckGo: {fallback_error}")

    return [], preferred or "searxng", "; ".join(errors) or "Keine Webquellen gefunden."


def estimate_value_range(item: dict[str, Any], ai_context: str = "") -> tuple[int, int]:
    text = " ".join(
        str(item.get(key) or "").lower()
        for key in ["object_type", "object_class_name", "brand", "model"]
    ) + " " + ai_context.lower()
    ranges = [
        (("hebebuehne", "hebebühne", "lift"), (1500, 12000)),
        (("wuchtmaschine", "reifenmontiermaschine"), (800, 8000)),
        (("kompressor",), (300, 4000)),
        (("werkzeugwagen",), (100, 1500)),
        (("rtx 4070", "14700hx"), (1600, 2600)),
        (("rtx 4060", "13700hx", "13620h"), (1100, 2100)),
        (("rtx 3060", "rtx 3070", "12700h"), (700, 1600)),
        (("omen", "rog", "legion", "alienware", "gaming laptop"), (800, 1800)),
        (("notebook", "laptop", "thinkpad"), (120, 900)),
        (("iphone", "smartphone"), (300, 1200)),
        (("monitor", "display"), (40, 300)),
        (("nespresso", "kaffeemaschine", "kapselmaschine", "espresso"), (10, 90)),
        (("maus", "mouse", "tastatur", "keyboard"), (10, 80)),
        (("telefon", "dect"), (20, 300)),
        (("reifen", "radsatz"), (80, 700)),
    ]
    base_min, base_max = 50, 500
    for needles, price_range in ranges:
        if any(needle in text for needle in needles):
            base_min, base_max = price_range
            break

    multiplier = {
        "neu": 1.0,
        "sehr_gut": 0.8,
        "gut": 0.65,
        "gebraucht": 0.45,
        "reparaturbeduerftig": 0.2,
        "defekt": 0.08,
        "aussondern": 0.03,
    }.get(str(item.get("condition") or "gebraucht"), 0.45)
    return max(1, round(base_min * multiplier)), max(1, round(base_max * multiplier))


def parse_euro_amount(raw: str) -> float | None:
    text = raw.strip().replace("\u00a0", " ")
    text = text.replace(".", "").replace(",", ".") if "," in text else text.replace(",", "")
    try:
        value = float(text)
    except ValueError:
        return None
    if value < 1 or value > 50000:
        return None
    return value


def extract_price_candidates(text: str) -> list[float]:
    normalized = normalize_deep_dive_text(text)
    patterns = [
        r"(?:€|eur|euro)\s*(\d{1,5}(?:[.,]\d{2})?)",
        r"(\d{1,5}(?:[.,]\d{2})?)\s*(?:€|eur|euro)",
    ]
    values: list[float] = []
    for pattern in patterns:
        for match in re.finditer(pattern, normalized, flags=re.IGNORECASE):
            value = parse_euro_amount(match.group(1))
            if value is not None:
                values.append(value)
    return values


USED_MARKET_TERMS = (
    "gebraucht",
    "used",
    "second hand",
    "second-hand",
    "kleinanzeigen",
    "ebay",
    "willhaben",
    "refurbished",
    "generalüberholt",
    "generaluberholt",
)

NEW_PRICE_TERMS = (
    "neu",
    "neupreis",
    "uvp",
    "shop",
    "warenkorb",
    "lieferzeit",
    "zzgl",
)


def price_market_kind(source: dict[str, Any], text: str) -> str:
    haystack = f"{text} {source_host(str(source.get('url') or ''))}".lower()
    if any(term in haystack for term in USED_MARKET_TERMS):
        return "gebraucht"
    if any(term in haystack for term in NEW_PRICE_TERMS):
        return "neupreis"
    return "unbekannt"


def unique_tokens(tokens: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        normalized = token.strip().lower()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(normalized)
    return result


def source_match_text(source: dict[str, Any]) -> str:
    return normalize_deep_dive_text(
        " ".join(
            str(source.get(key) or "")
            for key in ["title", "snippet", "url", "excerpt"]
        )
    ).lower()


def is_machine_like_item(item: dict[str, Any], ai_context: str = "") -> bool:
    text = normalize_deep_dive_text(
        " ".join(str(item.get(key) or "") for key in ["object_type", "object_class_name", "specification", "brand", "model"])
        + " "
        + ai_context
    ).lower()
    return any(
        term in text
        for term in [
            "hebeb",
            "lift",
            "montiermaschine",
            "reifenmontiermaschine",
            "wuchtmaschine",
            "kompressor",
            "maschine",
            "werkstatt",
            "buehne",
            "bÃ¼hne",
        ]
    )


def source_reference_match(source: dict[str, Any], item: dict[str, Any], ai_context: str = "") -> dict[str, Any]:
    text = source_match_text(source)
    brand_tokens = search_tokens(str(item.get("brand") or ""))
    model_tokens = search_tokens(" ".join(str(item.get(key) or "") for key in ["model", "serial_number"]))
    spec_tokens = [
        token
        for token in search_tokens(str(item.get("specification") or "") + " " + ai_context)
        if any(ch.isdigit() for ch in token) or len(token) >= 5
    ]
    object_tokens = search_tokens(" ".join(str(item.get(key) or "") for key in ["object_type", "object_class_name"]))
    exact_terms = unique_tokens(model_tokens + brand_tokens + spec_tokens[:6] + object_tokens[:4])
    matched_terms = [token for token in exact_terms if token in text]
    missing_terms = [token for token in exact_terms if token not in text]
    source_kind = classify_deep_dive_source(source)
    market_kind = price_market_kind(source, text)
    used_market = source_kind == "gebrauchtmarkt" or market_kind == "gebraucht"
    brand_match = not brand_tokens or any(token in text for token in brand_tokens)
    model_match_count = sum(1 for token in model_tokens if token in text)
    spec_match_count = sum(1 for token in spec_tokens if token in text)
    object_match = any(token in text for token in object_tokens)
    machine_like = is_machine_like_item(item, ai_context)
    has_strong_model = bool(model_tokens)
    exact = False
    reason = "Kein belastbarer Produktabgleich."
    if used_market and has_strong_model and model_match_count >= max(1, min(len(model_tokens), 2)) and brand_match:
        exact = True
        reason = "Eindeutig: Modell/Typ und Hersteller passen zur Gebrauchtquelle."
    elif used_market and has_strong_model and model_match_count == len(model_tokens) and not brand_tokens:
        exact = True
        reason = "Eindeutig: Modell/Typ passt zur Gebrauchtquelle."
    elif used_market and machine_like and has_strong_model and object_match and model_match_count >= 1 and spec_match_count >= 1:
        exact = True
        reason = "Eindeutig genug: Maschinenklasse, Typenschild-/Modellhinweis und technische Daten passen."
    if exact:
        return {
            "reference_match": "exact",
            "reference_status": "referenzpreis_verfuegbar",
            "match_score": min(1.0, 0.55 + 0.12 * len(matched_terms)),
            "match_reason": reason,
            "matched_terms": matched_terms,
            "missing_terms": missing_terms[:6],
            "source_kind": source_kind,
        }
    if used_market and (object_match or model_match_count or spec_match_count) and not (machine_like and not has_strong_model):
        return {
            "reference_match": "similar",
            "reference_status": "preisspanne_pruefen",
            "match_score": min(0.74, 0.28 + 0.1 * len(matched_terms)),
            "match_reason": "Ähnlicher Gebrauchtmarkt-Treffer, aber Hersteller/Modell/Typ sind nicht eindeutig genug.",
            "matched_terms": matched_terms,
            "missing_terms": missing_terms[:6],
            "source_kind": source_kind,
        }
    if used_market and object_match:
        return {
            "reference_match": "similar",
            "reference_status": "preisspanne_pruefen",
            "match_score": min(0.62, 0.22 + 0.08 * len(matched_terms)),
            "match_reason": "Nur die Objektart passt. Kein Referenzpreis ohne passenden Modell-/Typnachweis.",
            "matched_terms": matched_terms,
            "missing_terms": missing_terms[:6],
            "source_kind": source_kind,
        }
    return {
        "reference_match": "weak",
        "reference_status": "keine_referenz",
        "match_score": min(0.35, 0.08 * len(matched_terms)),
        "match_reason": reason,
        "matched_terms": matched_terms,
        "missing_terms": missing_terms[:6],
        "source_kind": source_kind,
    }


def fetch_source_excerpt(url: str) -> str:
    try:
        response = httpx.get(
            url,
            headers={"User-Agent": "Mozilla/5.0 Inventar-App-Raumtest/0.1"},
            timeout=5.0,
            follow_redirects=True,
        )
        content_type = response.headers.get("content-type", "").lower()
        if response.status_code >= 400 or ("text/html" not in content_type and "text/plain" not in content_type):
            return ""
        text = response.text[:240_000]
        meta_parts = re.findall(
            r'<meta[^>]+(?:property|name)=["\'](?:og:price:amount|product:price:amount|twitter:data1|description)["\'][^>]+content=["\']([^"\']+)["\']',
            text,
            flags=re.IGNORECASE,
        )
        clean_text = re.sub(r"<script\b.*?</script>|<style\b.*?</style>", " ", text, flags=re.IGNORECASE | re.DOTALL)
        clean_text = re.sub(r"<[^>]+>", " ", clean_text)
        return " ".join(meta_parts + [clean_text[:40_000]])
    except Exception:
        return ""


def source_price_candidates(sources: list[dict[str, Any]], item: dict[str, Any], ai_context: str = "") -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for source in sources[:5]:
        text = " ".join([str(source.get("title") or ""), str(source.get("snippet") or "")])
        market_kind = price_market_kind(source, text)
        prices = extract_price_candidates(text)
        source_for_match = source
        if not prices and any(hint in source_host(str(source.get("url") or "")) for hint in SHOP_SOURCE_HINTS):
            excerpt = fetch_source_excerpt(str(source.get("url") or ""))
            if excerpt:
                market_kind = price_market_kind(source, f"{text} {excerpt[:2000]}")
                source_for_match = {**source, "excerpt": excerpt[:2000]}
            prices = extract_price_candidates(excerpt)
        match = source_reference_match(source_for_match, item, ai_context)
        for price in prices[:4]:
            candidates.append({
                "value": price,
                "source": source.get("url"),
                "title": source.get("title"),
                "market_kind": market_kind,
                **match,
            })
    object_class = normalize_bga_object_class(item.get("object_class_name"), item.get("object_type"))
    if object_class in {"Computermaus", "Tastatur", "Kaffeemaschine"}:
        candidates = [candidate for candidate in candidates if candidate.get("market_kind") == "gebraucht"]
    limit = plausible_value_limit(object_class)
    if limit:
        candidates = [candidate for candidate in candidates if float(candidate["value"]) <= max(limit * 1.5, limit + 40)]
    baseline_min, _ = estimate_value_range(item, ai_context)
    class_floor = {
        "Computermaus": 10,
        "Tastatur": 15,
        "Monitor": 25,
        "Telefon": 80,
        "Drucker": 25,
        "Scanner": 25,
        "Kaffeemaschine": 10,
    }.get(object_class, 1)
    if any(term in (" ".join(str(item.get(key) or "").lower() for key in ["object_type", "brand", "model"]) + " " + ai_context.lower()) for term in ["iphone", "smartphone"]):
        class_floor = 150
    low_outlier_floor = max(class_floor, baseline_min * 0.45)
    candidates = [candidate for candidate in candidates if float(candidate["value"]) >= low_outlier_floor]
    return candidates[:12]


def numeric_or_none(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def select_value_references(item: dict[str, Any], ai_context: str = "", limit: int = 3) -> list[dict[str, Any]]:
    query_text = " ".join(
        normalize_deep_dive_text(item.get(key))
        for key in ["object_type", "object_class_name", "brand", "model", "serial_number", "specification", "construction_year", "condition"]
    )
    query_tokens = set(tokenize_reference_text(f"{query_text} {ai_context}"))
    rows = fetch_all(
        """
        SELECT id, item_id, object_class_name, object_type, brand, model, serial_number,
               condition, corrected_json, notes, created_at
        FROM ai_learning_examples
        WHERE approved = true
        ORDER BY created_at DESC
        LIMIT 300
        """
    )
    scored: list[tuple[int, dict[str, Any]]] = []
    current = {key: normalize_deep_dive_text(item.get(key)).lower() for key in ["object_type", "object_class_name", "brand", "model", "specification", "construction_year", "condition"]}
    for row in rows:
        corrected = row.get("corrected_json") or {}
        value_estimate = numeric_or_none(corrected.get("value_estimate"))
        estimated_age = numeric_or_none(corrected.get("estimated_age_years"))
        if value_estimate is None and estimated_age is None:
            continue
        reference_text = " ".join(
            normalize_deep_dive_text(value)
            for value in [
                row.get("object_class_name"),
                row.get("object_type"),
                row.get("brand"),
                row.get("model"),
                row.get("serial_number"),
                row.get("condition"),
                corrected.get("specification"),
                corrected.get("construction_year"),
                corrected.get("object_type"),
                corrected.get("brand"),
                corrected.get("model"),
            ]
        )
        reference_tokens = set(tokenize_reference_text(reference_text))
        score = len(query_tokens & reference_tokens)
        reasons: list[str] = []

        def add_exact(field: str, label: str, weight: int) -> None:
            nonlocal score
            current_value = current.get(field) or ""
            reference_value = normalize_deep_dive_text(corrected.get(field) or row.get(field)).lower()
            if current_value and reference_value and current_value == reference_value:
                score += weight
                reasons.append(label)

        add_exact("object_type", "gleiche Bezeichnung", 5)
        add_exact("brand", "gleiche Marke", 5)
        add_exact("model", "gleiches Modell", 7)
        add_exact("construction_year", "gleiches Baujahr", 3)
        add_exact("condition", "gleicher Zustand", 2)
        if current.get("object_class_name") and normalize_deep_dive_text(row.get("object_class_name")).lower() == current["object_class_name"]:
            score += 2
            reasons.append("gleiche Klasse")
        specification_overlap = query_tokens & set(tokenize_reference_text(str(corrected.get("specification") or "")))
        if specification_overlap:
            score += min(4, len(specification_overlap))
            reasons.append("ähnliche Spezifikation")
        if score < 7:
            continue
        scored.append((
            score,
            {
                "id": str(row.get("id")),
                "item_id": str(row.get("item_id")) if row.get("item_id") else None,
                "object_class": row.get("object_class_name"),
                "object_type": row.get("object_type"),
                "brand": row.get("brand"),
                "model": row.get("model"),
                "specification": corrected.get("specification"),
                "construction_year": corrected.get("construction_year"),
                "condition": row.get("condition") or corrected.get("condition"),
                "value_estimate": value_estimate,
                "estimated_age_years": estimated_age,
                "match_score": score,
                "match_reason": ", ".join(dict.fromkeys(reasons)) or "ähnliche geprüfte Eingaben",
                "notes": row.get("notes"),
                "created_at": row.get("created_at").isoformat() if hasattr(row.get("created_at"), "isoformat") else str(row.get("created_at") or ""),
                "source": "gepruefte_wertreferenz",
            },
        ))
    scored.sort(key=lambda entry: entry[0], reverse=True)
    return [row for _, row in scored[:limit]]


def estimate_value_from_web_legacy(item: dict[str, Any], ai_context: str, sources: list[dict[str, Any]]) -> dict[str, Any]:
    price_candidates = source_price_candidates(sources, item, ai_context)
    object_class = normalize_bga_object_class(item.get("object_class_name"), item.get("object_type"))
    condition = str(item.get("condition") or "gebraucht")
    condition_factor = {
        "neu": 1.0,
        "sehr_gut": 0.95,
        "gut": 0.90,
        "gebraucht": 0.85,
        "reparaturbeduerftig": 0.25,
        "defekt": 0.10,
        "aussondern": 0.04,
    }.get(condition, 0.85)
    if price_candidates:
        sorted_values = sorted(float(candidate["value"]) for candidate in price_candidates)
        basis = sorted_values[min(len(sorted_values) - 1, max(0, len(sorted_values) // 3))]
        estimate = max(1, round(basis * condition_factor))
        return {
            "estimated_value": estimate,
            "estimated_value_range": {"min": max(1, round(estimate * 0.75)), "max": max(1, round(estimate * 1.25))},
            "estimated_value_confidence": 0.72,
            "estimated_value_reason": f"Aus {len(price_candidates)} Web-Preisfund(en) konservativ abgeleitet; Zustand berücksichtigt.",
            "value_requires_review": True,
            "price_candidates": price_candidates[:5],
        }
    if object_class in {"Computermaus", "Tastatur", "Kaffeemaschine"}:
        return {
            "estimated_value": None,
            "estimated_value_range": {"min": None, "max": None},
            "estimated_value_confidence": 0.0,
            "estimated_value_reason": "Kein belastbarer Gebrauchtmarktpreis gefunden; Wert bleibt offen.",
            "value_requires_review": True,
            "price_candidates": [],
        }
    value_min, value_max = estimate_value_range(item, ai_context)
    estimate = conservative_value_estimate(value_min, value_max)
    text = " ".join(str(item.get(key) or "").lower() for key in ["object_type", "object_class_name", "brand", "model"]) + " " + ai_context.lower()
    if sources and any(term in text for term in ["iphone", "smartphone"]):
        estimate = max(300, estimate)
    return bga_value_guardrail(item, ai_context, estimate)


def estimate_value_from_web(item: dict[str, Any], ai_context: str, sources: list[dict[str, Any]]) -> dict[str, Any]:
    price_candidates = source_price_candidates(sources, item, ai_context)
    condition = str(item.get("condition") or "gebraucht")
    condition_factor = {
        "neu": 1.0,
        "sehr_gut": 1.0,
        "gut": 1.0,
        "gebraucht": 1.0,
        "reparaturbeduerftig": 0.35,
        "defekt": 0.15,
        "aussondern": 0.05,
    }.get(condition, 1.0)
    trusted_candidates = [
        candidate for candidate in price_candidates
        if candidate.get("market_kind") == "gebraucht" and candidate.get("reference_match") == "exact"
    ]
    similar_candidates = [
        candidate for candidate in price_candidates
        if candidate.get("market_kind") == "gebraucht" and candidate.get("reference_match") == "similar"
    ]
    if trusted_candidates:
        sorted_values = sorted(float(candidate["value"]) for candidate in trusted_candidates)
        basis = sorted_values[min(len(sorted_values) - 1, max(0, len(sorted_values) // 3))]
        estimate = max(1, round(basis * condition_factor))
        return {
            "estimated_value": estimate,
            "estimated_value_range": {"min": max(1, round(estimate * 0.75)), "max": max(1, round(estimate * 1.25))},
            "estimated_value_confidence": 0.82,
            "estimated_value_reason": f"Referenzpreis aus {len(trusted_candidates)} eindeutig passendem Gebrauchtmarkt-Treffer(n). Zustand bleibt fachlich zu prüfen.",
            "value_requires_review": True,
            "price_candidates": price_candidates[:8],
            "value_source": "gebrauchtmarkt_referenz",
            "valuation_state": "reference_available",
            "reference_price_available": True,
            "reference_price_label": "Referenzpreis verfügbar",
            "selected_price_reference": trusted_candidates[0],
        }
    if similar_candidates:
        sorted_values = sorted(float(candidate["value"]) for candidate in similar_candidates)
        return {
            "estimated_value": None,
            "estimated_value_range": {"min": max(1, round(sorted_values[0] * 0.8)), "max": max(1, round(sorted_values[-1] * 1.2))},
            "estimated_value_confidence": 0.35,
            "estimated_value_reason": "Ähnliche Gebrauchtmarkt-Treffer gefunden, aber kein eindeutiger Hersteller-/Modell-/Typ-Match. Kein Referenzpreis.",
            "value_requires_review": True,
            "price_candidates": price_candidates[:8],
            "value_source": "preisspanne_pruefen",
            "valuation_state": "range_review",
            "reference_price_available": False,
            "reference_price_label": "Preisspanne prüfen",
        }
    return {
        "estimated_value": None,
        "estimated_value_range": {"min": None, "max": None},
        "estimated_value_confidence": 0.0,
        "estimated_value_reason": "Kein eindeutiger Gebrauchtmarkt-Referenzpreis gefunden; Wert bleibt offen.",
        "value_requires_review": True,
        "price_candidates": price_candidates[:8],
        "value_source": "keine_referenz",
        "valuation_state": "no_reference",
        "reference_price_available": False,
        "reference_price_label": "Kein Referenzpreis verfügbar",
    }


def bga_value_guardrail(item: dict[str, Any], ai_context: str, estimated_value: int) -> dict[str, Any]:
    text = " ".join(
        str(item.get(key) or "").lower()
        for key in ["object_type", "object_class_name", "brand", "model"]
    ) + " " + ai_context.lower()
    rules = [
        (("nespresso", "kaffeemaschine", "kapselmaschine", "espresso"), 80, "Kaffeemaschine: Wert nur mit belastbarer Gebrauchtmarktquelle übernehmen."),
        (("computermaus", " maus", "mouse"), 100, "Computermaus: vierstellige Werte sind unplausibel."),
        (("tastatur", "keyboard"), 250, "Tastatur: nur niedriger bis mittlerer Gebrauchtwert plausibel."),
        (("monitor", "display"), 600, "Standardmonitor: vierstellige Werte nur mit sehr belastbarer Spezialbegründung."),
        (("drucker",), 800, "Drucker: Ausreißer ohne Modell-/Marktbeleg nicht übernehmen."),
        (("bürostuhl", "buerostuhl", "stuhl"), 800, "Bürostuhl: Wert stark markenabhängig und prüfpflichtig."),
        (("schreibtisch", "regal"), 800, "Büroausstattung: ohne belastbare Marke kein hoher KI-Wert."),
        (("iphone", "smartphone"), 1400, "Smartphone: Wert stark modell- und zustandsabhängig."),
        (("ladegerät", "ladegeraet"), 250, "Ladegerät: hoher Wert ohne Typenschild/Modell nicht plausibel."),
    ]
    for needles, max_value, reason in rules:
        if any(needle in text for needle in needles):
            if any(term in text for term in ["nespresso", "kaffeemaschine", "kapselmaschine", "espresso", "computermaus", " maus", "mouse", "tastatur", "keyboard"]):
                return {
                    "estimated_value": None,
                    "estimated_value_range": {"min": None, "max": None},
                    "estimated_value_confidence": 0.0,
                    "estimated_value_reason": f"{reason} Ohne geprüfte Gebrauchtquelle bleibt der Wert offen.",
                    "value_requires_review": True,
                }
            if estimated_value > max_value:
                return {
                    "estimated_value": None,
                    "estimated_value_range": {"min": None, "max": None},
                    "estimated_value_confidence": 0.0,
                    "estimated_value_reason": f"{reason} KI-Wert {estimated_value} EUR wurde verworfen.",
                    "value_requires_review": True,
                }
            return {
                "estimated_value": estimated_value,
                "estimated_value_range": {"min": max(1, round(estimated_value * 0.8)), "max": max(1, round(estimated_value * 1.2))},
                "estimated_value_confidence": 0.55,
                "estimated_value_reason": f"Konservative Heuristik für {reason.split(':', 1)[0]}; manuell prüfen.",
                "value_requires_review": True,
            }
    if estimated_value >= 1000 and not any(
        term in text
        for term in ["hebeb", "wuchtmaschine", "reifenmontiermaschine", "kompressor", "maschine", "werkzeugwagen", "rtx", "gaming laptop"]
    ):
        return {
            "estimated_value": None,
            "estimated_value_range": {"min": None, "max": None},
            "estimated_value_confidence": 0.0,
            "estimated_value_reason": f"Hoher KI-Wert {estimated_value} EUR ohne passende Objektklasse/Beleg verworfen.",
            "value_requires_review": True,
        }
    return {
        "estimated_value": estimated_value,
        "estimated_value_range": {"min": max(1, round(estimated_value * 0.8)), "max": max(1, round(estimated_value * 1.2))},
        "estimated_value_confidence": 0.45 if estimated_value >= 1000 else 0.6,
        "estimated_value_reason": "Konservative Heuristik aus Objektart, Zustand und Referenzhinweisen; fachlich prüfen.",
        "value_requires_review": True,
    }


def is_tire_item(item: dict[str, Any], ai_context: str = "") -> bool:
    text = " ".join(
        str(item.get(key) or "").lower()
        for key in ["object_type", "object_class_name", "brand", "model"]
    ) + " " + ai_context.lower()
    return any(term in text for term in ["reifen", "radsatz", "sommerreifen", "winterreifen", "ganzjahresreifen", "dot "])


def tire_dot_age_years(item: dict[str, Any], ai_context: str = "") -> float | None:
    text = " ".join(
        [str(item.get(key) or "") for key in ["object_type", "brand", "model", "serial_number"]]
        + [ai_context]
    )
    match = re.search(r"\bDOT\D{0,8}(\d{2})(\d{2})\b", text, flags=re.IGNORECASE)
    if not match and re.search(r"\bdot\b", text, flags=re.IGNORECASE):
        match = re.search(r"\b(\d{2})(\d{2})\b", text)
    if not match:
        if item.get("estimated_age_years") is not None and item.get("age_source") == "dot":
            try:
                return max(0.0, float(item["estimated_age_years"]))
            except (TypeError, ValueError):
                return None
        return None
    week = int(match.group(1))
    year_suffix = int(match.group(2))
    if week < 1 or week > 53:
        return None
    production_year = 2000 + year_suffix if year_suffix < 80 else 1900 + year_suffix
    current_year = datetime.utcnow().year
    current_week = int(datetime.utcnow().strftime("%V"))
    age = (current_year - production_year) + ((current_week - week) / 52)
    return round(max(0.0, age), 1)


def tire_profile_depth_mm(ai_context: str = "") -> float | None:
    patterns = [
        r"profil(?:tiefe)?\D{0,12}(\d{1,2}(?:[,.]\d)?)\s*mm",
        r"(\d{1,2}(?:[,.]\d)?)\s*mm\D{0,12}profil",
    ]
    for pattern in patterns:
        match = re.search(pattern, ai_context, flags=re.IGNORECASE)
        if match:
            try:
                return float(match.group(1).replace(",", "."))
            except ValueError:
                return None
    return None


def tire_season(item: dict[str, Any], ai_context: str = "") -> str:
    text = " ".join(str(part or "").lower() for part in [item.get("object_type"), item.get("model"), ai_context])
    if any(term in text for term in ["winter", "m+s", "m+s", "alpin", "snow"]):
        return "winter"
    if any(term in text for term in ["ganzjahr", "allseason", "all season"]):
        return "ganzjahr"
    return "sommer"


def tire_age_factor(age_years: float) -> float:
    if age_years <= 0:
        return 1.0
    factor = 0.75
    additional_years = max(0, int(age_years))
    if age_years > 1:
        additional_years = int(age_years - 1)
        factor *= 0.85 ** additional_years
    return max(0.12, factor)


def tire_profile_factor(profile_mm: float | None, season: str) -> tuple[float, str]:
    if profile_mm is None:
        return 0.85, "Profiltiefe fehlt: konservativer Zusatzabschlag 15%."
    if profile_mm >= 7:
        return 1.0, "Volles/gutes Profil: kein Profilabschlag; DOT-Alter bleibt wertmindernd."
    if profile_mm >= 5:
        return 0.85, "Profil 5-6,9 mm: Zusatzabschlag 15%."
    if profile_mm >= 4:
        return 0.70, "Profil 4-4,9 mm: Zusatzabschlag 30%."
    if profile_mm >= 3:
        factor = 0.55 if season == "sommer" else 0.35
        return factor, "Profil kritisch: deutlicher Zusatzabschlag."
    return 0.15, "Profil unter Praxisgrenze: nur sehr geringer Restwert."


def tire_condition_factor(condition: str | None) -> float:
    return {
        "neu": 1.0,
        "sehr_gut": 0.95,
        "gut": 0.9,
        "gebraucht": 0.8,
        "reparaturbeduerftig": 0.45,
        "defekt": 0.10,
        "aussondern": 0.03,
    }.get(str(condition or "gebraucht"), 0.8)


def conservative_tire_estimate(item: dict[str, Any], ai_context: str, sources: list[dict[str, str]]) -> dict[str, Any]:
    value_min, value_max = estimate_value_range({**item, "condition": "neu"}, ai_context)
    # Annäherung an günstigsten heutigen Neupreis: unteres Ende aus Web-/Referenzheuristik, dann konservativ.
    lowest_new_price = max(1, round(value_min * 0.9))
    dot_age = tire_dot_age_years(item, ai_context)
    estimated_age = dot_age if dot_age is not None else conservative_age_estimate(estimate_age_years(item, sources, ai_context))
    age_factor = tire_age_factor(float(estimated_age or 0))
    season = tire_season(item, ai_context)
    profile_mm = tire_profile_depth_mm(ai_context)
    profile_factor, profile_note = tire_profile_factor(profile_mm, season)
    condition_factor = tire_condition_factor(item.get("condition"))
    safety_factor = 0.90
    estimated_value = max(1, round(lowest_new_price * age_factor * profile_factor * condition_factor * safety_factor))
    return {
        "estimated_value": estimated_value,
        "estimated_age_years": round(float(estimated_age or 0), 1) if estimated_age is not None else None,
        "estimated_value_range": {"min": max(1, round(estimated_value * 0.8)), "max": max(1, round(estimated_value * 1.2))},
        "tire_valuation": {
            "lowest_new_price_basis": lowest_new_price,
            "dot_age_years": round(dot_age, 1) if dot_age is not None else None,
            "profile_depth_mm": profile_mm,
            "season": season,
            "age_factor": round(age_factor, 3),
            "profile_factor": round(profile_factor, 3),
            "condition_factor": round(condition_factor, 3),
            "safety_factor": safety_factor,
            "policy": "Günstigster plausibler Neupreis minus DOT-Alter, Profil, Zustand und Sicherheitsabschlag.",
        },
        "notes": (
            "Konservative Reifenbewertung: volles Profil hebt DOT-Alter nicht auf. "
            f"{profile_note} Lagerreifen werden nach DOT-Alter bewertet, auch wenn sie ungefahren wirken."
        ),
    }


def conservative_value_estimate(value_min: int, value_max: int) -> int:
    # Kaufprüfung: realistisch bleiben, aber bei Unsicherheit nicht zu hoch bewerten.
    lower_quartile = value_min + ((value_max - value_min) * 0.25)
    return max(1, round(lower_quartile * 0.9))


def parse_construction_year(item: dict[str, Any]) -> int | None:
    value = normalize_deep_dive_text(item.get("construction_year"))
    if not value:
        return None
    match = re.search(r"\b(19[8-9][0-9]|20[0-3][0-9])\b", value)
    if not match:
        return None
    year = int(match.group(1))
    current_year = datetime.utcnow().year
    if year > current_year + 1:
        return None
    return year


def construction_year_age(item: dict[str, Any]) -> float | None:
    year = parse_construction_year(item)
    if year is None:
        return None
    return round(max(0.0, datetime.utcnow().year - year), 1)


def estimate_age_years(item: dict[str, Any], sources: list[dict[str, str]], ai_context: str = "") -> float | None:
    construction_age = construction_year_age(item)
    if construction_age is not None:
        return construction_age
    if item.get("estimated_age_years") is not None and item.get("age_source") != "schaetzung":
        try:
            return round(float(item["estimated_age_years"]), 1)
        except (TypeError, ValueError):
            pass
    text = " ".join(
        [str(item.get(key) or "") for key in ["object_type", "brand", "model", "serial_number"]]
        + [source.get("title", "") for source in sources]
        + [ai_context]
    )
    lower_text = text.lower()
    if "14700hx" in lower_text or "rtx 4070" in lower_text:
        return 2.0
    if "13700hx" in lower_text or "13620h" in lower_text or "rtx 4060" in lower_text:
        return 3.0
    if "12700h" in lower_text or "rtx 3060" in lower_text or "rtx 3070" in lower_text:
        return 4.0
    iphone_years = {
        "iphone 16": 2024,
        "iphone 15": 2023,
        "iphone 14": 2022,
        "iphone 13": 2021,
        "iphone 12": 2020,
        "iphone 11": 2019,
    }
    for marker, release_year in iphone_years.items():
        if marker in lower_text:
            return max(0.0, round(datetime.utcnow().year - release_year, 1))
    if "redragon" in lower_text and "m908" in lower_text:
        return max(0.0, round(datetime.utcnow().year - 2018, 1))
    if "agile-splendor" in lower_text or ("ips" in lower_text and "monitor" in lower_text):
        return 2.0
    years = [int(value) for value in re.findall(r"\b(20[0-2][0-9]|19[8-9][0-9])\b", text)]
    if years:
        return max(0.0, round(datetime.utcnow().year - max(years), 1))
    return None


def conservative_age_estimate(value: float | None) -> float | None:
    if value is None:
        return None
    return round(max(0.0, value + 1.0), 1)


def build_deep_dive_result(item_id: str) -> dict[str, Any]:
    item = fetch_one(
        """
        SELECT i.*, oc.name AS object_class_name
        FROM inventory_items i
        LEFT JOIN object_classes oc ON oc.id = i.object_class_id
        WHERE i.id = %s
        """,
        (item_id,),
    )
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    ai_context = latest_ai_context(item_id)
    value_references = select_value_references(item, ai_context)
    best_value_reference = value_references[0] if value_references else None
    sources, search_provider, web_search_error, search_queries = collect_search_sources(item, ai_context)
    query = search_queries[0] if search_queries else deep_dive_query(item, ai_context)
    if not sources:
        reference_value = numeric_or_none(best_value_reference.get("value_estimate")) if best_value_reference else None
        reference_age = numeric_or_none(best_value_reference.get("estimated_age_years")) if best_value_reference else None
        construction_age = construction_year_age(item)
        if best_value_reference and (reference_value is not None or reference_age is not None):
            reference_value_range = {
                "min": max(1, round(reference_value * 0.9)) if reference_value is not None else None,
                "max": max(1, round(reference_value * 1.1)) if reference_value is not None else None,
            }
            reference_age_value = construction_age if construction_age is not None else reference_age
            return {
                "ai_stage": "deep_dive",
                "estimated_by_ai": True,
                "estimation_policy": "verified_reference",
                "web_search_performed": False,
                "search_provider": search_provider,
                "search_queries": search_queries,
                "query": query,
                "research_basis": deep_dive_research_basis(item),
                "technical_context": ai_context,
                "sources": [],
                "web_search_error": web_search_error or "Keine belastbare Webquelle gefunden; geprüfte Referenz genutzt.",
                "estimated_age_years": reference_age_value,
                "age_source": "gepruefte_wertreferenz" if construction_age is None and reference_age is not None else "baujahr",
                "age_verification_status": "geschaetzt",
                "age_confidence": 0.78 if construction_age is None and reference_age is not None else 0.9,
                "age_reason": "Aus geprüfter Wertreferenz übernommen." if construction_age is None and reference_age is not None else "Aus eingegebenem Baujahr abgeleitet.",
                "age_requires_review": True,
                "estimated_value": round(reference_value) if reference_value is not None else None,
                "estimated_value_range": reference_value_range,
                "estimated_value_confidence": 0.82 if reference_value is not None else 0.0,
                "estimated_value_reason": f"Aus geprüfter Wertreferenz abgeleitet: {best_value_reference.get('match_reason')}.",
                "value_requires_review": True,
                "value_source": "gepruefte_wertreferenz",
                "valuation_state": "reference_available",
                "reference_price_available": reference_value is not None,
                "reference_price_label": "Referenzpreis verfügbar" if reference_value is not None else "Kein Referenzpreis verfügbar",
                "value_reference_used": best_value_reference,
                "matching_value_references": value_references,
                "notes": "Geprüfte interne Referenz genutzt. Wert und Alter bleiben prüfpflichtige Vorschläge.",
                "manual_review_required": True,
                **deep_dive_dossier(
                    item,
                    ai_context,
                    [],
                    estimated_value=round(reference_value) if reference_value is not None else None,
                    estimated_value_range=reference_value_range,
                    estimated_value_confidence=0.82 if reference_value is not None else 0.0,
                    value_source="gepruefte_wertreferenz",
                    valuation_state="reference_available" if reference_value is not None else "no_reference",
                    reference_price_available=reference_value is not None,
                    reference_price_label="Referenzpreis verfügbar" if reference_value is not None else "Kein Referenzpreis verfügbar",
                    estimated_age_years=reference_age_value,
                ),
            }
        return {
            "ai_stage": "deep_dive",
            "estimated_by_ai": True,
            "estimation_policy": "source_required",
            "web_search_performed": False,
            "search_provider": search_provider,
            "search_queries": search_queries,
            "query": query,
            "research_basis": deep_dive_research_basis(item),
            "technical_context": ai_context,
            "sources": [],
            "web_search_error": web_search_error or "Keine belastbare Webquelle gefunden.",
            "estimated_age_years": None,
            "age_source": "unbekannt",
            "age_verification_status": "offen",
            "age_confidence": 0.0,
            "age_reason": "Keine belastbare Webquelle oder sichtbare Altersgrundlage gefunden.",
            "age_requires_review": True,
            "estimated_value": None,
            "estimated_value_range": {"min": None, "max": None},
            "estimated_value_confidence": 0.0,
            "estimated_value_reason": "Keine belastbare Webquelle gefunden; Wert bleibt offen.",
            "value_requires_review": True,
            "value_source": "keine_webquelle",
            "valuation_state": "no_reference",
            "reference_price_available": False,
            "reference_price_label": "Kein Referenzpreis verfügbar",
            "value_reference_used": None,
            "matching_value_references": value_references,
            "notes": "Websuche ohne verwertbare Quelle. Preis und Alter werden nicht geraten und müssen manuell geprüft werden.",
            "manual_review_required": True,
            **deep_dive_dossier(
                item,
                ai_context,
                [],
                value_source="keine_webquelle",
                valuation_state="no_reference",
                reference_price_available=False,
                reference_price_label="Kein Referenzpreis verfügbar",
            ),
        }
    if is_tire_item(item, ai_context):
        tire_result = conservative_tire_estimate(item, ai_context, sources)
        value_min = tire_result["estimated_value_range"]["min"]
        value_max = tire_result["estimated_value_range"]["max"]
        estimated_value = tire_result["estimated_value"]
        estimated_age = tire_result["estimated_age_years"]
        notes = tire_result["notes"]
        extra_result = {
            "tire_valuation": tire_result["tire_valuation"],
            "valuation_state": "range_review",
            "reference_price_available": False,
            "reference_price_label": "Preisspanne prüfen",
            "value_source": "reifen_heuristik",
        }
    else:
        guarded_value = estimate_value_from_web(item, ai_context, sources)
        value_range = guarded_value.get("estimated_value_range") or {}
        estimated_value = guarded_value.get("estimated_value")
        value_min = value_range.get("min")
        value_max = value_range.get("max")
        value_source = str(guarded_value.get("value_source") or "keine_referenz")
        estimated_value_confidence = float(guarded_value.get("estimated_value_confidence") or 0)
        estimated_value_reason = "KI-Webpreise werden nicht automatisch als Gebrauchtmarktwert übernommen. Ohne geprüfte Referenz bleibt der Wert offen."
        estimated_value_reason = str(guarded_value.get("estimated_value_reason") or "Kein eindeutiger Referenzpreis gefunden.")
        value_reference_used = None
        reference_value = numeric_or_none(best_value_reference.get("value_estimate")) if best_value_reference else None
        if reference_value is not None and best_value_reference and int(best_value_reference.get("match_score") or 0) >= 9:
            estimated_value = round(reference_value)
            value_min = max(1, round(reference_value * 0.9))
            value_max = max(1, round(reference_value * 1.1))
            guarded_value = {
                **guarded_value,
                "estimated_value": estimated_value,
                "estimated_value_range": {"min": value_min, "max": value_max},
                "estimated_value_confidence": max(float(guarded_value.get("estimated_value_confidence") or 0), 0.82),
                "estimated_value_reason": f"Aus geprüfter Wertreferenz abgeleitet: {best_value_reference.get('match_reason')}. Webquellen wurden zusätzlich geprüft.",
                "value_requires_review": True,
                "value_source": "gepruefte_wertreferenz",
                "valuation_state": "reference_available",
                "reference_price_available": True,
                "reference_price_label": "Referenzpreis verfügbar",
            }
        if reference_value is not None and best_value_reference and int(best_value_reference.get("match_score") or 0) >= 9:
            value_source = "gepruefte_wertreferenz"
            estimated_value_confidence = 0.82
            estimated_value_reason = f"Aus geprüfter Wertreferenz abgeleitet: {best_value_reference.get('match_reason')}."
            value_reference_used = best_value_reference
        elif guarded_value.get("reference_price_available") and guarded_value.get("estimated_value") is not None:
            value_source = str(guarded_value.get("value_source") or "gebrauchtmarkt_referenz")
            estimated_value_confidence = float(guarded_value.get("estimated_value_confidence") or 0.82)
            estimated_value_reason = str(guarded_value.get("estimated_value_reason") or "Eindeutige Gebrauchtmarkt-Referenz gefunden.")
            value_reference_used = guarded_value.get("selected_price_reference")
        else:
            estimated_value = None
        raw_age = estimate_age_years(item, sources, ai_context)
        age_from_construction_year = construction_year_age(item)
        manual_age = item.get("estimated_age_years") is not None and item.get("age_source") not in {None, "", "schaetzung", "unbekannt"}
        reference_age = numeric_or_none(best_value_reference.get("estimated_age_years")) if best_value_reference else None
        estimated_age = raw_age if (age_from_construction_year is not None or manual_age) else reference_age if reference_age is not None else conservative_age_estimate(raw_age)
        age_reason = (
            "Aus eingegebenem Baujahr abgeleitet."
            if age_from_construction_year is not None
            else "Aus manuell geprüftem Alter übernommen."
            if manual_age and estimated_age is not None
            else "Aus geprüfter Wertreferenz übernommen."
            if reference_age is not None and estimated_age is not None
            else "Aus sichtbarem Modellhinweis abgeleitet."
            if estimated_age is not None
            else "Keine belastbare Altersgrundlage erkannt."
        )
        notes = "Vorsichtige KI-Schätzung aus Objektangaben, Zustand und Referenzhinweisen. Unsichere oder unplausible Alter-/Wertangaben bleiben leer und müssen manuell geprüft werden."
        extra_result = {
            "estimated_value_confidence": estimated_value_confidence,
            "estimated_value_reason": estimated_value_reason,
            "value_requires_review": True,
            "value_reference_used": value_reference_used,
            "matching_value_references": value_references,
            "price_candidates": guarded_value.get("price_candidates") or [],
            "valuation_state": guarded_value.get("valuation_state") or "no_reference",
            "reference_price_available": bool(guarded_value.get("reference_price_available")),
            "reference_price_label": guarded_value.get("reference_price_label") or "Kein Referenzpreis verfügbar",
            "selected_price_reference": guarded_value.get("selected_price_reference"),
            "age_confidence": 0.8 if reference_age is not None and age_from_construction_year is None else 0.55 if estimated_age is not None else 0.0,
            "age_reason": age_reason,
            "age_requires_review": True,
            "value_source": value_source,
        }
    return {
        "ai_stage": "deep_dive",
        "estimated_by_ai": True,
        "estimation_policy": "konservativ",
        "web_search_performed": bool(sources),
        "search_provider": search_provider,
        "search_queries": search_queries,
        "query": query,
        "research_basis": deep_dive_research_basis(item),
        "technical_context": ai_context,
        "sources": sources,
        "web_search_error": web_search_error,
        "estimated_age_years": estimated_age,
        "age_source": "schaetzung",
        "age_verification_status": "geschaetzt",
        "estimated_value": estimated_value,
        "estimated_value_range": {"min": value_min, "max": value_max},
        "value_source": "ki_web_schaetzung",
        "notes": notes,
        "manual_review_required": True,
        **deep_dive_dossier(
            item,
            ai_context,
            sources,
            estimated_value=estimated_value,
            estimated_value_range={"min": value_min, "max": value_max},
            estimated_value_confidence=float(extra_result.get("estimated_value_confidence") or 0),
            value_source=str(extra_result.get("value_source") or "ki_web_schaetzung"),
            price_candidates=extra_result.get("price_candidates") or [],
            valuation_state=str(extra_result.get("valuation_state") or "no_reference"),
            reference_price_available=bool(extra_result.get("reference_price_available")),
            reference_price_label=str(extra_result.get("reference_price_label") or "Kein Referenzpreis verfügbar"),
            estimated_age_years=estimated_age,
        ),
        **extra_result,
    }


def process_deep_dive_item(item_id: str, expected_generation: Any = None) -> None:
    if ai_job_cancelled(item_id, expected_generation):
        audit("ai_deep_dive_cancelled", "inventory_item", item_id, {"stage": "deep_dive", "position": "before_start", "expected_generation": expected_generation})
        return
    if not update_ai_item_status(item_id, ai_status_for_mode("review", "running"), expected_generation, final_only_guard=True):
        audit("ai_deep_dive_cancelled", "inventory_item", item_id, {"stage": "deep_dive", "position": "before_running_status", "expected_generation": expected_generation})
        return
    try:
        result = build_deep_dive_result(item_id)
        if ai_job_cancelled(item_id, expected_generation):
            audit("ai_deep_dive_cancelled", "inventory_item", item_id, {"stage": "deep_dive", "position": "before_write", "expected_generation": expected_generation})
            return
        if expected_generation is None:
            row = execute(
                """
                INSERT INTO ai_results (item_id, ai_type, model_used, prompt_version, input_sources, result_json, confidence, status)
                VALUES (%s, 'deep_dive', %s, 'phase1-deep-dive-v1', %s::jsonb, %s::jsonb, %s, 'completed')
                RETURNING *
                """,
                (
                    item_id,
                    "web-search+heuristic",
                    '["item_fields","web_search","reference_heuristic"]',
                    json_string(result),
                    0.55,
                ),
            )
        else:
            expected_session = expected_session_generation(expected_generation)
            expected_item = expected_item_generation(expected_generation)
            row = execute(
                """
                INSERT INTO ai_results (item_id, ai_type, model_used, prompt_version, input_sources, result_json, confidence, status)
                SELECT i.id, 'deep_dive', %s, 'phase1-deep-dive-v1', %s::jsonb, %s::jsonb, %s, 'completed'
                FROM inventory_items i
                JOIN inventory_sessions s ON s.id = i.session_id
                WHERE i.id = %s
                  AND COALESCE(s.ai_cancel_generation, 0) = %s
                  AND (%s IS NULL OR COALESCE(i.ai_cancel_generation, 0) = %s)
                RETURNING *
                """,
                (
                    "web-search+heuristic",
                    '["item_fields","web_search","reference_heuristic"]',
                    json_string(result),
                    0.55,
                    item_id,
                    expected_session,
                    expected_item,
                    expected_item,
                ),
            )
        if not row:
            audit("ai_deep_dive_cancelled", "inventory_item", item_id, {"stage": "deep_dive", "position": "before_write_race", "expected_generation": expected_generation})
            return
        if not update_ai_item_status(item_id, ai_status_for_mode("review", "done"), expected_generation, final_only_guard=True):
            audit("ai_deep_dive_cancelled", "inventory_item", item_id, {"stage": "deep_dive", "position": "before_status", "expected_generation": expected_generation})
            return
        audit("ai_deep_dive_created", "inventory_item", item_id, result)
    except Exception as exc:
        update_ai_item_status(item_id, ai_status_for_mode("review", "open"), expected_generation, final_only_guard=True)
        audit("ai_deep_dive_failed", "inventory_item", item_id, {"error": type(exc).__name__, "message": str(exc)[:240]})
        raise


FAST_AI_BLOCKED_FIELDS = {
    "estimated_age_years",
    "estimated_value",
    "estimated_value_eur",
    "estimated_value_confidence",
    "estimated_value_reason",
    "value_estimate",
    "value_source",
    "value_requires_review",
    "age_confidence",
    "age_reason",
    "age_requires_review",
    "age_source",
    "age_verification_status",
}


def strip_fast_ai_estimates(suggestion: dict[str, Any]) -> dict[str, Any]:
    cleaned = dict(suggestion)
    for key in FAST_AI_BLOCKED_FIELDS:
        cleaned.pop(key, None)

    suggested_fields = cleaned.get("suggested_fields")
    if isinstance(suggested_fields, dict):
        cleaned["suggested_fields"] = {
            key: value
            for key, value in suggested_fields.items()
            if key not in FAST_AI_BLOCKED_FIELDS and key not in {"estimated_value", "estimated_age_years", "value_estimate"}
        }

    bga_detection = cleaned.get("bga_detection")
    if isinstance(bga_detection, dict):
        cleaned_bga = dict(bga_detection)
        for key in FAST_AI_BLOCKED_FIELDS:
            cleaned_bga.pop(key, None)
        cleaned["bga_detection"] = cleaned_bga

    cleaned["ai_latency_profile"] = "mobile_fast"
    cleaned["estimation_policy"] = "no_age_or_value_in_mobile_fast"
    return cleaned


def status_state_for_item_status(scope: str, status: str | None) -> str:
    if scope == "deep_dive":
        if status == "ki_pruefung_wartet":
            return "queued"
        if status == "ki_pruefung_laeuft":
            return "running"
        return "idle"
    if scope == "review":
        if status == "ki_pruefung_wartet":
            return "queued"
        if status == "ki_pruefung_laeuft":
            return "running"
        if status == "ki_pruefung_fertig":
            return "completed"
        return "idle"
    if status == "ki_schnell_wartet":
        return "queued"
    if status == "ki_schnell_laeuft":
        return "running"
    if status == "ki_schnell_fertig":
        return "completed"
    return "idle"


def ai_result_type_for_scope(scope: str) -> str:
    if scope == "deep_dive":
        return "deep_dive"
    if scope == "review":
        return "review_vision"
    return "quick_vision"


def ai_updated_fields(result: dict[str, Any] | None) -> list[str]:
    if not isinstance(result, dict):
        return []
    fields: set[str] = set()
    suggested_fields = result.get("suggested_fields")
    if isinstance(suggested_fields, dict):
        fields.update(str(key) for key, value in suggested_fields.items() if value)
    for key in ("object_type", "object_name", "specification", "serial_number", "construction_year", "condition", "suggested_remark"):
        if result.get(key):
            fields.add("object_type" if key == "object_name" else "remark" if key == "suggested_remark" else key)
    nameplate = result.get("nameplate_extraction")
    if isinstance(nameplate, dict):
        for key, mapped in {
            "suggested_object_type": "object_type",
            "suggested_specification": "specification",
            "serial_number": "serial_number",
            "construction_year": "construction_year",
            "suggested_remark": "remark",
        }.items():
            if nameplate.get(key):
                fields.add(mapped)
    return sorted(fields)


def first_ai_text(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def normalize_ai_condition(value: Any) -> str | None:
    text = str(value or "").strip().lower().replace(" ", "_").replace("-", "_")
    mapping = {
        "neu": "neu",
        "sehr_gut": "sehr_gut",
        "sehrgut": "sehr_gut",
        "gut": "gut",
        "gebraucht": "gebraucht",
        "normal": "gebraucht",
        "reparaturbeduerftig": "reparaturbeduerftig",
        "reparaturbedürftig": "reparaturbeduerftig",
        "defekt": "defekt",
        "aussondern": "aussondern",
        "unklar": "unklar",
    }
    return mapping.get(text)


def bga_uvv_status_from_ai(object_type: str, object_class: str, specification: str) -> str | None:
    text = f"{object_type} {object_class} {specification}".lower()
    uvv_required_tokens = (
        "hebebühne", "hebebuehne", "kompressor", "kran", "winde", "hubwagen", "stapler",
        "maschine", "reifenmontier", "wuchtmaschine", "werkstatt", "druckbehälter", "druckbehaelter",
    )
    uvv_not_required_tokens = (
        "smartphone", "iphone", "telefon", "computermaus", "maus", "tastatur", "keyboard",
        "monitor", "drucker", "scanner", "laptop", "notebook", "pc", "schreibtisch",
        "bürostuhl", "buerostuhl", "stuhl", "regal", "schrank", "kaffeemaschine",
    )
    if any(token in text for token in uvv_required_tokens):
        return None
    if any(token in text for token in uvv_not_required_tokens):
        return "nicht_uvv_pflichtig"
    return None


def object_class_id_for_ai(object_class: str, object_type: str) -> str | None:
    for value in (object_class, object_type):
        normalized = normalize_bga_object_class(value)
        if not normalized or normalized == "Unklar":
            continue
        row = fetch_one(
            """
            SELECT id
            FROM object_classes
            WHERE lower(name) = lower(%s) OR lower(slug) = lower(%s)
            LIMIT 1
            """,
            (normalized, normalized),
        )
        if row:
            return str(row["id"])
    return None


def apply_ai_suggestion_to_empty_item_fields(item_id: str, suggestion: dict[str, Any]) -> dict[str, Any]:
    item = fetch_one("SELECT * FROM inventory_items WHERE id = %s", (item_id,))
    if not item:
        return {}
    fields = suggestion.get("suggested_fields") if isinstance(suggestion.get("suggested_fields"), dict) else {}
    detection = suggestion.get("bga_detection") if isinstance(suggestion.get("bga_detection"), dict) else {}
    if isinstance(detection.get("suggested_fields"), dict):
        fields = {**detection["suggested_fields"], **fields}
    nameplate = (
        detection.get("nameplate_extraction")
        if isinstance(detection.get("nameplate_extraction"), dict)
        else suggestion.get("nameplate_extraction") if isinstance(suggestion.get("nameplate_extraction"), dict) else {}
    )
    object_type = first_ai_text(fields.get("object_type"), suggestion.get("object_type"), suggestion.get("object_name"), detection.get("object_name"), nameplate.get("suggested_object_type"))
    specification = first_ai_text(fields.get("specification"), suggestion.get("specification"), detection.get("specification"), nameplate.get("suggested_specification"))
    serial_number = first_ai_text(fields.get("serial_number"), suggestion.get("serial_number"), detection.get("serial_number"), nameplate.get("serial_number"))
    construction_year = first_ai_text(fields.get("construction_year"), suggestion.get("construction_year"), nameplate.get("construction_year"))
    brand = first_ai_text(suggestion.get("brand"), suggestion.get("manufacturer"), detection.get("brand"), detection.get("manufacturer"), nameplate.get("manufacturer"))
    model = first_ai_text(suggestion.get("model"), detection.get("model"), nameplate.get("model"), nameplate.get("type_designation"))
    remark = first_ai_text(fields.get("remark"), suggestion.get("suggested_remark"), detection.get("suggested_remark"), nameplate.get("suggested_remark"))
    condition = normalize_ai_condition(first_ai_text(fields.get("condition"), suggestion.get("condition"), suggestion.get("condition_guess"), detection.get("condition_guess")))
    object_class = first_ai_text(suggestion.get("object_class"), detection.get("object_class"))

    updates: dict[str, Any] = {}
    for key, value in {
        "object_type": object_type,
        "specification": specification,
        "serial_number": serial_number,
        "construction_year": construction_year,
        "brand": brand,
        "model": model,
        "remark": remark,
    }.items():
        if value and not str(item.get(key) or "").strip():
            updates[key] = value
    if condition and item.get("condition") in {None, "", "unklar", "gebraucht"}:
        updates["condition"] = condition
    if item.get("uvv_status") in {None, "", "unklar"}:
        uvv_status = bga_uvv_status_from_ai(object_type, object_class, specification)
        if uvv_status:
            updates["uvv_status"] = uvv_status
    if not item.get("object_class_id"):
        object_class_id = object_class_id_for_ai(object_class, object_type)
        if object_class_id:
            updates["object_class_id"] = object_class_id
    if not updates:
        return {}
    fields_sql = ", ".join(f"{key} = %s" for key in updates)
    execute(
        f"UPDATE inventory_items SET {fields_sql}, updated_at = now() WHERE id = %s RETURNING id",
        tuple(updates.values()) + (item_id,),
    )
    run_inventory_rework_check(item_id)
    return updates


def update_ai_item_status(item_id: str, status: str, expected_generation: Any = None, final_only_guard: bool = False) -> dict[str, Any] | None:
    expected_session = expected_session_generation(expected_generation)
    if expected_session is None:
        final_guard = "AND status <> 'finalisiert'" if final_only_guard else ""
        return execute(
            f"UPDATE inventory_items SET status = %s, updated_at = now() WHERE id = %s {final_guard} RETURNING id",
            (status, item_id),
        )
    expected_item = expected_item_generation(expected_generation)
    final_guard = "AND i.status <> 'finalisiert'" if final_only_guard else ""
    return execute(
        f"""
        UPDATE inventory_items i
        SET status = %s,
            updated_at = now()
        FROM inventory_sessions s
        WHERE i.id = %s
          AND s.id = i.session_id
          AND COALESCE(s.ai_cancel_generation, 0) = %s
          AND (%s IS NULL OR COALESCE(i.ai_cancel_generation, 0) = %s)
          {final_guard}
        RETURNING i.id
        """,
        (status, item_id, expected_session, expected_item, expected_item),
    )


def process_ai_item(item_id: str, mode: str = "fast", expected_generation: Any = None) -> None:
    normalized_mode = "review" if mode == "review" else "fast"
    if ai_job_cancelled(item_id, expected_generation):
        audit("ai_result_cancelled", "inventory_item", item_id, {"stage": normalized_mode, "position": "before_start", "expected_generation": expected_generation})
        return
    if not update_ai_item_status(item_id, ai_status_for_mode(normalized_mode, "running"), expected_generation):
        audit("ai_result_cancelled", "inventory_item", item_id, {"stage": normalized_mode, "position": "before_running_status", "expected_generation": expected_generation})
        return
    suggestion = build_ai_suggestion(item_id, normalized_mode)
    if normalized_mode == "fast":
        suggestion = strip_fast_ai_estimates(suggestion)
    if ai_job_cancelled(item_id, expected_generation):
        audit("ai_result_cancelled", "inventory_item", item_id, {"stage": normalized_mode, "position": "before_write", "expected_generation": expected_generation})
        return
    model_used = suggestion.pop("_model_used", "phase1-stub")
    confidence = float(suggestion.get("confidence") or 0)
    needs_review_ai = normalized_mode == "fast" and (
        confidence < 0.78
        or bool(suggestion.get("missing_fields"))
        or bool(suggestion.get("required_evidence_missing"))
    )
    suggestion["ai_stage"] = normalized_mode
    suggestion["needs_review_ai"] = needs_review_ai
    if expected_generation is None:
        row = execute(
            """
            INSERT INTO ai_results (item_id, ai_type, model_used, prompt_version, input_sources, result_json, confidence)
            VALUES (%s, %s, %s, %s, %s::jsonb, %s::jsonb, %s)
            RETURNING *
            """,
            (
                item_id,
                "quick_vision" if normalized_mode == "fast" else "review_vision",
                model_used,
                "phase1-fast-v2" if normalized_mode == "fast" else "phase1-review-v2",
                '["photos","audio"]',
                json_string(suggestion),
                confidence,
            ),
        )
    else:
        expected_session = expected_session_generation(expected_generation)
        expected_item = expected_item_generation(expected_generation)
        row = execute(
            """
            INSERT INTO ai_results (item_id, ai_type, model_used, prompt_version, input_sources, result_json, confidence)
            SELECT i.id, %s, %s, %s, %s::jsonb, %s::jsonb, %s
            FROM inventory_items i
            JOIN inventory_sessions s ON s.id = i.session_id
            WHERE i.id = %s
              AND COALESCE(s.ai_cancel_generation, 0) = %s
              AND (%s IS NULL OR COALESCE(i.ai_cancel_generation, 0) = %s)
            RETURNING *
            """,
            (
                "quick_vision" if normalized_mode == "fast" else "review_vision",
                model_used,
                "phase1-fast-v2" if normalized_mode == "fast" else "phase1-review-v2",
                '["photos","audio"]',
                json_string(suggestion),
                confidence,
                item_id,
                expected_session,
                expected_item,
                expected_item,
            ),
        )
    if not row:
        audit("ai_result_cancelled", "inventory_item", item_id, {"stage": normalized_mode, "position": "before_write_race", "expected_generation": expected_generation})
        return
    if ai_job_cancelled(item_id, expected_generation):
        audit("ai_result_cancelled", "inventory_item", item_id, {"stage": normalized_mode, "position": "before_tasks", "expected_generation": expected_generation})
        return
    applied_fields = apply_ai_suggestion_to_empty_item_fields(item_id, suggestion)
    create_rework_tasks(item_id, suggestion)
    execute(
        """
        UPDATE item_accounting_data
        SET accounting_status = CASE WHEN %s THEN 'buchhaltung_pruefen' ELSE accounting_status END
        WHERE item_id = %s
        RETURNING id
        """,
        (suggestion["requires_accounting_review"], item_id),
    )
    final_status = ai_status_for_mode(normalized_mode, "open" if needs_review_ai else "done")
    if ai_job_cancelled(item_id, expected_generation):
        audit("ai_result_cancelled", "inventory_item", item_id, {"stage": normalized_mode, "position": "before_status", "expected_generation": expected_generation})
        return
    if not update_ai_item_status(item_id, final_status, expected_generation):
        audit("ai_result_cancelled", "inventory_item", item_id, {"stage": normalized_mode, "position": "before_final_status", "expected_generation": expected_generation})
        return
    audit("ai_result_created", "inventory_item", item_id, {"stage": normalized_mode, "status": final_status, "applied_fields": applied_fields, **suggestion})




@app.post("/items/{item_id}/ai/run")
def run_ai(item_id: str, background_tasks: BackgroundTasks, mode: str = "fast") -> dict[str, Any]:
    require_item_session_open(item_id)
    active = current_active_ai_job(item_id)
    if active:
        return {
            "item_id": item_id,
            "status": active.get("status"),
            "stage": "review" if mode == "review" else "fast",
            "message": "KI läuft bereits für diesen Artikel",
            "already_running": True,
        }
    if not has_ai_object_photo(item_id):
        return {
            "item_id": item_id,
            "status": "skipped",
            "stage": "review" if mode == "review" else "fast",
            "message": "KI-Vorschlag erst nach Objektfoto möglich. Bitte Objektfoto aufnehmen und erneut starten.",
        }
    normalized_mode = "review" if mode == "review" else "fast"
    expected_generation = ai_job_generation_for_item(item_id)
    queued_status = ai_status_for_mode(normalized_mode, "queued")
    update_ai_item_status(item_id, queued_status, expected_generation)
    background_tasks.add_task(process_ai_item, item_id, normalized_mode, expected_generation)
    audit("ai_job_queued", "inventory_item", item_id, {"status": queued_status, "stage": normalized_mode, "expected_generation": expected_generation})
    label = "Prüf-KI" if normalized_mode == "review" else "Schnell-KI"
    return {"item_id": item_id, "status": queued_status, "stage": normalized_mode, "message": f"{label} läuft im Hintergrund"}

@app.get("/items/{item_id}/ai/status")
def get_ai_status(item_id: str, scope: str = "fast") -> dict[str, Any]:
    require_item_session_open(item_id)
    normalized_scope = "deep_dive" if scope in {"deep_dive", "deep-dive"} else "review" if scope == "review" else "fast"
    repair_stale_ai_statuses(item_id=item_id)
    item = fetch_one(
        "SELECT id, object_type, status, updated_at FROM inventory_items WHERE id = %s",
        (item_id,),
    )
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    result_type = ai_result_type_for_scope(normalized_scope)
    latest = fetch_one(
        """
        SELECT result_json, created_at, status
        FROM ai_results
        WHERE item_id = %s AND ai_type = %s
        ORDER BY created_at DESC
        LIMIT 1
        """,
        (item_id, result_type),
    )
    result = latest.get("result_json") if latest else None
    state = status_state_for_item_status(normalized_scope, item.get("status"))
    if result and state == "idle":
        state = "completed"
    messages = {
        "fast": {
            "queued": "Schnell-KI startet.",
            "running": "Schnell-KI erkennt Bezeichnung und Typenschild.",
            "completed": "Vorschlag bereit.",
            "cancelled": "KI wurde abgebrochen.",
            "idle": "Noch kein Schnell-KI-Lauf.",
        },
        "review": {
            "queued": "Pruef-KI startet.",
            "running": "Pruef-KI prueft den Datensatz.",
            "completed": "Pruefvorschlag bereit.",
            "cancelled": "KI wurde abgebrochen.",
            "idle": "Noch kein Pruef-KI-Lauf.",
        },
        "deep_dive": {
            "queued": "KI-Webrecherche startet.",
            "running": "KI-Webrecherche sammelt Quellen.",
            "completed": "Recherche-Dossier bereit.",
            "cancelled": "KI-Webrecherche wurde abgebrochen.",
            "idle": "Noch keine KI-Webrecherche.",
        },
    }
    return {
        "item_id": item_id,
        "scope": normalized_scope,
        "state": state,
        "message": messages[normalized_scope].get(state, "KI-Status unbekannt."),
        "result_preview": result,
        "updated_fields": ai_updated_fields(result),
        "can_cancel": state in {"queued", "running"},
        "started_at": item.get("updated_at") if state in {"queued", "running"} else None,
        "completed_at": latest.get("created_at") if latest and result else None,
        "item_status": item.get("status"),
    }


@app.post("/items/{item_id}/ai/deep-dive")
def run_ai_deep_dive(item_id: str, background_tasks: BackgroundTasks) -> dict[str, Any]:
    require_item_session_open(item_id)
    active = current_active_ai_job(item_id)
    if active:
        return {
            "item_id": item_id,
            "stage": "deep_dive",
            "status": active.get("status"),
            "message": "KI läuft bereits für diesen Artikel",
            "already_running": True,
        }
    expected_generation = ai_job_generation_for_item(item_id)
    update_ai_item_status(item_id, ai_status_for_mode("review", "queued"), expected_generation, final_only_guard=True)
    background_tasks.add_task(process_deep_dive_item, item_id, expected_generation)
    audit("ai_deep_dive_queued", "inventory_item", item_id, {"stage": "deep_dive", "expected_generation": expected_generation})
    return {
        "item_id": item_id,
        "stage": "deep_dive",
        "status": "queued",
        "message": "KI Deep Dive gestartet: Websuche, Alters- und Wertschätzung laufen im Hintergrund",
    }


@app.post("/items/{item_id}/ai/cancel")
def cancel_ai_item(item_id: str) -> dict[str, Any]:
    require_item_session_open(item_id)
    row = execute(
        f"""
        UPDATE inventory_items
        SET status = 'ki_pruefung_offen',
            ai_cancel_generation = COALESCE(ai_cancel_generation, 0) + 1,
            ai_cancelled_at = now(),
            updated_at = now()
        WHERE id = %s
          AND status IN ({AI_ACTIVE_STATUS_SQL})
          AND status <> 'finalisiert'
        RETURNING id, object_type, status, updated_at
        """,
        (item_id,),
    )
    audit("ai_job_cancelled", "inventory_item", item_id, {"status": row.get("status") if row else "not_active"})
    return {
        "item_id": item_id,
        "status": row.get("status") if row else "not_active",
        "message": "KI-Prozess abgebrochen. Du kannst ihn später neu starten.",
    }


@app.post("/sessions/{session_id}/ai/cancel")
def cancel_session_ai(session_id: str) -> dict[str, Any]:
    session = fetch_one("SELECT id, status FROM inventory_sessions WHERE id = %s", (session_id,))
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    active_rows = fetch_all(
        f"""
        SELECT id
        FROM inventory_items
        WHERE session_id = %s
          AND status IN ({AI_ACTIVE_STATUS_SQL})
        """,
        (session_id,),
    )
    generation_row = execute(
        """
        UPDATE inventory_sessions
        SET ai_cancel_generation = COALESCE(ai_cancel_generation, 0) + 1,
            ai_cancelled_at = now()
        WHERE id = %s
        RETURNING id, ai_cancel_generation, ai_cancelled_at
        """,
        (session_id,),
    )
    fetch_all(
        f"""
        UPDATE inventory_items
        SET status = 'ki_pruefung_offen',
            updated_at = now()
        WHERE session_id = %s
          AND status IN ({AI_ACTIVE_STATUS_SQL})
          AND status <> 'finalisiert'
        RETURNING id
        """,
        (session_id,),
    )
    generation = int((generation_row or {}).get("ai_cancel_generation") or 0)
    audit(
        "session_ai_cancelled",
        "inventory_session",
        session_id,
        {"cancelled": len(active_rows), "ai_cancel_generation": generation},
    )
    return {
        "session_id": session_id,
        "cancelled": len(active_rows),
        "ai_cancel_generation": generation,
        "status": "ki_pruefung_offen",
        "message": f"{len(active_rows)} KI-Prozess(e) im Raum gestoppt.",
    }


@app.post("/sessions/{session_id}/ai/review")
def run_session_review_ai(session_id: str, background_tasks: BackgroundTasks) -> dict[str, Any]:
    session = fetch_one("SELECT id, status, COALESCE(ai_cancel_generation, 0) AS ai_cancel_generation FROM inventory_sessions WHERE id = %s", (session_id,))
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session["status"] != "open":
        raise HTTPException(status_code=409, detail="Raum ist abgeschlossen")

    repair_stale_ai_statuses(session_id=session_id)
    rows = fetch_all(
        f"""
        SELECT i.id, COALESCE(i.ai_cancel_generation, 0) AS ai_cancel_generation
        FROM inventory_items i
        WHERE session_id = %s
          AND finalized_at IS NULL
          AND locked_at IS NULL
          AND status NOT IN ({AI_ACTIVE_STATUS_SQL}, 'finalisiert')
          AND EXISTS (
            SELECT 1 FROM item_photos p
            WHERE p.item_id = i.id AND p.photo_type IN ('object', 'object_front')
          )
        ORDER BY created_at
        """,
        (session_id,),
    )
    expected_generation = int(session.get("ai_cancel_generation") or 0)
    for row in rows:
        item_id = str(row["id"])
        item_generation = {"session": expected_generation, "item": int(row.get("ai_cancel_generation") or 0)}
        update_ai_item_status(item_id, ai_status_for_mode("review", "queued"), item_generation)
        background_tasks.add_task(process_ai_item, item_id, "review", item_generation)
    audit("session_ai_review_queued", "inventory_session", session_id, {"queued": len(rows), "expected_generation": expected_generation})
    return {"session_id": session_id, "queued": len(rows), "status": "ki_pruefung_wartet"}


def json_string(value: Any) -> str:
    import json

    return json.dumps(value, default=str)


def excel_value(value: Any) -> Any:
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def format_excel_datetime(value: Any) -> str | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def label_status(value: Any) -> str:
    labels = {
        "erfasst": "Erfasst",
        "ki_vorgefuellt": "KI vorgefüllt",
        "nacharbeit_erfasser": "Noch zu ergänzen: Erfasser",
        "nacharbeit_pruefer": "Noch zu ergänzen: Prüfer",
        "nacharbeit_buchhaltung": "Später auswerten",
        "nacharbeit_technik": "Noch zu ergänzen: Technik",
        "finalisierbar": "Finalisierbar",
        "geprueft": "Geprüft",
        "finalisiert": "Finalisiert",
        "ki_schnell_wartet": "Schnell-KI wartet",
        "ki_schnell_laeuft": "Schnell-KI läuft",
        "ki_schnell_fertig": "Schnell-KI fertig",
        "ki_pruefung_offen": "Prüf-KI offen",
        "ki_pruefung_wartet": "Prüf-KI wartet",
        "ki_pruefung_laeuft": "Prüf-KI läuft",
        "ki_pruefung_fertig": "Prüf-KI fertig",
        "open": "Offen",
        "closed": "Abgeschlossen",
        "hochgeladen": "Hochgeladen",
    }
    return labels.get(str(value or ""), str(value or ""))


def label_role(value: Any) -> str:
    roles = {
        "Buchhaltung": "Spätere kaufmännische Auswertung",
        "Auswertung": "Spätere Auswertung",
        "Prüfer": "Prüfer",
        "Erfasser": "Erfasser",
        "Technik": "Technik",
    }
    return roles.get(str(value or ""), str(value or "Nicht zugeordnet"))


def label_field(value: Any) -> str:
    raw = str(value or "").strip()
    replacements = {
        "serial_number": "Seriennummer",
        "serial number": "Seriennummer",
        "seriennummer": "Seriennummer",
        "Seriennummer": "Seriennummer",
        "model": "Modell",
        "modell": "Modell",
        "Modell": "Modell",
        "Foto/Nachweis": "Foto oder Nachweisfoto",
        "Typenschildfoto": "Typenschildfoto",
        "object": "Objektfoto",
        "object_front": "Objektfoto",
        "object_back": "Rückseite/Detail",
        "type_plate": "Typenschild",
        "uvv_label": "UVV-Siegel",
        "condition_detail": "Zustandsdetail",
        "other": "Sonstiges Foto",
        "tire_overview": "Gesamtfoto Reifen/Radsatz",
        "tread": "Profilbild",
        "tire_size": "Reifengröße",
        "rim": "Felge",
        "damage": "Schaden",
        "dot_photo": "DOT-Foto",
        "dot": "DOT-Foto",
        "DOT-Foto": "DOT-Foto",
        "Anschaffungsdatum": "Anschaffungsdatum später klären",
        "Buchwert": "Wert später klären",
        "Anlagenummer": "Zuordnung später klären",
        "Anlagenummer/Buchwert": "Wert/Zuordnung später klären",
        "Wert/Zuordnung später klären": "Wert/Zuordnung später klären",
    }
    if raw in replacements:
        return replacements[raw]
    label = raw.replace("_", " ").strip()
    return f"{label[:1].upper()}{label[1:]}" if label else "offen"


def label_field_list(value: Any) -> str:
    parts = [part.strip() for part in str(value or "").split(",") if part.strip()]
    return ", ".join(label_field(part) for part in parts) if parts else ""


def export_value(row: dict[str, Any], key: str) -> Any:
    if key == "lfd_photo":
        return row.get("sequence_number") or row.get("inventory_id") or row.get("temporary_id") or ""
    if key == "function_ok_yes":
        return "X" if row.get("function_ok") == "ja" else ""
    if key == "function_ok_no":
        return "X" if row.get("function_ok") == "nein" else ""
    if key == "bga_remark":
        notes = [str(row.get("condition_note") or "").strip(), str(row.get("remark") or "").strip()]
        if row.get("function_ok") == "nicht_geprueft":
            notes.append("Funktion nicht geprüft")
        if row.get("uvv_status") in {"nicht_vorhanden", "unklar"}:
            notes.append(export_value(row, "uvv_status_label"))
        return " | ".join(note for note in notes if note)
    if key == "captured_by_name":
        return row.get("created_by_name") or row.get("session_started_by_name") or "Nicht erfasst"
    if key == "reviewed_by_name":
        return row.get("reviewed_by_name") or ("Prüfer nicht gesetzt" if row.get("manual_reviewed_at") else "")
    if key == "status_label":
        return label_status(row.get("status"))
    if key == "review_status_label":
        return label_status(row.get("review_status"))
    if key == "session_status_label":
        return label_status(row.get("session_status"))
    if key == "condition_label":
        return str(row.get("condition") or "").replace("_", " ")
    if key == "function_ok_label":
        return {
            "ja": "Ja",
            "nein": "Nein",
            "nicht_geprueft": "Nicht geprüft",
        }.get(str(row.get("function_ok") or ""), row.get("function_ok"))
    if key == "uvv_status_label":
        return {
            "vorhanden": "UVV vorhanden",
            "nicht_vorhanden": "UVV nicht vorhanden",
            "nicht_uvv_pflichtig": "nicht UVV-pflichtig",
            "unklar": "unklar",
        }.get(str(row.get("uvv_status") or ""), row.get("uvv_status"))
    if key == "uvv_valid_until":
        if row.get("uvv_valid_until"):
            return format_excel_datetime(row.get("uvv_valid_until"))
        if row.get("uvv_status") == "nicht_uvv_pflichtig":
            return "nicht UVV-pflichtig"
        if row.get("uvv_status") == "nicht_vorhanden":
            return "UVV nicht vorhanden"
        return ""
    if key == "inspection_book_label":
        return {
            "ja": "Ja",
            "nein": "Nein",
            "nicht_erforderlich": "Nicht erforderlich",
            "unklar": "Unklar",
        }.get(str(row.get("inspection_book_available") or ""), row.get("inspection_book_available"))
    if key == "type_plate_status_label":
        return {
            "vorhanden": "vorhanden",
            "nicht_vorhanden": "nicht vorhanden",
            "uebersprungen": "übersprungen",
            "nicht_geprueft": "nicht geprüft",
        }.get(str(row.get("type_plate_status") or ""), row.get("type_plate_status"))
    if key == "value_provenance":
        if row.get("latest_deep_dive_at") and row.get("value_estimate") is not None:
            return "KI-Schätzung, konservativ"
        return "Manuell/ungeklärt" if row.get("value_estimate") is not None else ""
    if key == "age_provenance":
        if row.get("latest_deep_dive_at") and row.get("estimated_age_years") is not None:
            return "KI-Altersschätzung, konservativ"
        if row.get("age_source") and row.get("age_source") != "unbekannt":
            return label_field(row.get("age_source"))
        return ""
    if key == "ai_value_note":
        if row.get("latest_deep_dive_at"):
            return "KI-Wert ist ein konservativer Richtwert und muss bei Bedarf fachlich bestätigt werden."
        return ""
    if key == "open_task_roles_label":
        return ", ".join(label_role(part.strip()) for part in str(row.get("open_task_roles") or "").split(",") if part.strip())
    if key == "open_task_fields_label":
        return label_field_list(row.get("open_task_fields"))
    if key in {
        "created_at",
        "updated_at",
        "finalized_at",
        "session_created_at",
        "session_started_at",
        "latest_ai_at",
        "latest_deep_dive_at",
        "manual_reviewed_at",
        "first_photo_uploaded_at",
        "last_photo_uploaded_at",
        "captured_at",
    }:
        return format_excel_datetime(row.get(key))
    value = row.get(key)
    if isinstance(value, bool):
        return "ja" if value else "nein"
    return value


def filename_part(value: str | None, fallback: str) -> str:
    raw = (value or fallback).strip().lower()
    cleaned = "".join(ch if ch.isalnum() else "-" for ch in raw)
    while "--" in cleaned:
        cleaned = cleaned.replace("--", "-")
    return cleaned.strip("-") or fallback


EXPORT_COLUMNS = [
    ("lfd. Nr. / Foto", "lfd_photo"),
    ("Bezeichnung", "object_type"),
    ("Typ / Spezifikation", "specification"),
    ("Baujahr", "construction_year"),
    ("Zustand", "condition_label"),
    ("Funktion i. O. Ja", "function_ok_yes"),
    ("Funktion i. O. Nein", "function_ok_no"),
    ("UVV bis", "uvv_valid_until"),
    ("Bemerkung", "bga_remark"),
]


def export_query(where_sql: str = "", params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    return fetch_all(
        f"""
        SELECT i.*, oc.name AS object_class_name,
          s.status AS session_status, s.created_at AS session_created_at,
          s.started_at AS session_started_at, s.closed_at AS session_closed_at,
          l.name AS location_name, b.name AS building_name, r.name AS room_name,
          session_starter.display_name AS session_started_by_name,
          creator.display_name AS created_by_name,
          reviewer.display_name AS reviewed_by_name,
          finalizer.display_name AS finalized_by_name,
          (
            SELECT ar.created_at
            FROM ai_results ar
            WHERE ar.item_id = i.id
            ORDER BY ar.created_at DESC
            LIMIT 1
          ) AS latest_ai_at,
          (
            SELECT ar.model_used
            FROM ai_results ar
            WHERE ar.item_id = i.id
            ORDER BY ar.created_at DESC
            LIMIT 1
          ) AS latest_ai_model,
          (
            SELECT ar.created_at
            FROM ai_results ar
            WHERE ar.item_id = i.id AND ar.ai_type = 'deep_dive'
            ORDER BY ar.created_at DESC
            LIMIT 1
          ) AS latest_deep_dive_at,
          (
            SELECT al.created_at
            FROM audit_log al
            WHERE al.entity_type = 'inventory_item'
              AND al.entity_id = i.id
              AND al.action IN ('item_changed', 'status_changed', 'item_finalized')
            ORDER BY al.created_at DESC
            LIMIT 1
          ) AS manual_reviewed_at,
          EXISTS(SELECT 1 FROM item_photos p WHERE p.item_id = i.id AND p.photo_type IN ('object','object_front')) AS has_object_photo,
          EXISTS(SELECT 1 FROM item_photos p WHERE p.item_id = i.id AND p.photo_type IN ('nameplate','type_plate')) AS has_nameplate_photo,
          EXISTS(SELECT 1 FROM item_photos p WHERE p.item_id = i.id AND p.photo_type = 'dot') AS has_dot_photo,
          (
            SELECT p.id
            FROM item_photos p
            WHERE p.item_id = i.id AND p.photo_type IN ('object','object_front')
            ORDER BY p.uploaded_at DESC
            LIMIT 1
          ) AS object_photo_id,
          (
            SELECT COALESCE(p.original_path, p.stamped_path)
            FROM item_photos p
            WHERE p.item_id = i.id AND p.photo_type IN ('object','object_front')
            ORDER BY p.uploaded_at DESC
            LIMIT 1
          ) AS object_photo_path,
          (SELECT min(p.uploaded_at) FROM item_photos p WHERE p.item_id = i.id) AS first_photo_uploaded_at,
          (SELECT max(p.uploaded_at) FROM item_photos p WHERE p.item_id = i.id) AS last_photo_uploaded_at,
          (SELECT count(*)::int FROM item_photos p WHERE p.item_id = i.id) AS photo_count,
          (SELECT count(*)::int FROM item_audio_notes a WHERE a.item_id = i.id) AS audio_count,
          (SELECT count(*)::int FROM accounting_tasks t WHERE t.item_id = i.id AND t.status = 'open') AS open_task_count,
          (
            SELECT string_agg(DISTINCT t.assigned_role, ', ' ORDER BY t.assigned_role)
            FROM accounting_tasks t
            WHERE t.item_id = i.id AND t.status = 'open'
          ) AS open_task_roles,
          (
            SELECT string_agg(t.missing_field, ', ' ORDER BY t.created_at)
            FROM accounting_tasks t
            WHERE t.item_id = i.id AND t.status = 'open'
          ) AS open_task_fields
        FROM inventory_items i
        JOIN inventory_sessions s ON s.id = i.session_id
        JOIN locations l ON l.id = i.location_id
        JOIN buildings b ON b.id = i.building_id
        JOIN rooms r ON r.id = i.room_id
        LEFT JOIN object_classes oc ON oc.id = i.object_class_id
        LEFT JOIN users session_starter ON session_starter.id = s.started_by
        LEFT JOIN users creator ON creator.id = i.created_by
        LEFT JOIN users reviewer ON reviewer.id = i.reviewed_by
        LEFT JOIN users finalizer ON finalizer.id = i.finalized_by
        {where_sql}
        ORDER BY l.name, b.name, r.name, i.created_at DESC
        """,
        params,
    )


def refreshed_export_rows(where_sql: str = "", params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    rows = export_query(where_sql, params)
    refreshed = False
    for row in rows:
        if row.get("id") and row.get("inventory_type") == "bga":
            run_inventory_rework_check(str(row["id"]))
            refreshed = True
    rows = export_query(where_sql, params) if refreshed else rows
    return normalize_export_open_tasks(rows)


def bga_task_satisfied_by_item(task: dict[str, Any], item: dict[str, Any]) -> bool:
    if item.get("inventory_type") != "bga":
        return False
    field = label_field(task.get("missing_field")).lower()
    function_resolved = item.get("function_ok") == "ja"
    uvv_resolved = item.get("uvv_status") in {"vorhanden", "nicht_uvv_pflichtig"}
    inspection_book_resolved = item.get("inspection_book_available") in {"ja", "nein", "nicht_erforderlich"}
    if "funktion" in field and "uvv" in field:
        return function_resolved and uvv_resolved
    if "funktion" in field:
        return function_resolved
    if "uvv" in field:
        return uvv_resolved
    if "prüfbuch" in field or "pruefbuch" in field:
        return inspection_book_resolved
    return False


def export_task_relevant(task: dict[str, Any], item: dict[str, Any]) -> bool:
    if bga_task_satisfied_by_item(task, item):
        return False
    if item.get("inventory_type") == "bga" and "prüfbuch" in label_field(task.get("missing_field")).lower():
        return False
    return True


def normalize_export_open_tasks(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    for item in rows:
        if not item.get("id"):
            continue
        item_tasks = fetch_all(
            """
            SELECT *
            FROM accounting_tasks
            WHERE item_id = %s AND status = 'open'
            ORDER BY created_at
            """,
            (item["id"],),
        )
        relevant = [task for task in item_tasks if export_task_relevant(task, item)]
        item["open_task_count"] = len(relevant)
        roles = sorted({str(task.get("assigned_role") or "") for task in relevant if task.get("assigned_role")})
        item["open_task_roles"] = ", ".join(roles)
        item["open_task_fields"] = ", ".join(str(task.get("missing_field") or "") for task in relevant if task.get("missing_field"))
    return rows


def task_action(role: str | None, field: str | None) -> str:
    role_label = label_role(role)
    field_label = label_field(field)
    if role in {"Buchhaltung", "Auswertung"}:
        return f"{field_label} nach der Raumaufnahme auswerten"
    if role == "Erfasser":
        return f"{field_label} im Raum ergänzen"
    if role == "Technik":
        return f"{field_label} technisch prüfen"
    return f"{field_label} für {role_label} prüfen"


def task_reason(role: str | None) -> str:
    if role in {"Buchhaltung", "Auswertung"}:
        return "Blockiert die schnelle Erfassung nicht; dient der späteren Auswertung."
    if role == "Erfasser":
        return "Fehlt für belastbaren Foto-/Objektnachweis im Raum."
    if role == "Technik":
        return "Technischer Nachweis oder Prüffrist ist noch offen."
    return "Punkt ist für saubere Prüfung noch offen."


def fetch_export_tasks(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    tasks: list[dict[str, Any]] = []
    for item in rows:
        if not item.get("id"):
            continue
        item_tasks = fetch_all(
            """
            SELECT *
            FROM accounting_tasks
            WHERE item_id = %s AND status = 'open'
            ORDER BY created_at
            """,
            (item["id"],),
        )
        for task in item_tasks:
            if not export_task_relevant(task, item):
                continue
            if item.get("inventory_type") == "bga" and "prüfbuch" in label_field(task.get("missing_field")).lower():
                continue
            task["item"] = item
            tasks.append(task)
    return tasks


def fetch_export_photos(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    photos: list[dict[str, Any]] = []
    for item in rows:
        if not item.get("id"):
            continue
        item_photos = fetch_all(
            """
            SELECT id, photo_type, original_path, stamped_path, uploaded_at, taken_at
            FROM item_photos
            WHERE item_id = %s
            ORDER BY uploaded_at ASC
            """,
            (item["id"],),
        )
        for photo in item_photos:
            photo["item"] = item
            photos.append(photo)
    return photos


def fetch_export_audit_rows(rows: list[dict[str, Any]], session_id: str | None, entity_id: str | None) -> list[dict[str, Any]]:
    ids = [str(row["id"]) for row in rows if row.get("id")]
    if session_id:
        ids.append(session_id)
    elif entity_id:
        ids.append(entity_id)
    audit_rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_id in ids:
        if row_id in seen:
            continue
        seen.add(row_id)
        audit_rows.extend(fetch_all("SELECT * FROM audit_log WHERE entity_id = %s ORDER BY created_at ASC", (row_id,)))
    return sorted(audit_rows, key=lambda row: row.get("created_at") or datetime.min)


EXCEL_DETAIL_PHOTO_TYPES = {"type_plate", "nameplate", "uvv_label", "condition_detail", "dot", "tread", "tire_size", "damage"}


def excel_photo_limit(photo_type: Any) -> int:
    return 2400 if str(photo_type or "") in EXCEL_DETAIL_PHOTO_TYPES else 1600


def excel_photo_source(photo: dict[str, Any]) -> Path | None:
    raw_path = photo.get("original_path") or photo.get("stamped_path")
    if not raw_path:
        return None
    try:
        path = safe_upload_path(str(raw_path))
    except HTTPException:
        return None
    return path if path.exists() else None


def prepare_excel_photo(photo: dict[str, Any]) -> tuple[Path | None, int | None, int | None, str]:
    source = excel_photo_source(photo)
    if not source:
        return None, None, None, "Original fehlt"
    limit = excel_photo_limit(photo.get("photo_type"))
    digest = hashlib.sha256(f"{source}-{source.stat().st_mtime_ns}-{limit}".encode("utf-8")).hexdigest()[:16]
    target = Path(settings.upload_root, "temp", f"excel-{digest}.jpg")
    try:
        if target.exists():
            with PillowImage.open(target) as cached:
                return target, cached.width, cached.height, "Original für Excel optimiert"
        with PillowImage.open(source) as image:
            image = ImageOps.exif_transpose(image)
            image.thumbnail((limit, limit), PillowImage.Resampling.LANCZOS)
            if image.mode not in {"RGB", "L"}:
                image = image.convert("RGB")
            target.parent.mkdir(parents=True, exist_ok=True)
            image.save(target, format="JPEG", quality=90, optimize=True)
            return target, image.width, image.height, "Original für Excel optimiert"
    except Exception:
        return source, None, None, "Original unverändert eingebettet"


def audit_action_label(action: Any) -> str:
    labels = {
        "session_started": "Raum-Session gestartet",
        "device_joined": "Handy gekoppelt",
        "item_created": "Gegenstand angelegt",
        "photo_uploaded": "Foto hochgeladen",
        "audio_saved": "Sprachnotiz gespeichert",
        "ai_job_queued": "KI gestartet",
        "ai_result_created": "KI-Vorschlag erstellt",
        "ai_deep_dive_created": "KI Deep Dive erstellt",
        "item_changed": "Gegenstand manuell geändert",
        "status_changed": "Status geändert",
        "rework_requested": "Nacharbeit gesetzt",
        "item_finalized": "Gegenstand finalisiert",
        "session_closed": "Raum abgeschlossen",
        "export_created": "Excel-Export erstellt",
    }
    return labels.get(str(action or ""), str(action or ""))


def build_excel_workbook(
    title: str,
    rows: list[dict[str, Any]],
    summary: dict[str, Any],
    session_id: str | None = None,
    entity_id: str | None = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventurliste"
    header_fill = PatternFill("solid", fgColor="D9E2D6")
    thin_side = Side(style="thin", color="6B7280")
    list_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.merge_cells("A1:I1")
    ws["A1"] = "Manuelle Zählliste"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:I2")
    ws["A2"] = "Betriebs- und Geschäftsausstattung"
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = Alignment(horizontal="center")

    captured_by = next((export_value(row, "captured_by_name") for row in rows if row.get("id")), summary.get("erfasst durch") or summary.get("Aufnehmer") or "")
    captured_at = summary.get("Datum") or summary.get("Erstes Objekt aufgenommen am") or format_excel_datetime(datetime.now())
    ws["A4"] = "Standort"
    ws["B4"] = " / ".join(str(part) for part in [summary.get("Betrieb"), summary.get("Gebäude"), summary.get("Raum")] if part)
    ws["D4"] = "erfasst durch"
    ws["E4"] = captured_by
    ws["G4"] = "Datum"
    ws["H4"] = captured_at
    for cell_ref in ["A4", "D4", "G4"]:
        ws[cell_ref].font = Font(bold=True)

    header_row = 7
    for column_index, (header, _) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = list_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=header_row + 1):
        ws.append([excel_value(export_value(row, key)) for _, key in EXPORT_COLUMNS])
        insert_excel_photo(ws, row, row_index)
        for column_index, (_, key) in enumerate(EXPORT_COLUMNS, start=1):
            if key in {"value_provenance", "age_provenance", "ai_value_note", "latest_ai_at", "latest_deep_dive_at", "latest_ai_model"}:
                ws.cell(row=row_index, column=column_index).fill = PatternFill("solid", fgColor="D9EAFE")
            cell = ws.cell(row=row_index, column=column_index)
            cell.border = list_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = max(ws.row_dimensions[row_index].height or 0, 46)

    ws.freeze_panes = "A8"
    if rows:
        ws.auto_filter.ref = f"A{header_row}:I{header_row + len(rows)}"
    ws.print_title_rows = "1:7"
    ws.print_area = f"A1:I{max(header_row + len(rows), header_row + 1)}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    widths = [18, 28, 34, 12, 16, 16, 16, 16, 48]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    summary_ws = wb.create_sheet("Übersicht", 0)
    summary_ws.append(["Export", title])
    summary_ws.append(["Erzeugt am", excel_value(datetime.now())])
    for key, value in summary.items():
        summary_ws.append([key, excel_value(value)])
    summary_ws["A1"].font = Font(bold=True)
    summary_ws["B1"].font = Font(bold=True)
    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 42

    task_ws = wb.create_sheet("Offene Punkte - Nacharbeit")
    task_ws.append(["Inventar-ID", "Objektart", "Rolle", "Feld / Thema", "Status", "Betrieb", "Gebäude", "Raum"])
    for row in rows:
        if not row.get("open_task_count"):
            continue
        task_ws.append([
            row.get("inventory_id"),
            row.get("object_type"),
            row.get("open_task_roles"),
            row.get("open_task_fields"),
            row.get("review_status"),
            row.get("location_name"),
            row.get("building_name"),
            row.get("room_name"),
        ])
    task_ws.freeze_panes = "A2"
    task_ws.auto_filter.ref = task_ws.dimensions
    for index in range(1, 9):
        task_ws.column_dimensions[get_column_letter(index)].width = 24

    if "Offene Punkte - Nacharbeit" in wb.sheetnames:
        del wb["Offene Punkte - Nacharbeit"]
    task_ws = wb.create_sheet("Offene Punkte - Nacharbeit")
    task_ws.append([
        "Priorität", "Inventar-ID", "Objekt", "Klasse", "Raum", "Verantwortlich",
        "Aufgabe", "Warum relevant", "Blockiert Finalisierung", "Status", "Erstellt am", "Erledigt am", "Hinweis"
    ])
    for task in fetch_export_tasks(rows):
        item = task["item"]
        role = task.get("assigned_role")
        field = task.get("missing_field")
        blocks = "ja" if role in {"Erfasser", "Prüfer", "Technik"} else "nein"
        task_ws.append([
            task.get("priority") or "normal",
            item.get("inventory_id") or item.get("temporary_id"),
            item.get("object_type") or "Unbekanntes Objekt",
            item.get("object_class_name"),
            item.get("room_name"),
            label_role(role),
            task_action(role, field),
            task_reason(role),
            blocks,
            label_status(task.get("status")),
            format_excel_datetime(task.get("created_at")),
            format_excel_datetime(task.get("completed_at")),
            task.get("comment") or "",
        ])
    task_ws.freeze_panes = "A2"
    task_ws.auto_filter.ref = task_ws.dimensions
    for index in range(1, 14):
        task_ws.column_dimensions[get_column_letter(index)].width = 24

    photos_ws = wb.create_sheet("Fotos - Nachweise")
    photos_ws.append([
        "Bild", "Objekt-ID", "Laufende Nr.", "Objekt", "Fotoart", "Dateiname",
        "Original-Link/Pfad", "Excel-Bildquelle", "Pixel", "Aufgenommen am", "Hochgeladen am"
    ])
    for row_index, photo in enumerate(fetch_export_photos(rows), start=2):
        item = photo["item"]
        source_path = excel_photo_source(photo)
        image_path, width, height, source_note = prepare_excel_photo(photo)
        photos_ws.append([
            " ",
            item.get("inventory_id") or item.get("temporary_id"),
            item.get("sequence_number"),
            item.get("object_type"),
            label_field(photo.get("photo_type")),
            source_path.name if source_path else "",
            str(source_path) if source_path else "",
            source_note,
            f"{width} x {height}" if width and height else "",
            format_excel_datetime(photo.get("taken_at")),
            format_excel_datetime(photo.get("uploaded_at")),
        ])
        if source_path:
            link_cell = photos_ws.cell(row=row_index, column=7)
            link_cell.hyperlink = f"/uploads/photos/{photo['id']}"
            link_cell.style = "Hyperlink"
        if image_path and image_path.exists():
            try:
                image = ExcelImage(str(image_path))
                max_width, max_height = 300, 210
                scale = min(max_width / image.width, max_height / image.height, 1)
                image.width = int(image.width * scale)
                image.height = int(image.height * scale)
                image.anchor = f"A{row_index}"
                photos_ws.add_image(image)
                photos_ws.row_dimensions[row_index].height = 168
            except Exception:
                photos_ws.cell(row=row_index, column=1, value="Bild nicht lesbar")
    photos_ws.freeze_panes = "A2"
    photos_ws.auto_filter.ref = photos_ws.dimensions
    for index, width in enumerate([34, 24, 12, 28, 18, 36, 72, 28, 18, 22, 22], start=1):
        photos_ws.column_dimensions[get_column_letter(index)].width = width

    protocol_ws = wb.create_sheet("Protokoll")
    protocol_ws.append(["Zeitpunkt", "Aktion", "Entität", "Benutzer", "Details"])
    for audit_row in fetch_export_audit_rows(rows, session_id, entity_id):
        protocol_ws.append([
            format_excel_datetime(audit_row.get("created_at")),
            audit_action_label(audit_row.get("action")),
            audit_row.get("entity_type"),
            audit_row.get("user_id") or "",
            str(audit_row.get("new_value_json") or "")[:500],
        ])
    protocol_ws.freeze_panes = "A2"
    protocol_ws.auto_filter.ref = protocol_ws.dimensions
    for index, width in enumerate([22, 28, 22, 24, 80], start=1):
        protocol_ws.column_dimensions[get_column_letter(index)].width = width

    append_accounting_sheet(wb, rows)
    return wb


def insert_excel_photo(ws: Any, row: dict[str, Any], row_index: int) -> None:
    cell = ws.cell(row=row_index, column=1)
    current_value = cell.value
    photo_path = row.get("object_photo_path")
    if not photo_path:
        return
    path = Path(str(photo_path))
    if not path.exists():
        cell.value = f"{current_value or ''} / Foto fehlt".strip()
        return
    try:
        image_path, _, _, _ = prepare_excel_photo({"original_path": str(path), "photo_type": "object_front"})
        image = ExcelImage(str(image_path or path))
    except Exception:
        cell.value = f"{current_value or ''} / Foto nicht lesbar".strip()
        return

    max_side = 82
    scale = min(max_side / image.width, max_side / image.height, 1)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    image.anchor = f"A{row_index}"
    ws.add_image(image)
    ws.row_dimensions[row_index].height = 68
    cell.value = current_value
    if row.get("object_photo_id"):
        cell.hyperlink = f"/uploads/photos/{row['object_photo_id']}"
        cell.style = "Hyperlink"


def save_export(
    *,
    title: str,
    rows: list[dict[str, Any]],
    filename: str,
    session_id: str | None,
    entity_type: str,
    entity_id: str | None,
    export_type: str,
    summary: dict[str, Any],
) -> dict[str, Any]:
    ensure_upload_dirs()
    path = Path(settings.upload_root, "exports", filename)
    wb = build_excel_workbook(title, rows, summary, session_id=session_id, entity_id=entity_id)
    wb.save(path)
    export = execute(
        "INSERT INTO exports (tenant_id, session_id, export_type, file_path) VALUES ((SELECT tenant_id FROM inventory_sessions WHERE id = %s), %s, %s, %s) RETURNING *",
        (session_id, session_id, export_type, str(path)),
    )
    audit("export_created", entity_type, entity_id, {"export": export, "rows": len(rows), "title": title})
    return export


DAMAGE_PHOTO_LABELS = {
    "front": "Frontansicht",
    "side": "Rückansicht",
    "serial_number": "Typenschild (Seriennummer und BJ)",
    "uvv_sticker": "UVV-Aufkleber (wenn vorhanden)",
    "damage_detail_1": "Schaden 1",
    "damage_detail_2": "Schaden 2",
}

DAMAGE_PHOTO_TYPES = tuple(DAMAGE_PHOTO_LABELS.keys())


def request_tenant_id(request: Request) -> str | None:
    auth = getattr(request.state, "auth", {}) or {}
    tenant_id = auth.get("tenant_id") or default_tenant_id()
    return str(tenant_id) if tenant_id else None


def parse_client_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def damage_article_value(article: DamageArticleSnapshot, field: str) -> Any:
    data = article.model_dump()
    return data.get(field)


def damage_report_row(report_id: str) -> dict[str, Any]:
    row = fetch_one("SELECT * FROM damage_reports WHERE id = %s", (report_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Schadensfall nicht gefunden")
    return row


def delete_damage_files(paths: list[str]) -> int:
    upload_root = Path(settings.upload_root, "damage").resolve()
    removed = 0
    for raw_path in paths:
        try:
            path = Path(str(raw_path)).resolve()
            if upload_root not in path.parents or not path.exists() or not path.is_file():
                continue
            path.unlink()
            removed += 1
        except Exception:
            continue
    return removed


def find_damage_report_conflict(tenant_id: str | None, article_no: str, client_report_id: str, source_device_id: str | None) -> dict[str, Any] | None:
    existing_article = fetch_one(
        """
        SELECT *
        FROM damage_reports
        WHERE tenant_id IS NOT DISTINCT FROM %s AND article_no = %s
        LIMIT 1
        """,
        (tenant_id, article_no),
    )
    if not existing_article:
        return None
    if existing_article.get("client_report_id") == client_report_id:
        return None
    if source_device_id and existing_article.get("source_device_id") == source_device_id:
        return None
    return existing_article


def upsert_damage_report(payload: DamageReportPayload, tenant_id: str | None) -> dict[str, Any]:
    article_no = str(payload.article_no or "").strip()
    if not article_no:
        raise HTTPException(status_code=422, detail="Artikelnummer fehlt")
    existing = fetch_one(
        """
        SELECT *
        FROM damage_reports
        WHERE tenant_id IS NOT DISTINCT FROM %s AND client_report_id = %s
          AND (%s::text IS NULL OR source_device_id = %s)
        LIMIT 1
        """,
        (tenant_id, payload.client_report_id, payload.source_device_id, payload.source_device_id),
    ) or fetch_one(
        """
        SELECT *
        FROM damage_reports
        WHERE tenant_id IS NOT DISTINCT FROM %s AND article_no = %s
        LIMIT 1
        """,
        (tenant_id, article_no),
    )
    captured_at = parse_client_datetime(payload.created_at)
    article = payload.article
    entry_type = "free" if str(payload.entry_type or "").strip() == "free" else "catalog"
    free_reference = (payload.free_reference or "").strip() or None
    values = {
        "session_id": payload.session_id,
        "source_device_id": payload.source_device_id,
        "article_no": article_no,
        "nr": damage_article_value(article, "nr") or article_no,
        "buchungskreis": damage_article_value(article, "buchungskreis"),
        "anlagenbezeichnung": damage_article_value(article, "anlagenbezeichnung"),
        "aktivdatum": str(damage_article_value(article, "aktivdatum_iso") or damage_article_value(article, "aktivdatum") or ""),
        "alter": damage_article_value(article, "alter"),
        "entry_type": entry_type,
        "free_reference": free_reference,
        "team_name": (payload.team_name or "Team 1").strip() or "Team 1",
        "damage_description": payload.description.strip(),
        "uvv_sticker_present": payload.uvv_sticker_present if payload.uvv_sticker_present in {"ja", "nein", "unklar"} else "unklar",
        "captured_at": captured_at,
    }
    if existing:
        row = execute(
            """
            UPDATE damage_reports
            SET session_id = COALESCE(%s, session_id),
                source_device_id = COALESCE(%s, source_device_id),
                article_no = %s,
                nr = %s,
                buchungskreis = %s,
                anlagenbezeichnung = %s,
                aktivdatum = %s,
                alter = %s,
                entry_type = %s,
                free_reference = %s,
                team_name = %s,
                damage_description = %s,
                uvv_sticker_present = %s,
                captured_at = COALESCE(%s, captured_at),
                updated_at = now()
            WHERE id = %s
            RETURNING *
            """,
            (
                values["session_id"],
                values["source_device_id"],
                values["article_no"],
                values["nr"],
                values["buchungskreis"],
                values["anlagenbezeichnung"],
                values["aktivdatum"],
                values["alter"],
                values["entry_type"],
                values["free_reference"],
                values["team_name"],
                values["damage_description"],
                values["uvv_sticker_present"],
                values["captured_at"],
                existing["id"],
            ),
        )
        audit("damage_report_updated", "damage_report", str(row["id"]), row)
        return row
    row = execute(
        """
        INSERT INTO damage_reports (
          tenant_id, client_report_id, session_id, source_device_id, article_no, nr, buchungskreis,
          anlagenbezeichnung, aktivdatum, alter, entry_type, free_reference, team_name, damage_description,
          uvv_sticker_present, captured_at
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING *
        """,
        (
            tenant_id,
            payload.client_report_id,
            values["session_id"],
            values["source_device_id"],
            values["article_no"],
            values["nr"],
            values["buchungskreis"],
            values["anlagenbezeichnung"],
            values["aktivdatum"],
            values["alter"],
            values["entry_type"],
            values["free_reference"],
            values["team_name"],
            values["damage_description"],
            values["uvv_sticker_present"],
            values["captured_at"],
        ),
    )
    audit("damage_report_created", "damage_report", str(row["id"]), row)
    return row


async def save_damage_photo(
    *,
    report_id: str,
    tenant_id: str | None,
    session_id: str | None,
    file: UploadFile,
    photo_type: str,
    client_photo_id: str | None,
    source_device_id: str | None,
) -> tuple[dict[str, Any], bool]:
    if photo_type not in DAMAGE_PHOTO_TYPES:
        raise HTTPException(status_code=422, detail="Fotoart ist unbekannt")
    ensure_upload_dirs()
    damage_report_row(report_id)
    existing = fetch_one(
        """
        SELECT *
        FROM damage_photos
        WHERE damage_report_id = %s
          AND (
            photo_type = %s OR
            (%s::text IS NOT NULL AND source_device_id = %s AND client_photo_id = %s)
          )
        ORDER BY uploaded_at DESC
        LIMIT 1
        """,
        (report_id, photo_type, client_photo_id, source_device_id, client_photo_id),
    )
    data = await file.read()
    mime_type, suffix = validate_photo_upload(file, data)
    digest = hashlib.sha256(data).hexdigest()
    filename = f"{report_id}-{photo_type}-{digest[:12]}{suffix}"
    path = Path(settings.upload_root, "damage", filename)
    path.write_bytes(data)
    width = None
    height = None
    try:
        with PillowImage.open(path) as image:
            image = ImageOps.exif_transpose(image)
            width, height = image.width, image.height
    except Exception:
        pass
    metadata = json.dumps({"mime_type": mime_type, "size": len(data), "filename": file.filename})
    if existing:
        old_path = Path(str(existing.get("original_path") or ""))
        row = execute(
            """
            UPDATE damage_photos
            SET session_id = COALESCE(%s, session_id),
                client_photo_id = COALESCE(%s, client_photo_id),
                source_device_id = COALESCE(%s, source_device_id),
                original_path = %s,
                original_hash = %s,
                width = %s,
                height = %s,
                uploaded_at = now(),
                metadata_json = %s::jsonb
            WHERE id = %s
            RETURNING *
            """,
            (session_id, client_photo_id, source_device_id, str(path), digest, width, height, metadata, existing["id"]),
        )
        try:
            if old_path != path and old_path.exists() and old_path.is_file():
                old_path.unlink()
        except Exception:
            pass
        audit("damage_photo_updated", "damage_report", report_id, row)
        return row, True
    row = execute(
        """
        INSERT INTO damage_photos (
          tenant_id, damage_report_id, session_id, client_photo_id, source_device_id, photo_type,
          original_path, original_hash, width, height, metadata_json
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)
        RETURNING *
        """,
        (tenant_id, report_id, session_id, client_photo_id, source_device_id, photo_type, str(path), digest, width, height, metadata),
    )
    audit("damage_photo_uploaded", "damage_report", report_id, row)
    return row, False


@app.post("/damage-reports/sync")
async def sync_damage_report(request: Request, payload: str = Form(...), files: list[UploadFile] = File(default=[])) -> dict[str, Any]:
    try:
        parsed = DamageSyncPayload(**json.loads(payload))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Schadenspaket ist kein gültiges JSON") from exc
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Schadenspaket ist unvollständig: {exc}") from exc
    if not parsed.report.description.strip():
        raise HTTPException(status_code=422, detail="Schadensbeschreibung fehlt")
    auth = getattr(request.state, "auth", {}) or {}
    if auth.get("kind") == "mobile_session":
        parsed.report.session_id = str(auth.get("session_id") or "") or parsed.report.session_id
        parsed.report.source_device_id = str(auth.get("device_id") or "") or parsed.report.source_device_id
    tenant_id = request_tenant_id(request)
    required_photo_types = {"front", "damage_detail_1"}
    if parsed.report.uvv_sticker_present == "ja":
        required_photo_types.add("uvv_sticker")
    submitted_photo_types = {photo.photo_type for photo in parsed.photos}
    existing_report = fetch_one(
        """
        SELECT id
        FROM damage_reports
        WHERE tenant_id IS NOT DISTINCT FROM %s AND article_no = %s
        LIMIT 1
        """,
        (tenant_id, str(parsed.report.article_no or "").strip()),
    )
    existing_photo_types = {
        str(row["photo_type"])
        for row in fetch_all(
            "SELECT photo_type FROM damage_photos WHERE damage_report_id = %s",
            (existing_report["id"],),
        )
    } if existing_report else set()
    missing_photo_types = [
        DAMAGE_PHOTO_LABELS[photo_type]
        for photo_type in DAMAGE_PHOTO_TYPES
        if photo_type in required_photo_types
        and photo_type not in submitted_photo_types
        and photo_type not in existing_photo_types
    ]
    if missing_photo_types:
        raise HTTPException(status_code=422, detail=f"Pflichtfoto fehlt: {', '.join(missing_photo_types)}")
    report = upsert_damage_report(parsed.report, tenant_id)
    photo_results: list[dict[str, Any]] = []
    for index, meta in enumerate(parsed.photos):
        if index >= len(files):
            photo_results.append({
                "client_photo_id": meta.client_photo_id,
                "photo_type": meta.photo_type,
                "status": "failed",
                "error": "Datei fehlt im Sync-Paket",
            })
            continue
        try:
            row, already_exists = await save_damage_photo(
                report_id=str(report["id"]),
                tenant_id=tenant_id,
                session_id=parsed.report.session_id,
                file=files[index],
                photo_type=meta.photo_type,
                client_photo_id=meta.client_photo_id,
                source_device_id=parsed.report.source_device_id,
            )
            photo_results.append({
                "client_photo_id": meta.client_photo_id,
                "photo_type": meta.photo_type,
                "status": "already_exists" if already_exists else "synced",
                "server_photo_id": str(row["id"]),
            })
        except Exception as exc:
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            photo_results.append({
                "client_photo_id": meta.client_photo_id,
                "photo_type": meta.photo_type,
                "status": "failed",
                "error": detail,
            })
    return {
        "server_report_id": str(report["id"]),
        "client_report_id": parsed.report.client_report_id,
        "article_no": str(report["article_no"]),
        "status": "updated",
        "photo_results": photo_results,
    }


@app.get("/damage-reports")
def list_damage_reports(request: Request) -> list[dict[str, Any]]:
    tenant_id = request_tenant_id(request)
    return fetch_all(
        """
        SELECT r.*,
               COALESCE(p.photo_count, 0)::int AS photo_count,
               COALESCE(p.photo_types, ARRAY[]::text[]) AS photo_types,
               COALESCE(p.photos, '[]'::jsonb) AS photos
        FROM damage_reports r
        LEFT JOIN (
          SELECT damage_report_id,
                 count(*)::int AS photo_count,
                 array_agg(photo_type ORDER BY photo_type) AS photo_types,
                 jsonb_agg(
                   jsonb_build_object(
                     'id', id::text,
                     'photo_type', photo_type,
                     'uploaded_at', uploaded_at
                   )
                   ORDER BY uploaded_at
                 ) AS photos
          FROM damage_photos
          GROUP BY damage_report_id
        ) p ON p.damage_report_id = r.id
        WHERE r.tenant_id IS NOT DISTINCT FROM %s
        ORDER BY r.updated_at DESC
        """,
        (tenant_id,),
    )


@app.delete("/damage-reports/{report_id}")
def delete_damage_report(report_id: str, request: Request) -> dict[str, Any]:
    try:
        UUID(report_id)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="Schadensfall-ID ist ungültig") from exc
    tenant_id = request_tenant_id(request)
    report = fetch_one(
        "SELECT * FROM damage_reports WHERE id = %s AND tenant_id IS NOT DISTINCT FROM %s",
        (report_id, tenant_id),
    )
    if not report:
        raise HTTPException(status_code=404, detail="Schadensfall nicht gefunden")
    photo_paths = [
        str(row["original_path"])
        for row in fetch_all("SELECT original_path FROM damage_photos WHERE damage_report_id = %s", (report_id,))
        if row.get("original_path")
    ]
    photo_count = len(photo_paths)
    execute("DELETE FROM damage_reports WHERE id = %s RETURNING id", (report_id,))
    removed_files = delete_damage_files(photo_paths)
    audit("damage_report_deleted", "damage_report", report_id, {
        "article_no": report.get("article_no"),
        "photo_count": photo_count,
        "removed_files": removed_files,
    })
    return {"deleted": True, "report_id": report_id, "photo_count": photo_count, "removed_files": removed_files}


def fetch_damage_export_rows(tenant_id: str | None) -> list[dict[str, Any]]:
    rows = fetch_all(
        """
        SELECT *
        FROM damage_reports
        WHERE tenant_id IS NOT DISTINCT FROM %s
        ORDER BY CASE WHEN article_no ~ '^[0-9]+$' THEN article_no::int END NULLS LAST, article_no
        """,
        (tenant_id,),
    )
    for row in rows:
        row["photos"] = fetch_all(
            """
            SELECT *
            FROM damage_photos
            WHERE damage_report_id = %s
            ORDER BY uploaded_at
            """,
            (row["id"],),
        )
    return rows


def damage_photo_for_type(row: dict[str, Any], photo_type: str) -> dict[str, Any] | None:
    for photo in row.get("photos") or []:
        if photo.get("photo_type") == photo_type:
            return photo
    return None


def insert_damage_excel_photo(ws: Any, row: dict[str, Any], row_index: int, column_index: int, photo_type: str) -> None:
    photo = damage_photo_for_type(row, photo_type)
    cell = ws.cell(row=row_index, column=column_index)
    if not photo:
        cell.value = "fehlt"
        return
    image_path, _, _, _ = prepare_excel_photo({"original_path": photo.get("original_path"), "photo_type": photo_type})
    if not image_path or not image_path.exists():
        cell.value = "Bild fehlt"
        return
    try:
        image = ExcelImage(str(image_path))
        max_width, max_height = 142, 104
        scale = min(max_width / image.width, max_height / image.height, 1)
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)
        image.anchor = f"{get_column_letter(column_index)}{row_index}"
        ws.add_image(image)
        cell.value = " "
        cell.hyperlink = f"/uploads/damage-photos/{photo['id']}"
        cell.style = "Hyperlink"
    except Exception:
        cell.value = "Bild nicht lesbar"


def build_damage_excel_workbook(rows: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Schadensliste"
    header_fill = PatternFill("solid", fgColor="0D1A2E")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="C6D6EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = [
        "Art",
        "Nr.",
        "Buchungskreis",
        "Anlagenbezeichnung",
        "Freier Hinweis",
        "Aktivdatum",
        "Alter",
        "Team",
        "Erfasst am",
        "Ge\u00e4ndert am",
        "Schadensbeschreibung",
        "UVV-Aufkleber vorhanden",
        "Frontansicht",
        "Rückansicht",
        "Typenschild (Seriennummer und BJ)",
        "UVV-Aufkleber (wenn vorhanden)",
        "Schaden 1",
        "Schaden 2",
    ]
    ws.append(headers)
    for column_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        entry_type = "Nicht in Liste" if row.get("entry_type") == "free" else "Listenartikel"
        ws.cell(row=row_index, column=1, value=entry_type)
        ws.cell(row=row_index, column=2, value=row.get("nr") or row.get("article_no"))
        ws.cell(row=row_index, column=3, value=row.get("buchungskreis"))
        ws.cell(row=row_index, column=4, value=row.get("anlagenbezeichnung"))
        ws.cell(row=row_index, column=5, value=row.get("free_reference"))
        ws.cell(row=row_index, column=6, value=row.get("aktivdatum"))
        ws.cell(row=row_index, column=7, value=float(row["alter"]) if row.get("alter") is not None else None)
        ws.cell(row=row_index, column=8, value=row.get("team_name"))
        ws.cell(row=row_index, column=9, value=format_excel_datetime(row.get("captured_at") or row.get("created_at")))
        ws.cell(row=row_index, column=10, value=format_excel_datetime(row.get("updated_at")))
        ws.cell(row=row_index, column=11, value=row.get("damage_description"))
        ws.cell(row=row_index, column=12, value=row.get("uvv_sticker_present"))
        for photo_offset, photo_type in enumerate(DAMAGE_PHOTO_TYPES, start=13):
            insert_damage_excel_photo(ws, row, row_index, photo_offset, photo_type)
        for column_index in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 92
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:R{len(rows) + 1}"
    widths = [16, 12, 16, 36, 30, 14, 10, 18, 20, 20, 48, 18, 22, 22, 24, 24, 24, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    summary_ws = wb.create_sheet("\u00dcbersicht", 0)
    summary_ws.append(["Export", "Schadensliste"])
    summary_ws.append(["Erzeugt am", excel_value(datetime.now())])
    summary_ws.append(["Nicht in Liste", len([row for row in rows if row.get("entry_type") == "free"])])
    summary_ws.append(["Schadensf\u00e4lle", len(rows)])
    summary_ws.append(["Teams", len({row.get("team_name") for row in rows if row.get("team_name")} )])
    summary_ws.column_dimensions["A"].width = 22
    summary_ws.column_dimensions["B"].width = 36
    summary_ws["A1"].font = Font(bold=True)
    summary_ws["B1"].font = Font(bold=True)
    return wb


@app.post("/damage-reports/export/excel")
def export_damage_excel(request: Request) -> dict[str, Any]:
    ensure_upload_dirs()
    tenant_id = request_tenant_id(request)
    rows = fetch_damage_export_rows(tenant_id)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"schadensliste-{timestamp}.xlsx"
    path = Path(settings.upload_root, "exports", filename)
    wb = build_damage_excel_workbook(rows)
    wb.save(path)
    export = execute(
        "INSERT INTO exports (tenant_id, session_id, export_type, file_path) VALUES (%s, NULL, %s, %s) RETURNING *",
        (tenant_id, "damage_excel", str(path)),
    )
    audit("damage_export_created", "damage_export", str(export["id"]), {"rows": len(rows), "file": str(path)})
    return export


@app.get("/items/{item_id}/ai-results")
def ai_results(item_id: str) -> list[dict[str, Any]]:
    return fetch_all("SELECT * FROM ai_results WHERE item_id = %s ORDER BY created_at DESC", (item_id,))


@app.get("/object-classes")
def object_classes() -> list[dict[str, Any]]:
    return fetch_all("SELECT * FROM object_classes ORDER BY name")


@app.get("/object-classes/{class_id}/requirements")
def requirements(class_id: str) -> list[dict[str, Any]]:
    return fetch_all("SELECT * FROM field_requirements WHERE object_class_id = %s ORDER BY sort_order", (class_id,))


@app.post("/object-classes")
def create_object_class(body: dict[str, Any]) -> dict[str, Any]:
    row = execute(
        """
        INSERT INTO object_classes (name, slug, description, default_commercial_category, requires_accounting_review)
        VALUES (%s, %s, %s, %s, %s) RETURNING *
        """,
        (
            body["name"],
            body["slug"],
            body.get("description"),
            body.get("default_commercial_category", "ungeklaert"),
            body.get("requires_accounting_review", False),
        ),
    )
    audit("object_class_created", "object_class", str(row["id"]), row)
    return row


@app.patch("/object-classes/{class_id}")
def update_object_class(class_id: str, body: dict[str, Any]) -> dict[str, Any]:
    row = execute(
        """
        UPDATE object_classes
        SET description = COALESCE(%s, description),
            default_commercial_category = COALESCE(%s, default_commercial_category),
            updated_at = now()
        WHERE id = %s RETURNING *
        """,
        (body.get("description"), body.get("default_commercial_category"), class_id),
    )
    return row


@app.get("/accounting/tasks")
def accounting_tasks() -> list[dict[str, Any]]:
    return fetch_all("SELECT * FROM accounting_tasks ORDER BY created_at DESC")


@app.patch("/accounting/tasks/{task_id}")
def patch_accounting_task(task_id: str, body: dict[str, Any]) -> dict[str, Any]:
    row = execute(
        """
        UPDATE accounting_tasks
        SET status = COALESCE(%s, status), comment = COALESCE(%s, comment),
            completed_at = CASE WHEN %s = 'completed' THEN now() ELSE completed_at END
        WHERE id = %s RETURNING *
        """,
        (body.get("status"), body.get("comment"), body.get("status"), task_id),
    )
    audit("accounting_task_changed", "accounting_task", task_id, row)
    return row


@app.get("/items/{item_id}/accounting")
def item_accounting(item_id: str) -> dict[str, Any]:
    return fetch_one("SELECT * FROM item_accounting_data WHERE item_id = %s", (item_id,))


@app.patch("/items/{item_id}/accounting")
def patch_item_accounting(item_id: str, body: dict[str, Any]) -> dict[str, Any]:
    require_item_session_open(item_id)
    row = execute(
        """
        UPDATE item_accounting_data
        SET commercial_category = COALESCE(%s, commercial_category),
            cost_center = COALESCE(%s, cost_center),
            asset_number = COALESCE(%s, asset_number),
            book_value = COALESCE(%s, book_value),
            accounting_status = COALESCE(%s, accounting_status)
        WHERE item_id = %s RETURNING *
        """,
        (
            body.get("commercial_category"),
            body.get("cost_center"),
            body.get("asset_number"),
            body.get("book_value"),
            body.get("accounting_status"),
            item_id,
        ),
    )
    audit("accounting_changed", "inventory_item", item_id, row)
    return row


@app.post("/sessions/{session_id}/export/excel")
def export_excel(session_id: str) -> dict[str, Any]:
    session = get_session(session_id)
    session_inventory_type = session.get("inventory_type") or "bga"
    if session_inventory_type != "bga":
        raise HTTPException(
            status_code=409,
            detail=f"Excel-Export für {inventory_type_label(session_inventory_type)} ist vorbereitet, aber noch nicht aktiv.",
        )
    rows = refreshed_export_rows("WHERE i.session_id = %s", (session_id,))
    return save_export(
        title=f"Raumaufnahme {session.get('room_name') or session_id}",
        rows=rows,
        filename=f"raumaufnahme-{filename_part(session.get('room_name'), 'raum')}-{session_id[:8]}.xlsx",
        session_id=session_id,
        entity_type="inventory_session",
        entity_id=session_id,
        export_type="session_excel",
        summary={
            "Betrieb": session.get("location_name"),
            "Gebäude": session.get("building_name"),
            "Raum": session.get("room_name"),
            "Aufnehmer": rows[0].get("session_started_by_name") if rows else None,
            "Session gestartet am": format_excel_datetime(session.get("started_at") or session.get("created_at")),
            "Erstes Objekt aufgenommen am": format_excel_datetime(min((row.get("created_at") for row in rows if row.get("created_at")), default=None)),
            "Letzte Änderung am": format_excel_datetime(max((row.get("updated_at") for row in rows if row.get("updated_at")), default=None)),
            "Prüfer": "siehe Inventur-Spalten Prüfer und Manuell geprüft am",
            "Objekte": len(rows),
            "Offene Nacharbeiten": sum(int(row.get("open_task_count") or 0) for row in rows),
        },
    )
@app.post("/items/{item_id}/export/excel")
def export_item_excel(item_id: str) -> dict[str, Any]:
    item = get_item(item_id)
    rows = refreshed_export_rows("WHERE i.id = %s", (item_id,))
    if not rows:
        raise HTTPException(status_code=404, detail="Gegenstand nicht gefunden")
    row = rows[0]
    label = row.get("inventory_id") or row.get("temporary_id") or item_id
    return save_export(
        title=f"Aufnahme {label}",
        rows=rows,
        filename=f"aufnahme-{filename_part(str(label), 'gegenstand')}.xlsx",
        session_id=str(item["session_id"]),
        entity_type="inventory_item",
        entity_id=item_id,
        export_type="item_excel",
        summary={
            "Inventar-ID": row.get("inventory_id"),
            "Objektart": row.get("object_type"),
            "Betrieb": row.get("location_name"),
            "Raum": row.get("room_name"),
            "Offene Nacharbeiten": row.get("open_task_count"),
        },
    )


@app.post("/exports/excel")
def export_all_excel() -> dict[str, Any]:
    rows = refreshed_export_rows()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    locations = len({row.get("location_name") for row in rows if row.get("location_name")})
    rooms = len({(row.get("location_name"), row.get("building_name"), row.get("room_name")) for row in rows})
    return save_export(
        title="Gesamtaufstellung Inventur",
        rows=rows,
        filename=f"gesamtaufstellung-inventur-{timestamp}.xlsx",
        session_id=None,
        entity_type="inventory_export",
        entity_id=None,
        export_type="all_excel",
        summary={
            "Betriebe": locations,
            "Räume": rooms,
            "Objekte": len(rows),
            "Offene Nacharbeiten": sum(int(row.get("open_task_count") or 0) for row in rows),
            "Finalisiert": sum(1 for row in rows if row.get("review_status") == "finalisiert" or row.get("status") == "finalisiert"),
        },
    )


@app.get("/exports/{export_id}/download")
def download_export(export_id: str, request: Request) -> FileResponse:
    row = fetch_one("SELECT * FROM exports WHERE id = %s", (export_id,))
    if not row or not os.path.exists(row["file_path"]):
        raise HTTPException(status_code=404, detail="Export not found")
    auth = getattr(request.state, "auth", {}) or {}
    if auth.get("tenant_id") and row.get("tenant_id") and str(row.get("tenant_id")) != str(auth.get("tenant_id")):
        raise HTTPException(status_code=403, detail="Export gehört nicht zu diesem Mandanten")
    if (
        auth.get("kind") == "mobile_session"
        and row.get("export_type") != "damage_excel"
        and str(row.get("session_id") or "") != str(auth.get("session_id") or "")
    ):
        raise HTTPException(status_code=403, detail="Export gehört nicht zu dieser mobilen Session")
    return FileResponse(row["file_path"], filename=Path(row["file_path"]).name)


@app.get("/uploads/damage-photos/{photo_id}")
def download_damage_photo(photo_id: str, request: Request) -> FileResponse:
    row = fetch_one("SELECT * FROM damage_photos WHERE id = %s", (photo_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Schadensfoto nicht gefunden")
    auth = getattr(request.state, "auth", {}) or {}
    if auth.get("tenant_id") and row.get("tenant_id") and str(row.get("tenant_id")) != str(auth.get("tenant_id")):
        raise HTTPException(status_code=403, detail="Foto gehört nicht zu diesem Mandanten")
    path = safe_upload_path(row["original_path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="Fotodatei nicht gefunden")
    return FileResponse(path)


@app.get("/uploads/photos/{photo_id}")
def download_photo(photo_id: str, request: Request) -> FileResponse:
    row = fetch_one(
        """
        SELECT p.*, i.session_id
        FROM item_photos p
        JOIN inventory_items i ON i.id = p.item_id
        WHERE p.id = %s
        """,
        (photo_id,),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Foto nicht gefunden")
    auth = getattr(request.state, "auth", {}) or {}
    if auth.get("kind") == "mobile_session" and str(row.get("session_id")) != str(auth.get("session_id")):
        raise HTTPException(status_code=403, detail="Foto gehört nicht zu dieser mobilen Session")
    if auth.get("tenant_id") and row.get("tenant_id") and str(row.get("tenant_id")) != str(auth.get("tenant_id")):
        raise HTTPException(status_code=403, detail="Foto gehört nicht zu diesem Mandanten")
    path = safe_upload_path(row.get("stamped_path") or row["original_path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="Fotodatei nicht gefunden")
    return FileResponse(path)


@app.get("/items/{item_id}/audit-log")
def item_audit(item_id: str) -> list[dict[str, Any]]:
    return fetch_all("SELECT * FROM audit_log WHERE entity_id = %s ORDER BY created_at DESC", (item_id,))


@app.get("/sessions/{session_id}/audit-log")
def session_audit(session_id: str) -> list[dict[str, Any]]:
    return fetch_all("SELECT * FROM audit_log WHERE entity_id = %s OR entity_type = 'inventory_item' ORDER BY created_at DESC LIMIT 200", (session_id,))


@app.get("/sessions/{session_id}/events")
async def session_events(session_id: str):
    async def stream():
        import asyncio
        import json

        last_payload = ""
        while True:
            payload = json.dumps({"items": session_items(session_id)}, default=str)
            if payload != last_payload:
                yield f"data: {payload}\n\n"
                last_payload = payload
            await asyncio.sleep(2)

    return StreamingResponse(stream(), media_type="text/event-stream")


@app.get("/uploads/{file_id}")
def upload_placeholder(file_id: str) -> dict[str, str]:
    return {"message": "Phase 1 stores file paths internally. Signed file lookup is reserved for hardening.", "file_id": file_id}


# ---------------------------------------------------------------------------
# Inventur-Cockpit: Live-Steuerstand ueber alle Raeume (Roadmap Sprint 1).
# Bewusst als reiner Lese-Endpoint am Dateiende ergaenzt.
# ---------------------------------------------------------------------------

@app.get("/cockpit/overview")
def cockpit_overview() -> dict[str, Any]:
    sessions = fetch_all(
        """
        SELECT s.id, s.status, s.started_at, s.closed_at, s.inventory_type,
               r.name AS room_name, b.name AS building_name, l.name AS location_name
        FROM inventory_sessions s
        JOIN rooms r ON r.id = s.room_id
        JOIN buildings b ON b.id = s.building_id
        JOIN locations l ON l.id = s.location_id
        WHERE s.status = 'open' OR s.closed_at::date = CURRENT_DATE
        ORDER BY (s.status = 'open') DESC, s.started_at DESC
        LIMIT 24
        """
    )
    session_ids = [str(row["id"]) for row in sessions]
    stats_rows = fetch_all(
        """
        SELECT session_id,
               count(*)::int AS items,
               count(*) FILTER (WHERE review_status = 'finalisiert')::int AS finalized,
               count(*) FILTER (WHERE review_status LIKE 'nacharbeit%%')::int AS rework,
               count(*) FILTER (WHERE captured_at > now() - interval '60 minutes')::int AS last_hour,
               max(captured_at) AS last_capture_at,
               COALESCE(sum(value_estimate), 0)::float AS value_sum
        FROM inventory_items
        WHERE session_id = ANY(%s)
        GROUP BY session_id
        """,
        (session_ids,),
    ) if session_ids else []
    photo_rows = fetch_all(
        """
        SELECT i.session_id, count(DISTINCT i.id)::int AS with_photo
        FROM inventory_items i
        JOIN item_photos p ON p.item_id = i.id
        WHERE i.session_id = ANY(%s) AND p.photo_type IN ('object', 'object_front')
        GROUP BY i.session_id
        """,
        (session_ids,),
    ) if session_ids else []
    device_rows = fetch_all(
        """
        SELECT session_id, device_name, last_seen_at, pending_count
        FROM session_devices
        WHERE session_id = ANY(%s) AND revoked_at IS NULL
        ORDER BY last_seen_at DESC NULLS LAST
        """,
        (session_ids,),
    ) if session_ids else []

    stats = {str(row["session_id"]): row for row in stats_rows}
    photos = {str(row["session_id"]): row["with_photo"] for row in photo_rows}
    devices: dict[str, list[dict[str, Any]]] = {}
    for row in device_rows:
        devices.setdefault(str(row["session_id"]), []).append(
            {
                "device_name": row["device_name"],
                "last_seen_at": row["last_seen_at"],
                "pending_count": row["pending_count"],
            }
        )

    from datetime import datetime, timezone

    now = datetime.now(timezone.utc)
    rooms: list[dict[str, Any]] = []
    for session in sessions:
        sid = str(session["id"])
        stat = stats.get(sid, {})
        items = int(stat.get("items") or 0)
        started = session.get("started_at")
        hours = max(((now - started).total_seconds() / 3600.0) if started else 0.25, 0.25)
        rooms.append(
            {
                "session_id": sid,
                "status": session["status"],
                "room_name": session["room_name"],
                "building_name": session["building_name"],
                "location_name": session["location_name"],
                "inventory_type": session.get("inventory_type"),
                "started_at": session.get("started_at"),
                "items": items,
                "finalized": int(stat.get("finalized") or 0),
                "rework": int(stat.get("rework") or 0),
                "last_hour": int(stat.get("last_hour") or 0),
                "last_capture_at": stat.get("last_capture_at"),
                "value_sum": float(stat.get("value_sum") or 0.0),
                "with_photo": int(photos.get(sid, 0)),
                "per_hour": round(items / hours, 1),
                "devices": devices.get(sid, []),
            }
        )

    totals_row = fetch_one(
        """
        SELECT count(*) FILTER (WHERE captured_at::date = CURRENT_DATE)::int AS today,
               count(*) FILTER (WHERE captured_at > now() - interval '60 minutes')::int AS last_hour,
               COALESCE(sum(value_estimate) FILTER (WHERE captured_at::date = CURRENT_DATE), 0)::float AS value_today
        FROM inventory_items
        """
    ) or {}
    feed = fetch_all(
        """
        SELECT i.captured_at, i.object_type, i.sequence_number, r.name AS room_name
        FROM inventory_items i
        JOIN inventory_sessions s ON s.id = i.session_id
        JOIN rooms r ON r.id = s.room_id
        ORDER BY i.captured_at DESC
        LIMIT 12
        """
    )
    return {
        "totals": {
            "today": int(totals_row.get("today") or 0),
            "last_hour": int(totals_row.get("last_hour") or 0),
            "value_today": float(totals_row.get("value_today") or 0.0),
            "open_rooms": sum(1 for room in rooms if room["status"] == "open"),
            "devices_online": sum(
                1
                for room_devices in devices.values()
                for device in room_devices
                if device["last_seen_at"] is not None
            ),
        },
        "rooms": rooms,
        "feed": feed,
        "generated_at": now,
    }


class CockpitHeartbeatIn(BaseModel):
    pending_count: int = 0


@app.post("/sessions/{session_id}/devices/{device_id}/heartbeat")
def device_heartbeat(session_id: str, device_id: str, body: CockpitHeartbeatIn) -> dict[str, Any]:
    """Handy meldet sich periodisch: speist das Geräte-Panel im Cockpit."""
    row = execute(
        "UPDATE session_devices SET last_seen_at = now(), pending_count = %s WHERE id = %s AND session_id = %s RETURNING id, last_seen_at, pending_count",
        (max(0, int(body.pending_count)), device_id, session_id),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Gerät nicht gefunden")
    return row


# ---------------------------------------------------------------------------
# Buchhaltungs-Blatt (SKR51) im Excel-Export + Pflege-Endpoints.
# Empfehlungen basieren auf konfigurierbaren Grenzen (settings.gwg_*) und
# ersetzen keine steuerliche Beratung.
# ---------------------------------------------------------------------------

_DEFAULT_USEFUL_LIFE = {
    "hebebuehne": 8, "monitor": 3, "notebook": 3, "eingabegeraet": 3,
    "it_geraet": 3, "werkzeugwagen": 13,
}


def append_accounting_sheet(wb: "Workbook", rows: list[dict[str, Any]]) -> None:
    profiles = {
        str(row["object_class_id"]): row
        for row in fetch_all(
            """
            SELECT ap.*, oc.slug AS object_class_slug
            FROM accounting_profiles ap
            JOIN object_classes oc ON oc.id = ap.object_class_id
            """
        )
    }
    ws = wb.create_sheet("Buchhaltung (SKR51)")
    minor = settings.gwg_minor_limit_eur
    gwg = settings.gwg_limit_eur
    pool = settings.gwg_pool_limit_eur

    ws.merge_cells("A1:K1")
    ws["A1"] = "Buchhaltung – SKR51-Vorbereitung"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:K2")
    ws["A2"] = (
        f"Einordnung automatisch aus Zeitwert (Grenzen: {minor:.0f} / {gwg:.0f} / {pool:.0f} EUR, "
        "in den Servereinstellungen anpassbar). Konto & Nutzungsdauer je Objektklasse pflegt die "
        "Buchhaltung (accounting_profiles). Keine steuerliche Beratung."
    )
    ws["A2"].font = Font(italic=True, size=9, color="555555")

    headers = [
        "Inventar-Nr", "Bezeichnung", "Typ / Spezifikation", "Objektklasse",
        "SKR51-Konto", "Kostenstelle", "Baujahr", "ND (Jahre)", "Zeitwert EUR",
        "Einordnung (Empfehlung)", "Belegfotos",
    ]
    header_row = 4
    header_fill = PatternFill("solid", fgColor="0D1A2E")
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    first_data = header_row + 1
    current_year = datetime.now().year
    for offset, row in enumerate(rows):
        row_index = first_data + offset
        profile = profiles.get(str(row.get("object_class_id") or ""), {})
        account = profile.get("default_skr51_account") or ""
        life = profile.get("useful_life_years")
        if life is None:
            life = _DEFAULT_USEFUL_LIFE.get(str(profile.get("object_class_slug") or ""), 13)
        photo_count = int(row.get("photo_count") or 0) or (1 if row.get("object_photo_id") else 0)
        value = row.get("value_estimate")
        ws.cell(row=row_index, column=1, value=excel_value(row.get("inventory_id") or row.get("temporary_id")))
        ws.cell(row=row_index, column=2, value=excel_value(row.get("object_type")))
        ws.cell(row=row_index, column=3, value=excel_value(row.get("specification")))
        ws.cell(row=row_index, column=4, value=excel_value(row.get("object_class_name")))
        account_cell = ws.cell(row=row_index, column=5, value=account or "KONTO PFLEGEN")
        if not account:
            account_cell.font = Font(color="B45309", bold=True)
        ws.cell(row=row_index, column=6, value=excel_value(row.get("cost_center")))
        ws.cell(row=row_index, column=7, value=excel_value(row.get("construction_year")))
        ws.cell(row=row_index, column=8, value=float(life) if life is not None else None)
        value_cell = ws.cell(row=row_index, column=9, value=float(value) if value is not None else None)
        value_cell.number_format = '#,##0.00 "EUR"'
        # Lebende Formel: Buchhaltung kann den Zeitwert aendern, die
        # Einordnung rechnet in Excel sofort mit.
        ws.cell(
            row=row_index,
            column=10,
            value=(
                f'=IF(I{row_index}="","Wert fehlt – prüfen",'
                f'IF(I{row_index}<={minor:.0f},"Sofortaufwand (<= {minor:.0f})",'
                f'IF(I{row_index}<={gwg:.0f},"GWG Sofortabschreibung",'
                f'IF(I{row_index}<={pool:.0f},"Sammelposten möglich","Aktivieren + AfA"))))'
            ),
        )
        ws.cell(row=row_index, column=11, value=photo_count)

    last_data = first_data + max(len(rows) - 1, 0)
    summary_row = last_data + 2
    ws.cell(row=summary_row, column=8, value="Summe Zeitwert:").font = Font(bold=True)
    total_cell = ws.cell(row=summary_row, column=9, value=f"=SUM(I{first_data}:I{last_data})" if rows else 0)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00 "EUR"'
    ws.cell(row=summary_row + 1, column=8, value="Aktivieren (> Pool):").font = Font(bold=True)
    ws.cell(
        row=summary_row + 1,
        column=9,
        value=f'=SUMIF(I{first_data}:I{last_data},">{pool:.0f}")' if rows else 0,
    ).number_format = '#,##0.00 "EUR"'

    ws.freeze_panes = f"A{first_data}"
    if rows:
        ws.auto_filter.ref = f"A{header_row}:K{last_data}"
    for index, width in enumerate([20, 26, 30, 18, 14, 12, 10, 10, 14, 26, 10], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


@app.get("/accounting/profiles")
def list_accounting_profiles() -> list[dict[str, Any]]:
    return fetch_all(
        """
        SELECT ap.*, oc.name AS object_class_name, oc.slug AS object_class_slug
        FROM accounting_profiles ap
        JOIN object_classes oc ON oc.id = ap.object_class_id
        ORDER BY oc.name
        """
    )


class AccountingProfilePatch(BaseModel):
    default_skr51_account: str | None = None
    default_cost_center: str | None = None
    useful_life_years: float | None = None


@app.patch("/accounting/profiles/{profile_id}")
def patch_accounting_profile(profile_id: str, body: AccountingProfilePatch) -> dict[str, Any]:
    row = execute(
        """
        UPDATE accounting_profiles
        SET default_skr51_account = COALESCE(%s, default_skr51_account),
            default_cost_center = COALESCE(%s, default_cost_center),
            useful_life_years = COALESCE(%s, useful_life_years),
            updated_at = now()
        WHERE id = %s RETURNING *
        """,
        (body.default_skr51_account, body.default_cost_center, body.useful_life_years, profile_id),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Profil nicht gefunden")
    audit("accounting_profile_changed", "accounting_profile", profile_id, row)
    return row
